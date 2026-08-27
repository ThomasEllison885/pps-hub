"""The Monday pack's AR totals and P&L snapshot, in more detail.

Run: python -m pytest tests/test_office_ops_report_detail.py -v

Thomas, 2026-08-27: "Make the AR totals more inclusive of the details on the
A/R Summary. Make the P&L snapshot more inclusive of the details too. Give
insight into unusual increases or decreases."

Two halves.

**AR.** The parser already produced per-customer buckets, an operating total
excluding Bridges, and an overdue-weighted chase list. The pack printed the
five aging buckets and stopped. Everything else was computed and thrown away.

**P&L.** The parser read five numbers and discarded the line items, so "what
moved" could not be asked at all. It now keeps them — except compensation,
which is dropped at parse time and never stored. That exclusion is the part
worth guarding: the pack is written to Postgres, rendered into an Excel and
emailed, and CLAUDE.md's boundary says comp does not live in Hub content.
`test_no_compensation_line_survives_into_the_workbook` is that guard, and it
looks at the finished .xlsx bytes rather than at any intermediate.

"Unusual" is two thresholds, both required (Thomas's choice on 2026-08-27
over dollars-only and percent-only). Dollars alone lets a $400k line drifting
2% outrank a $12k line that doubled; percent alone fills the section with $40
accounts tripling.
"""
import io
import os
import sys

import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

import office_ops
import office_ops_generate as gen

# A QB-shaped export: section labels with no figures, leaf accounts, roll-ups,
# and three compensation lines that must not come out the other side.
PL_ROWS = [
    ['Pure Property Solutions LLC', None, None, None],
    ['Profit and Loss', None, None, None],
    ['January 1 - August 26, 2026 vs January 1 - August 26, 2025', None, None, None],
    [None, 'Jan 1 - Aug 26, 2026', 'Jan 1 - Aug 26, 2025', '% Change'],
    ['Income', None, None, None],
    ['   Services', 4210000, 3480000, 20.98],
    ['   Materials Reimbursed', 118000, 96000, 22.92],
    ['   Discounts given', -21000, -8500, -147.06],
    ['Total for Income', 4307000, 3567500, 20.73],
    ['Cost of Goods Sold', None, None, None],
    ['   Job Materials', 1480000, 1210000, 22.31],
    ['   Subcontractor Labor', 980000, 905000, 8.29],
    ['   Dumpsters', 61000, 33000, 84.85],
    ['   Equipment Rental', 24500, 26100, -6.13],
    ['   Job Payroll', 402000, 351000, 14.53],
    ['Total for Cost of Goods Sold', 2947500, 2525100, 16.73],
    ['Gross Profit', 1359500, 1042400, 30.42],
    ['Expenses', None, None, None],
    ['   Advertising - Digital', 88000, 12000, 633.33],
    ['   Advertising - Radio', 0, 47000, -100.00],
    ['   Insurance - General Liability', 71000, 64500, 10.08],
    ['   Fuel', 44200, 42800, 3.27],
    ['   Office Supplies', 9100, 8400, 8.33],
    ['   Software Subscriptions', 31000, 9800, 216.33],
    ['   Officer Compensation', 240000, 210000, 14.29],
    ['   Payroll Expenses', 318000, 291000, 9.28],
    ['   Meals', 6200, 3100, 100.00],
    ['   Bank Charges', 1450, 1380, 5.07],
    ['   Vehicle Repairs', 38500, 14200, 171.13],
    ['Total for Expenses', 847450, 704180, 20.35],
    ['Net Operating Income', 512050, 338220, 51.40],
    ['Net Income', 512050, 338220, 51.40],
]

COMP_LABELS = ('Officer Compensation', 'Payroll Expenses', 'Job Payroll')

CUSTOMERS = [
    {'customer': 'Bridges of Pine Creek - Phase 2', 'current': 0, '1_30': 0,
     '31_60': 12000, '61_90': 40000, '91_and_over': 118000, 'total': 170000},
    {'customer': 'Willow Run HOA', 'current': 22000, '1_30': 8000, '31_60': 0,
     '61_90': 0, '91_and_over': 0, 'total': 30000},
    {'customer': 'Cedar Ridge HOA', 'current': 0, '1_30': 0, '31_60': 0,
     '61_90': 9000, '91_and_over': 26000, 'total': 35000},
    {'customer': 'Maple Grove Condos', 'current': 41000, '1_30': 0, '31_60': 0,
     '61_90': 0, '91_and_over': 0, 'total': 41000},
    {'customer': 'Stonebrook Villas', 'current': 5000, '1_30': 14000,
     '31_60': 6000, '61_90': 0, '91_and_over': 0, 'total': 25000},
    {'customer': 'Harbor Point', 'current': 9000, '1_30': 0, '31_60': 0,
     '61_90': 0, '91_and_over': 4000, 'total': 13000},
]


def _pl_bytes():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Profit and Loss'
    for row in PL_ROWS:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(scope='module')
def pl():
    return office_ops.parse_pl_bytes('Profit_and_Loss.xlsx', _pl_bytes())


@pytest.fixture(scope='module')
def ar():
    return office_ops._build_summary(CUSTOMERS, 'Aug 26, 2026', 'xlsx')


def _labels(lines):
    return [line['label'] for line in lines]


# ── the P&L keeps its line items now ────────────────────────────────────────

def test_the_totals_still_parse_the_way_they_always_did(pl):
    """Adding detail must not move the five numbers the pack already used."""
    assert pl['income_ty'] == 4307000
    assert pl['income_py'] == 3567500
    assert pl['gross_profit_ty'] == 1359500
    assert pl['net_income_ty'] == 512050
    assert pl['cogs_ty'] == 2947500


def test_leaf_accounts_come_through(pl):
    labels = _labels(pl['lines'])
    for expected in ('Services', 'Job Materials', 'Dumpsters', 'Fuel',
                     'Vehicle Repairs'):
        assert expected in labels, expected


def test_rollup_rows_do_not(pl):
    """A category and its children in the same movers table counts the same
    dollars twice, which makes the table wrong rather than noisy."""
    labels = _labels(pl['lines'])
    for rollup in ('Total for Income', 'Total for Expenses', 'Gross Profit',
                   'Net Income', 'Total for Cost of Goods Sold'):
        assert rollup not in labels, rollup


def test_each_line_knows_which_section_it_came_from(pl):
    by_label = {line['label']: line['section'] for line in pl['lines']}
    assert by_label['Services'] == 'income'
    assert by_label['Job Materials'] == 'cogs'
    assert by_label['Fuel'] == 'expenses'


# ── the compensation boundary ───────────────────────────────────────────────

def test_compensation_lines_are_dropped_at_parse_time(pl):
    labels = _labels(pl['lines'])
    for comp in COMP_LABELS:
        assert comp not in labels, f'{comp} reached the pack'


def test_the_count_of_what_was_withheld_is_reported(pl):
    """A list that is silently incomplete reads as complete."""
    assert pl['withheld_comp_lines'] == len(COMP_LABELS)


def test_the_hint_list_is_matched_on_substrings():
    assert office_ops._is_comp_line('   Officer Compensation')
    assert office_ops._is_comp_line('Payroll Taxes')
    assert office_ops._is_comp_line("Owner's Draw")
    assert not office_ops._is_comp_line('Subcontractor Labor')
    assert not office_ops._is_comp_line('Dumpsters')


def test_the_insights_say_that_something_was_withheld(pl):
    text = gen._build_insights({'team_month': {}}, None, pl_summary=pl)
    assert 'compensation line(s) are deliberately not in this pack' in text


# ── what counts as unusual ──────────────────────────────────────────────────

def test_a_big_percentage_on_a_small_line_is_not_unusual(pl):
    """Bank Charges moved 5% on $1,450. Coffee tripling is not news."""
    labels = _labels(gen.pl_movers(pl))
    assert 'Bank Charges' not in labels


def test_a_big_dollar_move_at_a_small_percentage_is_not_unusual(pl):
    """Fuel moved $1,400 on $44k — under both floors. Subcontractor Labor
    moved $75,000, which clears the dollars easily, but only 8%."""
    labels = _labels(gen.pl_movers(pl))
    assert 'Fuel' not in labels
    assert 'Subcontractor Labor' not in labels, (
        '8% on a steady line is not a question for Monday')


def test_both_floors_are_required():
    lines = [
        {'label': 'Clears both', 'ty': 30000, 'py': 10000, 'section': 'expenses'},
        {'label': 'Dollars only', 'ty': 900000, 'py': 880000, 'section': 'expenses'},
        {'label': 'Percent only', 'ty': 1400, 'py': 300, 'section': 'expenses'},
    ]
    labels = _labels(gen.pl_movers({'lines': lines}))
    assert labels == ['Clears both']


def test_a_line_that_appeared_or_stopped_is_called_out(pl):
    movers = {m['label']: m for m in gen.pl_movers(pl)}
    assert movers['Advertising - Radio']['kind'] == 'stopped'
    assert movers['Advertising - Radio']['pct'] is None, 'no percent from zero'
    assert 'nothing this year' in gen._mover_sentence(movers['Advertising - Radio'])
    new = {'label': 'New Account', 'ty': 40000, 'py': 0, 'section': 'expenses'}
    got = gen.pl_movers({'lines': [new]})
    assert got and got[0]['kind'] == 'new'
    assert 'new this year' in gen._mover_sentence(got[0])


def test_an_appearance_too_small_to_matter_is_still_ignored():
    tiny = {'label': 'Petty', 'ty': 400, 'py': 0, 'section': 'expenses'}
    assert gen.pl_movers({'lines': [tiny]}) == []


def test_movers_are_ordered_by_dollars_within_a_section(pl):
    overhead = [m for m in gen.pl_movers(pl) if m['section'] == 'expenses']
    deltas = [abs(m['delta']) for m in overhead]
    assert deltas == sorted(deltas, reverse=True)


def test_the_cap_is_per_section_so_overhead_is_never_crowded_out():
    """Income lines are an order of magnitude bigger than overhead lines. One
    global top-N would be N revenue rows and none of the answer."""
    lines = [{'label': f'Rev {i}', 'ty': 1000000 + i, 'py': 100000,
              'section': 'income'} for i in range(10)]
    lines.append({'label': 'Small overhead', 'ty': 30000, 'py': 10000,
                  'section': 'expenses'})
    labels = _labels(gen.pl_movers({'lines': lines}))
    assert 'Small overhead' in labels
    assert len([l for l in labels if l.startswith('Rev')]) == gen.PL_MOVERS_LIMIT


def test_a_contra_line_reads_as_a_negative_not_a_typo():
    m = {'label': 'Discounts given', 'ty': -21000, 'py': -8500,
         'delta': -12500, 'pct': -1.47, 'kind': 'down', 'section': 'income'}
    sentence = gen._mover_sentence(m)
    assert '-$8,500' in sentence and '$-8,500' not in sentence


# ── the section in the pack ─────────────────────────────────────────────────

def test_the_movers_section_appears_and_is_grouped(pl):
    text = gen._build_insights({'team_month': {}}, None, pl_summary=pl)
    assert 'UNUSUAL MOVES (P&L, year over year)' in text
    assert 'Income lines' in text and 'Job costs' in text and 'Overhead' in text
    assert 'Vehicle Repairs' in text


def test_the_section_states_its_own_thresholds(pl):
    """A filtered list that does not say what it filtered reads as "nothing
    else moved"."""
    text = gen._build_insights({'team_month': {}}, None, pl_summary=pl)
    assert '$2,500' in text and '15%' in text


def test_a_quiet_year_says_so_rather_than_showing_an_empty_section():
    steady = {'lines': [{'label': 'Fuel', 'ty': 44200, 'py': 42800,
                         'section': 'expenses'}]}
    text = gen._build_insights({'team_month': {}}, None, pl_summary=steady)
    assert 'That is the finding, not a gap in the data' in text


def test_no_pl_means_no_movers_section():
    text = gen._build_insights({'team_month': {}}, None, pl_summary=None)
    assert 'UNUSUAL MOVES' not in text


def test_every_heading_the_builder_emits_is_one_the_sheet_styles(pl, ar):
    """The Excel styles headings by exact match against INSIGHT_SECTIONS. A
    heading that is not in that tuple renders as body text and the section
    disappears into the bullets above it.

    The builder emits headings through the SEC_* constants, so the two cannot
    drift; this checks the constants are all actually in the tuple, and that
    every one of them reaches a full render."""
    text = gen._build_insights({'team_month': {8: 400000}, 'invoice_count': 3,
                                'year': 2026,
                                'by_rep_month': {'Andy Potts': {8: 90000}}},
                               ar, pl_summary=pl,
                               notes_by_customer={'Cedar Ridge HOA': 'called'})
    constants = [v for k, v in vars(gen).items()
                 if k.startswith('SEC_') and isinstance(v, str)]
    assert constants, 'the section constants vanished'
    for const in constants:
        assert const in gen.INSIGHT_SECTIONS, f'unstyled heading: {const!r}'
        assert const in text, f'{const!r} never rendered'
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if line in gen.INSIGHT_SECTIONS:
            assert i == 0 or lines[i - 1] == '', (
                f'{line!r} has no blank line above it')


# ── AR: the details that were already computed ──────────────────────────────

def _ar_block(ar, pl=None):
    text = gen._build_insights({'team_month': {8: 400000}, 'invoice_count': 12,
                                'year': 2026}, ar, pl_summary=pl)
    out, started = [], False
    for line in text.splitlines():
        if line.startswith('A/R (company total)'):
            started = True
        elif started and line in gen.INSIGHT_SECTIONS:
            break
        if started:
            out.append(line)
    return '\n'.join(out)


def test_past_due_is_stated_as_a_number_not_left_to_be_added_up(ar):
    block = _ar_block(ar)
    assert 'Past due (anything out of Current): $237,000' in block
    assert '75% of AR' in block


def test_operating_ar_reaches_the_pack(ar):
    """It was computed for the AR draft and never shown here — and it is the
    number they actually steer on, because Bridges is its own conversation."""
    block = _ar_block(ar)
    assert 'Operating AR (excluding Bridges/BOPC): $144,000' in block


def test_the_pack_says_how_many_customers_and_how_many_are_in_91(ar):
    block = _ar_block(ar)
    assert '6 customer(s) carrying a balance' in block
    assert '3 of them have money in 91+' in block


def test_concentration_names_the_largest_balance(ar):
    block = _ar_block(ar)
    assert 'top 5 customers are 96% of AR' in block
    assert 'Bridges of Pine Creek - Phase 2 at $170,000' in block


def test_the_chase_list_is_named_not_just_totalled(ar):
    block = _ar_block(ar)
    assert 'Most overdue by weight' in block
    assert 'Cedar Ridge HOA' in block
    assert 'legacy/BOPC, handle with care' in block


def test_the_old_aging_line_is_still_there(ar):
    """More detail, not different detail."""
    block = _ar_block(ar)
    assert 'Total AR: $314,000' in block
    assert 'Aging: Current $77,000' in block


def test_no_ar_upload_still_says_what_to_do():
    text = gen._build_insights({'team_month': {}}, None)
    assert 'Upload A/R Aging Summary to fill AR.' in text


# ── the finished workbook ───────────────────────────────────────────────────

@pytest.fixture(scope='module')
def workbook(pl, ar):
    raw, insights, meta = gen.generate_from_qb([], ar_summary=ar,
                                               pl_summary=pl, year=2026)
    return raw, insights, meta


def test_the_workbook_still_builds(workbook):
    raw, insights, meta = workbook
    assert raw[:2] == b'PK', 'not a zip — openpyxl did not write a workbook'
    assert meta['pl_included'] is True


def test_the_pl_sheet_carries_the_movers_and_the_detail(workbook):
    import openpyxl
    raw, _insights, _meta = workbook
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert 'P&L Snapshot' in wb.sheetnames
    text = '\n'.join(
        str(c.value) for row in wb['P&L Snapshot'].iter_rows() for c in row
        if c.value is not None)
    assert 'Unusual moves year over year' in text
    assert 'All line items' in text
    assert 'Vehicle Repairs' in text
    assert 'compensation line(s) withheld' in text


def test_no_compensation_line_survives_into_the_workbook(workbook):
    """The guard that matters. The pack is stored in Postgres, rendered here
    and emailed; this looks at the bytes that leave the building."""
    raw, insights, _meta = workbook
    import zipfile
    blob = b''
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        for name in z.namelist():
            blob += z.read(name)
    haystack = blob.decode('utf-8', errors='ignore').lower() + insights.lower()
    # Name the hits and assert on the short list, never on `haystack` itself:
    # a failing `assert x not in <megabytes>` makes pytest render the whole
    # string, which looks like a hung suite rather than a failed test.
    found = [comp for comp in COMP_LABELS if comp.lower() in haystack]
    assert not found, f'compensation lines in the generated pack: {found}'
