"""Monday Numbers Excel — same layout/coloring as Stephanie's Monthly Outlook.

Takes an uploaded "2026 Monthly Outlook.xlsx" and produces a downloadable
Monday report: original sheets preserved + Insights sheet. Difference /
% Hit rows get the same green/red treatment as the live workbook.

Positive: fill #B7E1CD / #C6EFCE, font #006100
Negative: fill #F4C7C3 / #FFC7CE, font #9C0006
"""

from __future__ import annotations

import io
import re
from copy import copy
import hub_time
from datetime import datetime, timezone

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Match the Outlook file's conditional formatting palette (aRGB)
FILL_POS = PatternFill(start_color='FFB7E1CD', end_color='FFB7E1CD', fill_type='solid')
FILL_NEG = PatternFill(start_color='FFF4C7C3', end_color='FFF4C7C3', fill_type='solid')
# Hardcoded Excel "Good/Bad" style used on some Difference cells
FILL_POS_ALT = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')
FILL_NEG_ALT = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
FONT_POS = Font(color='FF006100')
FONT_NEG = Font(color='FF9C0006')
FILL_YELLOW = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
FILL_HEADER = PatternFill(start_color='FF1A5276', end_color='FF1A5276', fill_type='solid')
FONT_HEADER = Font(color='FFFFFFFF', bold=True, name='Calibri', size=12)
FONT_BODY = Font(name='Calibri', size=11)
THIN = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

MONTHS = (
    'Jan', 'Feb', 'March', 'April', 'May', 'June',
    'July', 'Aug', 'Sept', 'October', 'November', 'December',
)


def _apply_pos_neg_cell(cell, value):
    """Paint green if value > 0, red if value < 0 (same idea as Outlook CF)."""
    if value is None:
        return
    try:
        v = float(value)
    except (TypeError, ValueError):
        return
    if v > 0:
        cell.fill = FILL_POS_ALT
        cell.font = Font(
            name=cell.font.name or 'Calibri',
            size=cell.font.size or 11,
            bold=cell.font.bold,
            color='006100',
        )
    elif v < 0:
        cell.fill = FILL_NEG_ALT
        cell.font = Font(
            name=cell.font.name or 'Calibri',
            size=cell.font.size or 11,
            bold=cell.font.bold,
            color='9C0006',
        )


def _ensure_cf_on_range(ws, cell_range, greater_fill, less_fill, less_or_equal=False):
    """Add greaterThan / lessThan conditional formatting if not already present."""
    # Always add — Excel keeps multiple rules; ensures colors after our generation
    pos_rule = CellIsRule(
        operator='greaterThan',
        formula=['0'],
        fill=greater_fill,
        font=FONT_POS,
    )
    neg_op = 'lessThanOrEqual' if less_or_equal else 'lessThan'
    neg_rule = CellIsRule(
        operator=neg_op,
        formula=['0'],
        fill=less_fill,
        font=FONT_NEG,
    )
    ws.conditional_formatting.add(cell_range, pos_rule)
    ws.conditional_formatting.add(cell_range, neg_rule)


def _paint_difference_rows(ws, data_ws, label_col=1, start_col=2, end_col=13):
    """Find Difference / % Hit rows and paint using data_only values when available."""
    for r in range(1, min(ws.max_row or 1, 80) + 1):
        label = ws.cell(r, label_col).value
        if label is None:
            continue
        lab = str(label).strip().lower()
        if lab not in ('difference', '% hit or (miss)', 'hit/miss', 'hit/miss '):
            if 'hit or (miss)' not in lab and lab != 'hit/miss':
                continue
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            # Prefer evaluated value from data_only workbook
            val = None
            if data_ws is not None:
                val = data_ws.cell(r, c).value
            if val is None and cell.value is not None and not (
                isinstance(cell.value, str) and cell.value.startswith('=')
            ):
                val = cell.value
            _apply_pos_neg_cell(cell, val)


def _build_insights(data_wb, ar_summary=None):
    """Text insights from evaluated Monthly Outlook (+ optional AR totals)."""
    lines = []
    # Eastern — see the note in office_ops_generate._build_insights.
    now = hub_time.now().strftime('%Y-%m-%d %H:%M') + ' ET'
    lines.append(f'Office Ops · Monday Numbers insights · generated {now}')
    lines.append('')

    team = data_wb['Monthly Team'] if 'Monthly Team' in data_wb.sheetnames else None
    sales = data_wb['Monthly Sales'] if 'Monthly Sales' in data_wb.sheetnames else None

    # --- Team 2026 block starts ~row 9 ---
    if team:
        # Find 2026 header row
        year_row = None
        for r in range(1, 25):
            if team.cell(r, 1).value == 2026 or team.cell(r, 1).value == 2026.0:
                year_row = r
                break
        if year_row:
            goal_row = year_row + 2  # Goal $
            actual_row = year_row + 4  # Actual $
            diff_row = year_row + 5
            # Current month = last month with actual $ filled
            current_m = None
            ytd_actual = 0.0
            ytd_goal = 0.0
            month_notes = []
            for i, mname in enumerate(MONTHS):
                c = i + 2  # B=2
                actual = team.cell(actual_row, c).value
                goal = team.cell(goal_row, c).value
                if actual is None:
                    continue
                try:
                    actual_f = float(actual)
                    goal_f = float(goal or 0)
                except (TypeError, ValueError):
                    continue
                if actual_f == 0 and goal_f and i >= 7:  # empty future months often 0
                    # still count if explicitly zero mid-year? skip if goal exists and month is future
                    pass
                ytd_actual += actual_f
                ytd_goal += goal_f
                current_m = (mname, actual_f, goal_f, actual_f - goal_f)
                if goal_f:
                    pct = (actual_f - goal_f) / goal_f
                    status = 'HIT' if actual_f >= goal_f else 'MISS'
                    month_notes.append((mname, actual_f, goal_f, pct, status))

            lines.append('## Team (Monthly Team · 2026)')
            if current_m:
                mname, act, goal, diff = current_m
                lines.append(
                    f'- Latest month with actuals: **{mname}** — '
                    f'actual ${_money(act)} vs goal ${_money(goal)} '
                    f'({_pct((act - goal) / goal if goal else 0)})'
                )
            # Full year cell often col 14
            fy_actual = team.cell(actual_row, 14).value
            fy_goal = team.cell(goal_row, 14).value
            if fy_actual is not None:
                try:
                    lines.append(
                        f'- **YTD / Full-year actual (sheet):** ${_money(float(fy_actual))}'
                        + (
                            f' vs full-year goal ${_money(float(fy_goal))}'
                            if fy_goal is not None else ''
                        )
                    )
                except (TypeError, ValueError):
                    pass
            hits = [n for n in month_notes if n[4] == 'HIT']
            misses = [n for n in month_notes if n[4] == 'MISS']
            if hits:
                lines.append(
                    '- Months **at/above** goal: '
                    + ', '.join(f'{n[0]} ({_pct(n[3])})' for n in hits)
                )
            if misses:
                lines.append(
                    '- Months **below** goal: '
                    + ', '.join(f'{n[0]} ({_pct(n[3])})' for n in misses)
                )
            lines.append('')

    # --- Per consultant current month ---
    if sales:
        lines.append('## Consultants (Monthly Sales)')
        # Blocks of 6 rows: Historial, Goal, Actual, Difference, % Hit, blank
        # Labels in col A; scan for "* Actual"
        blocks = []
        for r in range(1, 45):
            lab = sales.cell(r, 1).value
            if not lab or not str(lab).endswith('Actual'):
                continue
            name = str(lab).replace(' Actual', '').strip()
            if name in ('PPS Historial',):
                continue
            # find latest month with actual
            latest = None
            for i, mname in enumerate(MONTHS):
                c = i + 2
                val = sales.cell(r, c).value
                if val is None:
                    continue
                try:
                    v = float(val)
                except (TypeError, ValueError):
                    continue
                goal = sales.cell(r - 1, c).value  # Goal row above Actual
                try:
                    g = float(goal) if goal not in (None, '$ -', '') else 0.0
                except (TypeError, ValueError):
                    g = 0.0
                latest = (mname, v, g)
            if latest:
                mname, v, g = latest
                diff = v - g
                status = 'hit' if (g and v >= g) or (not g and v > 0) else 'behind'
                blocks.append((name, mname, v, g, diff, status))
                lines.append(
                    f'- **{name}** ({mname}): ${_money(v)}'
                    + (f' vs ${_money(g)} goal ({_pct(diff / g if g else 0)})' if g else '')
                    + f' — {status}'
                )
        lines.append('')

    # --- AR totals only (Stephanie owns AR) ---
    lines.append('## A/R (company total — Office Ops)')
    if ar_summary and ar_summary.get('grand_total'):
        g = ar_summary['grand_total']
        lines.append(f'- **Total AR:** ${_money(g.get("total"))}')
        lines.append(
            f'- Current ${_money(g.get("current"))} · 1–30 ${_money(g.get("1_30"))} · '
            f'31–60 ${_money(g.get("31_60"))} · 61–90 ${_money(g.get("61_90"))} · '
            f'91+ ${_money(g.get("91_and_over"))}'
        )
        if ar_summary.get('bopc'):
            b = ar_summary['bopc']
            lines.append(
                f'- **BOPC / Bridges (included in total):** ${_money(b.get("total"))} '
                f'(91+: ${_money(b.get("91_and_over"))})'
            )
            op = ar_summary.get('operating_ex_bopc') or {}
            if op.get('total') is not None:
                lines.append(f'- **Operating AR (ex-BOPC, rough):** ${_money(op.get("total"))}')
        lines.append('- Collections: Stephanie owns A/R follow-up (not split by rep).')
    else:
        lines.append('- _Upload A/R Aging Summary on Office Ops to fill company AR totals._')
    lines.append('')
    lines.append('—')
    lines.append('Layout/colors match Monthly Outlook. Edit Actual $ in the live sheet; re-upload to regenerate.')
    return '\n'.join(lines)


def _money(n):
    try:
        return f'{float(n or 0):,.0f}'
    except (TypeError, ValueError):
        return '0'


def _pct(n):
    try:
        return f'{float(n):+.0%}'
    except (TypeError, ValueError):
        return 'n/a'


def generate_monday_report(outlook_bytes, ar_summary=None):
    """Return (xlsx_bytes, insights_text, meta dict)."""
    bio = io.BytesIO(outlook_bytes)
    # data_only for insights (uses cached values from Excel)
    try:
        data_wb = load_workbook(io.BytesIO(outlook_bytes), data_only=True)
    except Exception:
        data_wb = None

    wb = load_workbook(bio)

    # Reinforce conditional formatting on known ranges (same as source file)
    if 'Monthly Team' in wb.sheetnames:
        ws = wb['Monthly Team']
        data_ws = data_wb['Monthly Team'] if data_wb and 'Monthly Team' in data_wb.sheetnames else None
        _ensure_cf_on_range(ws, 'B6:M7', FILL_POS, FILL_NEG)
        _ensure_cf_on_range(ws, 'B14:M15', FILL_POS, FILL_NEG, less_or_equal=True)
        _paint_difference_rows(ws, data_ws)

    if 'Monthly Sales' in wb.sheetnames:
        ws = wb['Monthly Sales']
        data_ws = data_wb['Monthly Sales'] if data_wb and 'Monthly Sales' in data_wb.sheetnames else None
        _ensure_cf_on_range(ws, 'B5:M6', FILL_POS, FILL_NEG, less_or_equal=True)
        _ensure_cf_on_range(ws, 'B11:M12', FILL_POS, FILL_NEG, less_or_equal=True)
        _ensure_cf_on_range(ws, 'B17:M18', FILL_POS, FILL_NEG, less_or_equal=True)
        _ensure_cf_on_range(ws, 'B23:M24', FILL_POS, FILL_NEG, less_or_equal=True)
        _ensure_cf_on_range(ws, 'C29:M30', FILL_POS, FILL_NEG, less_or_equal=True)
        _paint_difference_rows(ws, data_ws)

    if 'Quarterly Breakdowns' in wb.sheetnames:
        ws = wb['Quarterly Breakdowns']
        data_ws = data_wb['Quarterly Breakdowns'] if data_wb and 'Quarterly Breakdowns' in data_wb.sheetnames else None
        _ensure_cf_on_range(ws, 'B4:E5', FILL_POS, FILL_NEG)
        _ensure_cf_on_range(ws, 'D8:D12', FILL_POS, FILL_NEG)
        _ensure_cf_on_range(ws, 'I8:I12', FILL_POS, FILL_NEG)
        _ensure_cf_on_range(ws, 'D15:D19', FILL_POS, FILL_NEG)
        _ensure_cf_on_range(ws, 'I15:I19', FILL_POS, FILL_NEG)
        _paint_difference_rows(ws, data_ws, end_col=9)

    insights = _build_insights(data_wb or wb, ar_summary=ar_summary)

    # Insights sheet (first)
    if 'Insights' in wb.sheetnames:
        del wb['Insights']
    ws_i = wb.create_sheet('Insights', 0)
    ws_i['A1'] = 'Monday Numbers · Insights'
    ws_i['A1'].font = FONT_HEADER
    ws_i['A1'].fill = FILL_HEADER
    ws_i.merge_cells('A1:B1')
    ws_i.column_dimensions['A'].width = 100
    for i, line in enumerate(insights.splitlines(), start=3):
        cell = ws_i.cell(i, 1, line)
        cell.font = FONT_BODY
        if line.startswith('## '):
            cell.font = Font(name='Calibri', size=12, bold=True, color='1A5276')

    # AR Totals sheet (simple company rollup)
    if 'AR Totals' in wb.sheetnames:
        del wb['AR Totals']
    ws_ar = wb.create_sheet('AR Totals', 1)
    ws_ar['A1'] = 'Company A/R (Stephanie owns collections)'
    ws_ar['A1'].font = FONT_HEADER
    ws_ar['A1'].fill = FILL_HEADER
    ws_ar.merge_cells('A1:C1')
    headers = ['Bucket', 'Amount']
    for i, h in enumerate(headers, 1):
        cell = ws_ar.cell(3, i, h)
        cell.font = Font(bold=True)
        cell.fill = FILL_YELLOW
    if ar_summary and ar_summary.get('grand_total'):
        g = ar_summary['grand_total']
        rows = [
            ('Current', g.get('current')),
            ('1–30', g.get('1_30')),
            ('31–60', g.get('31_60')),
            ('61–90', g.get('61_90')),
            ('91+', g.get('91_and_over')),
            ('TOTAL AR', g.get('total')),
        ]
        for i, (lab, amt) in enumerate(rows, 4):
            ws_ar.cell(i, 1, lab)
            cell = ws_ar.cell(i, 2, float(amt or 0))
            cell.number_format = '"$"#,##0'
            if lab == 'TOTAL AR':
                ws_ar.cell(i, 1).font = Font(bold=True)
                cell.font = Font(bold=True)
        r = 11
        if ar_summary.get('bopc'):
            ws_ar.cell(r, 1, 'BOPC / Bridges (in total)')
            c = ws_ar.cell(r, 2, float(ar_summary['bopc'].get('total') or 0))
            c.number_format = '"$"#,##0'
            r += 1
            op = ar_summary.get('operating_ex_bopc') or {}
            ws_ar.cell(r, 1, 'Operating AR ex-BOPC (rough)')
            c = ws_ar.cell(r, 2, float(op.get('total') or 0))
            c.number_format = '"$"#,##0'
    else:
        ws_ar['A4'] = 'Upload A/R Aging Summary on Office Ops to populate this sheet.'
    ws_ar.column_dimensions['A'].width = 36
    ws_ar.column_dimensions['B'].width = 16

    out = io.BytesIO()
    wb.save(out)
    raw = out.getvalue()
    meta = {
        'bytes': len(raw),
        'sheets': wb.sheetnames,
        'generated_at': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
    }
    return raw, insights, meta
