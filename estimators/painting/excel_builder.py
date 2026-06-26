"""PPS Exterior Painting Estimator — Excel workbook."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .calculator import calculate_painting_estimate

DARK_BLUE = '004C8C'
BLUE = '0096D6'
WARNING = 'FFF3CD'
WHITE = 'FFFFFF'


def build_estimate_excel(job, line_items, inputs, confidence=None):
    calc = calculate_painting_estimate(line_items, inputs)
    wb = Workbook()
    wb.remove(wb.active)

    bdr = Border(
        left=Side(style='thin', color='D0DCE8'),
        right=Side(style='thin', color='D0DCE8'),
        top=Side(style='thin', color='D0DCE8'),
        bottom=Side(style='thin', color='D0DCE8'),
    )

    ws1 = wb.create_sheet('1 – Job Summary')
    ws2 = wb.create_sheet('2 – Takeoff')
    ws3 = wb.create_sheet('3 – Bid Summary')

    # --- Job Summary ---
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 18
    ws1.merge_cells('B2:C2')
    ws1['B2'] = 'PPS — Exterior Painting Estimate'
    ws1['B2'].font = Font(bold=True, size=14, color=WHITE)
    ws1['B2'].fill = PatternFill('solid', start_color=DARK_BLUE)

    r = 4
    for lbl, val in [
        ('Property:', job.get('property_name', '')),
        ('Address:', job.get('address', '')),
        ('Estimator:', job.get('estimator', '')),
        ('Date:', job.get('date', '')),
        ('Report #:', job.get('report_number', '')),
    ]:
        ws1[f'B{r}'] = lbl
        ws1[f'B{r}'].font = Font(bold=True, color=DARK_BLUE)
        ws1[f'C{r}'] = val
        r += 1

    if confidence:
        from estimators.reliability import reliability_excel_lines
        for lbl, val in reliability_excel_lines(confidence):
            ws1[f'B{r}'] = lbl
            ws1[f'B{r}'].font = Font(bold=True, color=DARK_BLUE, size=10)
            ws1[f'C{r}'] = val
            ws1[f'C{r}'].font = Font(size=10)
            r += 1
        r += 1

    assump_row = r
    ws1.merge_cells(f'B{assump_row}:C{assump_row}')
    ws1[f'B{assump_row}'] = 'ASSUMPTIONS (yellow = editable)'
    ws1[f'B{assump_row}'].font = Font(bold=True, color=WHITE)
    ws1[f'B{assump_row}'].fill = PatternFill('solid', start_color=DARK_BLUE)

    assumptions = [
        ('Labor $/hour', inputs.get('labor_per_hour', 38), '_($* #,##0.00_)'),
        ('One-coat margin (%)', inputs.get('margin_one_coat_pct', 42), '0.0%'),
        ('Two-coat margin (%)', inputs.get('margin_two_coat_pct', 38), '0.0%'),
        ('Two-coat multiplier', inputs.get('two_coat_multiplier', 1.6), '0.0'),
    ]
    rr = assump_row + 1
    margin_one_row = margin_two_row = two_mult_row = labor_row = rr
    for lbl, val, fmt in assumptions:
        ws1[f'B{rr}'] = lbl
        ws1[f'C{rr}'] = val
        ws1[f'B{rr}'].border = bdr
        ws1[f'C{rr}'].border = bdr
        ws1[f'C{rr}'].number_format = fmt
        ws1[f'C{rr}'].fill = PatternFill('solid', start_color=WARNING)
        if 'One-coat margin' in lbl:
            margin_one_row = rr
        elif 'Two-coat margin' in lbl:
            margin_two_row = rr
        elif 'Two-coat multiplier' in lbl:
            two_mult_row = rr
        elif 'Labor' in lbl:
            labor_row = rr
        rr += 1

    bid_row = rr + 1
    ws1[f'B{bid_row}'] = 'One-Coat Bid'
    ws1[f'B{bid_row}'].font = Font(bold=True)
    ws1[f'C{bid_row}'] = "='3 – Bid Summary'!C13"
    ws1[f'C{bid_row}'].number_format = '_($* #,##0.00_)'
    ws1[f'B{bid_row + 1}'] = 'Two-Coat Bid'
    ws1[f'B{bid_row + 1}'].font = Font(bold=True, color=WHITE)
    ws1[f'C{bid_row + 1}'] = "='3 – Bid Summary'!D13"
    ws1[f'C{bid_row + 1}'].number_format = '_($* #,##0.00_)'
    ws1[f'B{bid_row + 1}'].fill = PatternFill('solid', start_color=DARK_BLUE)
    ws1[f'C{bid_row + 1}'].fill = PatternFill('solid', start_color=DARK_BLUE)
    ws1[f'C{bid_row + 1}'].font = Font(bold=True, color=WHITE)

    # --- Takeoff ---
    ws2.sheet_view.showGridLines = False
    for col, w in [('B', 14), ('C', 22), ('D', 10), ('E', 10), ('F', 10), ('G', 12), ('H', 12), ('I', 12)]:
        ws2.column_dimensions[col].width = w
    ws2['B2'] = 'TAKEOFF & LINE CALCULATIONS'
    ws2['B2'].font = Font(bold=True, color=DARK_BLUE)
    headers = [
        ('B', 'Type'), ('C', 'Category'), ('D', 'Measured'), ('E', 'Unit'),
        ('F', 'Hours'), ('G', 'Labor $'), ('H', 'Paint $'), ('I', 'Subtotal'),
    ]
    row = 4
    for col, h in headers:
        ws2[f'{col}{row}'] = h
        ws2[f'{col}{row}'].font = Font(bold=True, color=WHITE)
        ws2[f'{col}{row}'].fill = PatternFill('solid', start_color=BLUE)
    row += 1
    first_data = row
    for line in calc['lines']:
        ws2[f'B{row}'] = line['exterior_type']
        ws2[f'C{row}'] = line['category']
        ws2[f'D{row}'] = line['measured']
        ws2[f'E{row}'] = line['unit']
        ws2[f'F{row}'] = line['hours']
        ws2[f'G{row}'] = line['labor_cost']
        ws2[f'H{row}'] = line['paint_cost']
        ws2[f'I{row}'] = line['subtotal']
        for c in 'BCDEFGHI':
            ws2[f'{c}{row}'].border = bdr
        for c in 'DFGHI':
            ws2[f'{c}{row}'].number_format = '_($* #,##0.00_)' if c in 'GHI' else '0.00'
        row += 1
    last_data = row - 1
    if calc['lines']:
        ws2[f'G{row}'] = f'=SUM(G{first_data}:G{last_data})'
        ws2[f'H{row}'] = f'=SUM(H{first_data}:H{last_data})'
        ws2[f'I{row}'] = f'=SUM(I{first_data}:I{last_data})'
        ws2[f'F{row}'] = f'=SUM(F{first_data}:F{last_data})'
        ws2[f'B{row}'] = 'Grand Total'
        ws2[f'B{row}'].font = Font(bold=True)
        for c in 'FGHI':
            ws2[f'{c}{row}'].font = Font(bold=True)
            ws2[f'{c}{row}'].number_format = '_($* #,##0.00_)' if c in 'GHI' else '0.00'
        takeoff_labor_row = row
        takeoff_mat_row = row
        takeoff_total_row = row
    else:
        takeoff_labor_row = takeoff_mat_row = takeoff_total_row = 4

    # --- Bid Summary ---
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 16
    ws3['A1'] = 'Row Labels'
    ws3['B1'] = 'Total Material'
    ws3['C1'] = 'Total Labor'
    for c in 'ABC':
        ws3[f'{c}1'].font = Font(bold=True, color=WHITE)
        ws3[f'{c}1'].fill = PatternFill('solid', start_color=BLUE)

    r = 2
    type_start = r
    for ts in calc['by_type']:
        ws3[f'A{r}'] = ts['exterior_type']
        ws3[f'B{r}'] = ts['material']
        ws3[f'C{r}'] = ts['labor']
        r += 1
    type_end = r - 1
    ws3[f'A{r}'] = 'Grand Total'
    ws3[f'A{r}'].font = Font(bold=True)
    if calc['by_type']:
        ws3[f'B{r}'] = f'=SUM(B{type_start}:B{type_end})'
        ws3[f'C{r}'] = f'=SUM(C{type_start}:C{type_end})'
    else:
        ws3[f'B{r}'] = 0
        ws3[f'C{r}'] = 0
    grand_row = r

    ws3['B10'] = 'One Coat'
    ws3['C10'] = 'Two Coats'
    ws3['B10'].font = ws3['C10'].font = Font(bold=True)
    ws3['A11'] = 'Labor + Material'
    ws3['B11'] = f"=C{grand_row}+B{grand_row}"
    ws3['C11'] = f"=B11*'1 – Job Summary'!C{two_mult_row}"
    cost_row = 11
    ws3['A12'] = 'Margin %'
    ws3['B12'] = f"='1 – Job Summary'!C{margin_one_row}/100"
    ws3['C12'] = f"='1 – Job Summary'!C{margin_two_row}/100"
    ws3['B12'].number_format = ws3['C12'].number_format = '0.0%'
    ws3['A13'] = 'Bid Total'
    ws3['B13'] = f'=B11/(1-B12)'
    ws3['C13'] = f'=C11/(1-C12)'
    ws3['A14'] = 'Profit'
    ws3['B14'] = '=B13-B11'
    ws3['C14'] = '=C13-C11'
    for rr in (11, 13, 14):
        for cc in 'BC':
            ws3[f'{cc}{rr}'].number_format = '_($* #,##0.00_)'
    ws3['A13'].font = Font(bold=True, color=WHITE)
    ws3['B13'].font = Font(bold=True, color=WHITE)
    ws3['C13'].font = Font(bold=True, color=WHITE)
    ws3['A13'].fill = PatternFill('solid', start_color=DARK_BLUE)
    ws3['B13'].fill = PatternFill('solid', start_color=DARK_BLUE)
    ws3['C13'].fill = PatternFill('solid', start_color=DARK_BLUE)

    from estimators.excel_branding import brand_estimate_workbook
    brand_estimate_workbook(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf