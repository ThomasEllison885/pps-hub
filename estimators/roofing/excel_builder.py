"""PPS Roofing Estimator — Excel output (quick bid or full GAF material list)."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .calculator import calculate_materials, calculate_bid_summary
from .material_catalog import MATERIAL_LINES, GAF_DEFAULTS

DARK_BLUE = '004C8C'
BLUE = '0096D6'
LIGHT_BLUE = 'EBF6FC'
WHITE = 'FFFFFF'
GRAY_HDR = 'F2F7FB'
WARNING = 'FFF3CD'

REPORT_LABELS = {
    'bid_perfect': 'EagleView Bid Perfect',
    'premium': 'EagleView Premium',
    'roofr': 'Roofr',
    'unknown': 'Unknown',
}


def build_estimate_excel(job, measurements, inputs, pricing=None):
    pricing = pricing or {}
    report_type = measurements.get('report_type', 'premium')
    is_quick = report_type == 'bid_perfect'

    wb = Workbook()
    wb.remove(wb.active)

    if is_quick:
        summary = calculate_bid_summary(measurements, inputs)
        _build_quick_bid(wb, job, measurements, inputs, summary, pricing)
    else:
        qty = calculate_materials(measurements, inputs)
        _build_full_estimate(wb, job, measurements, inputs, qty, pricing)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _border():
    s = Side(style='thin', color='D0DCE8')
    return Border(left=s, right=s, top=s, bottom=s)


def _head(ws, row, col, text, span=4):
    from openpyxl.utils import get_column_letter
    end = get_column_letter(col + span - 1)
    start = get_column_letter(col)
    ws.merge_cells(f'{start}{row}:{end}{row}')
    c = ws[f'{start}{row}']
    c.value = text
    c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
    c.fill = PatternFill('solid', start_color=DARK_BLUE)
    c.alignment = Alignment(horizontal='left', indent=1)


def _money(ws, row, col):
    ws[f'{col}{row}'].number_format = '_($* #,##0.00_)'
    ws[f'{col}{row}'].alignment = Alignment(horizontal='right')


def _build_full_estimate(wb, job, measurements, inputs, qty, pricing):
    ws_sum = wb.create_sheet('1 – Job Summary')
    ws_meas = wb.create_sheet('2 – Measurements')
    ws_mat = wb.create_sheet('3 – Materials')
    ws_tot = wb.create_sheet('4 – Bid Total')

    waste_row = 14
    labor_row = 15
    dump_div_row = 16
    dump_cost_row = 17
    tax_row = 18
    margin_row = 19

    # --- Summary ---
    ws_sum.sheet_view.showGridLines = False
    ws_sum.column_dimensions['B'].width = 28
    ws_sum.column_dimensions['C'].width = 16
    ws_sum.merge_cells('B2:E2')
    ws_sum['B2'] = 'PURE PROPERTY SOLUTIONS — Roofing Estimate (GAF)'
    ws_sum['B2'].font = Font(bold=True, size=14, color=WHITE)
    ws_sum['B2'].fill = PatternFill('solid', start_color=DARK_BLUE)

    row = 4
    for lbl, val in [
        ('Property:', job.get('property_name', '')),
        ('Address:', job.get('address', '')),
        ('Estimator:', job.get('estimator', '')),
        ('Date:', job.get('date', '')),
        ('Report #:', measurements.get('report_number', '')),
        ('Report Type:', REPORT_LABELS.get(measurements.get('report_type'), '')),
    ]:
        ws_sum[f'B{row}'] = lbl
        ws_sum[f'B{row}'].font = Font(bold=True, color=DARK_BLUE)
        ws_sum[f'C{row}'] = val
        row += 1

    _head(ws_sum, 11, 2, 'ASSUMPTIONS — yellow cells drive formulas', 4)
    assumptions = [
        (12, 'Shingle Product', 'GAF Timberline HDZ', None),
        (13, 'Predominant Pitch', measurements.get('predominant_pitch', ''), None),
        (waste_row, 'Waste Factor (%)', inputs.get('waste_pct', 12), '0'),
        (labor_row, 'Labor $ / sq (order)', inputs.get('labor_per_sq', 60), '_($* #,##0_)'),
        (dump_div_row, 'Dump: squares per load', inputs.get('dump_divisor', 45), '0'),
        (dump_cost_row, 'Dump $ per load', inputs.get('dump_cost', 200), '_($* #,##0_)'),
        (tax_row, 'Material Tax (%)', inputs.get('tax_pct', 7.5), '0.0%'),
        (margin_row, 'Target Margin (%)', inputs.get('margin_pct', 25), '0.0%'),
        (20, 'Pipe Boots (manual)', inputs.get('pipe_boots', 0), '0'),
    ]
    bdr = _border()
    for rr, lbl, val, fmt in assumptions:
        ws_sum[f'B{rr}'] = lbl
        ws_sum[f'C{rr}'] = val
        for col in ('B', 'C'):
            ws_sum[f'{col}{rr}'].border = bdr
        if fmt:
            ws_sum[f'C{rr}'].number_format = fmt
            ws_sum[f'C{rr}'].fill = PatternFill('solid', start_color=WARNING)

    ws_sum['B22'] = 'Order Squares (with waste)'
    ws_sum['B22'].font = Font(bold=True)
    ws_sum['C22'] = qty['order_squares']

    # --- Measurements ---
    ws_meas.sheet_view.showGridLines = False
    _head(ws_meas, 2, 2, 'PARSED MEASUREMENTS', 3)
    meas_rows = [
        ('Total roof area (sq ft)', qty['roof_area_sqft']),
        ('Roof squares (measured)', qty['roof_area_squares']),
        ('Order squares (waste)', qty['order_squares']),
        ('Facets', measurements.get('facets', '')),
        ('Eaves (ft)', measurements.get('eaves_ft', '')),
        ('Rakes (ft)', measurements.get('rakes_ft', '')),
        ('Valleys (ft)', measurements.get('valleys_ft', '')),
        ('Ridges + hips (ft)', _ridge_cap_display(measurements)),
        ('Step flashing (ft)', measurements.get('step_flashing_ft', '')),
        ('Wall flashing (ft)', measurements.get('wall_flashing_ft', '')),
        ('Ice & water basis (ft)', qty['ice_water_lf']),
        ('Starter basis (ft)', qty['starter_lf']),
    ]
    r = 4
    for lbl, val in meas_rows:
        ws_meas[f'B{r}'] = lbl
        ws_meas[f'C{r}'] = val
        r += 1

    # --- Materials ---
    ws_mat.sheet_view.showGridLines = False
    ws_mat.column_dimensions['B'].width = 38
    ws_mat.column_dimensions['C'].width = 10
    ws_mat.column_dimensions['D'].width = 14
    ws_mat.column_dimensions['E'].width = 14
    _head(ws_mat, 2, 2, 'GAF MATERIAL LIST', 4)
    for col, txt in [('B', 'Item'), ('C', 'Qty'), ('D', 'Unit Price'), ('E', 'Extended')]:
        ws_mat[f'{col}4'] = txt
        ws_mat[f'{col}4'].font = Font(bold=True, color=WHITE)
        ws_mat[f'{col}4'].fill = PatternFill('solid', start_color=BLUE)

    r = 5
    first = r
    for key, label, unit, qty_key in MATERIAL_LINES:
        ws_mat[f'B{r}'] = label
        ws_mat[f'C{r}'] = qty.get(qty_key, 0)
        price = pricing.get(key)
        ws_mat[f'D{r}'] = price if price is not None else ''
        ws_mat[f'E{r}'] = f'=IF(D{r}="","",C{r}*D{r})'
        if price is None:
            ws_mat[f'D{r}'].fill = PatternFill('solid', start_color=WARNING)
        _money(ws_mat, r, 'DE')
        for col in 'BCDE':
            ws_mat[f'{col}{r}'].border = bdr
        r += 1

    mat_total_row = r
    ws_mat[f'B{mat_total_row}'] = 'Material Subtotal'
    ws_mat[f'B{mat_total_row}'].font = Font(bold=True)
    ws_mat[f'E{mat_total_row}'] = f'=SUM(E{first}:E{mat_total_row - 1})'
    _money(ws_mat, mat_total_row, 'E')

    # --- Totals ---
    ws_tot.sheet_view.showGridLines = False
    ws_tot.column_dimensions['B'].width = 32
    ws_tot.column_dimensions['C'].width = 16
    _head(ws_tot, 2, 2, 'BID TOTAL', 2)

    def tline(row, label, formula, yellow=False):
        ws_tot[f'B{row}'] = label
        ws_tot[f'C{row}'] = formula
        ws_tot[f'B{row}'].border = bdr
        ws_tot[f'C{row}'].border = bdr
        _money(ws_tot, row, 'C')
        if yellow:
            ws_tot[f'C{row}'].fill = PatternFill('solid', start_color=WARNING)

    tline(4, 'Material Subtotal', f"='3 – Materials'!E{mat_total_row}")
    tline(5, 'Material Tax', f"=C4*('1 – Job Summary'!C{tax_row}/100)")
    tline(6, 'Material Grand Total', '=C4+C5')
    tline(7, 'Labor', f"='1 – Job Summary'!C{labor_row}*'1 – Job Summary'!C22")
    tline(8, 'Dump / Haul', f"=ROUNDUP('1 – Job Summary'!C22/'1 – Job Summary'!C{dump_div_row},0)*'1 – Job Summary'!C{dump_cost_row}")
    tline(9, 'Cost Before Margin', '=C6+C7+C8', False)
    ws_tot['B9'].font = Font(bold=True)
    ws_tot['C9'].font = Font(bold=True)
    tline(11, 'Invoice (with margin)', f"=C9/(1-'1 – Job Summary'!C{margin_row}/100)")
    ws_tot['B11'].font = Font(bold=True, color=WHITE)
    ws_tot['C11'].font = Font(bold=True, color=WHITE)
    ws_tot['B11'].fill = PatternFill('solid', start_color=DARK_BLUE)
    ws_tot['C11'].fill = PatternFill('solid', start_color=DARK_BLUE)

    ws_sum['B24'] = 'Invoice Total'
    ws_sum['C24'] = "='4 – Bid Total'!C11"
    _money(ws_sum, 24, 'C')


def _ridge_cap_display(measurements):
    r = measurements.get('ridges_ft') or 0
    h = measurements.get('hips_ft') or 0
    return round(r + h, 1) if (r or h) else ''


def _build_quick_bid(wb, job, measurements, inputs, summary, pricing):
    ws = wb.create_sheet('Quick Bid')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    ws.merge_cells('B2:D2')
    ws['B2'] = 'PPS ROOFING — QUICK BID (Bid Perfect)'
    ws['B2'].font = Font(bold=True, size=14, color=WHITE)
    ws['B2'].fill = PatternFill('solid', start_color=DARK_BLUE)

    row = 4
    for lbl, val in [
        ('Property:', job.get('property_name', '')),
        ('Address:', job.get('address', '')),
        ('Estimator:', job.get('estimator', '')),
        ('Date:', job.get('date', '')),
        ('Report #:', measurements.get('report_number', '')),
        ('Waste %', inputs.get('waste_pct', 12)),
    ]:
        ws[f'B{row}'] = lbl
        ws[f'C{row}'] = val
        row += 1

    row += 1
    ws[f'B{row}'] = 'STRUCTURES'
    ws[f'B{row}'].font = Font(bold=True, color=DARK_BLUE)
    row += 1
    for col, txt in [('B', 'Structure'), ('C', 'Squares'), ('D', 'Order Sq')]:
        ws[f'{col}{row}'] = txt
        ws[f'{col}{row}'].font = Font(bold=True, color=WHITE)
        ws[f'{col}{row}'].fill = PatternFill('solid', start_color=BLUE)
    row += 1
    for s in summary.get('structures') or [{'label': 'Total', **summary}]:
        ws[f'B{row}'] = s.get('label', 'Structure')
        ws[f'C{row}'] = s.get('squares', summary.get('roof_area_squares'))
        ws[f'D{row}'] = s.get('order_squares', summary.get('order_squares'))
        row += 1

    row += 1
    mat_psq = inputs.get('material_per_sq', 65)
    labor_psq = inputs.get('labor_per_sq', 60)
    items = [
        ('Measured squares', summary['roof_area_squares']),
        ('Order squares (waste)', summary['order_squares']),
        (f'Material @ ${mat_psq}/sq', summary['material_cost']),
        (f'Labor @ ${labor_psq}/sq', summary['labor_cost']),
        (f'Dump ({summary["dump_loads"]} loads)', summary['dump_cost']),
        ('Tax', summary['tax']),
        ('Cost before margin', summary['cost_before_margin']),
        (f'Bid total ({summary["margin_pct"]}% margin)', summary['grand_total']),
    ]
    for lbl, val in items:
        ws[f'B{row}'] = lbl
        ws[f'C{row}'] = val
        if isinstance(val, float):
            _money(ws, row, 'C')
        row += 1

    ws.merge_cells(f'B{row + 1}:D{row + 1}')
    ws[f'B{row + 1}'] = (
        'Bid Perfect reports lack linear measurements. '
        'Upload Premium EagleView or Roofr for a full GAF material list.'
    )
    ws[f'B{row + 1}'].font = Font(size=9, italic=True, color='666666')