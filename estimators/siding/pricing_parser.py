"""Parse siding pricing uploads (CSV or Excel)."""
import csv
import io
import re

from .material_catalog import DETAIL_LINES, LIBRARY_LINES

# Friendly aliases → internal price keys used in Excel
KEY_ALIASES = {
    'siding': 'siding_sq',
    'siding_sq': 'siding_sq',
    'siding_primary_product': 'siding_sq',
    'vinyl siding': 'siding_sq',
    'ndx vinyl siding': 'siding_sq',
    'ndx vinyl siding_per_sq': 'siding_sq',
    'ndx woodsman': 'siding_sq',
    'starter': 'starter_piece',
    'starter_piece': 'starter_piece',
    'qa starter': 'starter_piece',
    'starter strip': 'starter_piece',
    'starter strip_receiver_track': 'starter_piece',
    'starter strip / receiver track': 'starter_piece',
    'j-channel': 'jchannel_piece',
    'j_channel': 'jchannel_piece',
    'j channel': 'jchannel_piece',
    'jchannel': 'jchannel_piece',
    'ndx 5/8 j channel': 'jchannel_piece',
    'ndx 5_8 j channel': 'jchannel_piece',
    'jchannel_piece': 'jchannel_piece',
    'inside corner': 'inside_corner_post',
    'inside corner post': 'inside_corner_post',
    'ndx inside corner 3/4': 'inside_corner_post',
    'ndx inside corner_3_4': 'inside_corner_post',
    'inside_corner_post': 'inside_corner_post',
    'outside corner': 'outside_corner_post',
    'outside corner post': 'outside_corner_post',
    "ndx 12' outside corner": 'outside_corner_post',
    'ndx 12_outside_corner': 'outside_corner_post',
    'outside_corner_post': 'outside_corner_post',
    'fascia': 'fascia_piece',
    'coil stock': 'fascia_piece',
    'roll of coil stock': 'fascia_piece',
    'fascia_piece': 'fascia_piece',
    'fascia cover_aluminum coil stock': 'fascia_piece',
    'utility trim': 'utility_piece',
    'universal trim': 'utility_piece',
    'ndx universal trim': 'utility_piece',
    'utility_piece': 'utility_piece',
    'under_sill_utility trim': 'utility_piece',
    'house wrap': 'housewrap_roll',
    'housewrap_roll': 'housewrap_roll',
    'house wrap_tyk 9sq roll': 'housewrap_roll',
    'house wrap tape': 'housewrap_tape',
    'housewrap_tape': 'housewrap_tape',
    'fan fold': 'fanfold_sq',
    'fanfold_sq': 'fanfold_sq',
    'fan fold insulation': 'fanfold_sq',
    'soffit': 'soffit_piece',
    'soffit_piece': 'soffit_piece',
    'j block': 'jblock_uniblock',
    'j block uniblock': 'jblock_uniblock',
    'jblock_uniblock': 'jblock_uniblock',
    'm block': 'jblock_mblock',
    'j block m block': 'jblock_mblock',
    'jblock_mblock': 'jblock_mblock',
    'exhaust vent': 'exhaust_vent',
    'exhaust_vent': 'exhaust_vent',
    'roofing nails': 'roofing_nails',
    'roofing_nails': 'roofing_nails',
    'cap nails': 'cap_nails',
    'cap_nails': 'cap_nails',
    'haul off': 'haul_off_building',
    'haul_off_building': 'haul_off_building',
}

# Substring hints (order matters — more specific first)
FUZZY_HINTS = [
    ('housewrap_tape', ('house wrap tape', 'housewrap tape')),
    ('housewrap_roll', ('house wrap', 'housewrap', 'tyvek')),
    ('jblock_mblock', ('m block', 'mblock')),
    ('jblock_uniblock', ('uniblock', 'j block')),
    ('outside_corner_post', ('outside corner',)),
    ('inside_corner_post', ('inside corner',)),
    ('jchannel_piece', ('j channel', 'j-channel', 'jchannel')),
    ('starter_piece', ('qa starter', 'starter strip', 'starter')),
    ('utility_piece', ('universal trim', 'utility trim', 'under-sill')),
    ('fascia_piece', ('coil stock', 'fascia cover')),
    ('exhaust_vent', ('exhaust vent',)),
    ('roofing_nails', ('roofing nail',)),
    ('cap_nails', ('cap nail',)),
    ('fanfold_sq', ('fan fold', 'fanfold')),
    ('soffit_piece', ('soffit',)),
    ('haul_off_building', ('haul off', 'dump fee')),
    ('siding_sq', ('woodsman', 'vinyl siding', 'sididng', 'siding')),
]

VALID_KEYS = {entry[0] for entry in DETAIL_LINES} | {entry[0] for entry in LIBRARY_LINES}


def _norm_header(h):
    return re.sub(r'[^a-z0-9]+', '_', (h or '').strip().lower()).strip('_')


def _build_label_map():
    mapping = {}
    for key, label, *_ in DETAIL_LINES:
        for variant in (label, label.replace(' White', ''), label.replace(' (per sq)', '')):
            mapping[variant.strip().lower()] = key
            mapping[_norm_header(variant)] = key
    for key, label, *_ in LIBRARY_LINES:
        for variant in (label, label.replace(' (per sq)', '')):
            mapping[variant.strip().lower()] = key
            mapping[_norm_header(variant)] = key
    for alias, key in KEY_ALIASES.items():
        mapping[alias] = key
        mapping[_norm_header(alias)] = key
    for key in VALID_KEYS:
        mapping[key] = key
        mapping[_norm_header(key)] = key
    return mapping


LABEL_MAP = _build_label_map()


def _parse_price(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace('$', '').replace(' ', '')
    if not s:
        return None
    # European decimal: 88,75
    if ',' in s and '.' not in s:
        s = s.replace(',', '.')
    else:
        s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None


def _clean_item_text(val):
    return re.sub(r'\s+', ' ', (val or '').strip())


def _resolve_key(raw_key, raw_name):
    candidates = []
    for val in (raw_key, raw_name, f'{raw_key} {raw_name}'.strip()):
        val = _clean_item_text(val)
        if val and val not in candidates:
            candidates.append(val)

    for candidate in candidates:
        c = candidate.lower()
        if c in LABEL_MAP:
            return LABEL_MAP[c]
        norm = _norm_header(c)
        if norm in LABEL_MAP:
            return LABEL_MAP[norm]

    combined = ' '.join(candidates).lower()
    for key, hints in FUZZY_HINTS:
        for hint in hints:
            if hint in combined:
                return key

    # Legacy template: raw key column may already be an internal key
    if raw_key and raw_key.strip() in VALID_KEYS:
        return raw_key.strip()

    return None


def _decode_text(raw):
    for encoding in ('utf-8-sig', 'utf-16', 'utf-16-le', 'latin-1'):
        try:
            return raw.decode(encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return raw.decode('utf-8', errors='replace')


def _dict_rows_from_text(text):
    """Parse delimited text; tries comma, tab, and semicolon."""
    if not text.strip():
        return [], []

    for delimiter in (',', '\t', ';', '|'):
        try:
            reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            if not reader.fieldnames or len(reader.fieldnames) < 2:
                continue
            rows = list(reader)
            if rows:
                return reader.fieldnames, rows
            # header-only file
            return reader.fieldnames, []
        except csv.Error:
            continue

    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=',\t;|')
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        return reader.fieldnames or [], list(reader)
    except csv.Error:
        return [], []


def _rows_from_xlsx(raw):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    grid = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None and str(cell).strip() for cell in row):
            grid.append(list(row))
    wb.close()
    if len(grid) < 2:
        return [], []

    header_idx = 0
    for i, row in enumerate(grid[:15]):
        norms = {_norm_header(str(c)) for c in row if c is not None}
        if norms & {'unit_price', 'item_price', 'price', 'cost', 'item', 'item_name', 'description', 'key'}:
            header_idx = i
            break

    headers = [str(c).strip() if c is not None else '' for c in grid[header_idx]]
    rows = []
    for row in grid[header_idx + 1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        record = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            record[h] = row[i] if i < len(row) else None
        rows.append(record)
    return headers, rows


def _column_map(fieldnames):
    fields = {_norm_header(h): h for h in fieldnames if h}
    key_col = (
        fields.get('key')
        or fields.get('item_key')
        or fields.get('item_name')
        or fields.get('item')
        or fields.get('description')
    )
    name_col = fields.get('description') or fields.get('item_name') or fields.get('item') or key_col
    price_col = (
        fields.get('unit_price')
        or fields.get('item_price')
        or fields.get('price')
        or fields.get('cost')
    )
    factor_col = (
        fields.get('qty_per_sq')
        or fields.get('quantity_per_sq')
        or fields.get('qty_per_square')
    )
    return key_col, name_col, price_col, factor_col


def _ingest_rows(rows, key_col, name_col, price_col, factor_col, result, start_row=2):
    for offset, row in enumerate(rows):
        i = start_row + offset
        if not row:
            continue

        raw_key = _clean_item_text(row.get(key_col, '') if key_col else '')
        if name_col and name_col != key_col:
            raw_name = _clean_item_text(row.get(name_col, ''))
        elif key_col:
            raw_name = raw_key
        else:
            raw_name = _clean_item_text(row.get(name_col, '') if name_col else '')

        if not raw_name and not raw_key:
            continue

        price = _parse_price(row.get(price_col) if price_col else None)
        factor_raw = row.get(factor_col) if factor_col else None
        factor = _parse_price(factor_raw)

        # Only treat as per-square library when qty_per_sq column exists and has a value
        if factor_col and factor is not None and raw_name:
            result['library'].append({
                'name': raw_name,
                'qty_per_sq': factor,
                'unit_price': price,
            })
            lib_key = _resolve_key(raw_key, raw_name)
            if lib_key and price is not None and lib_key not in result['prices']:
                result['prices'][lib_key] = price
            if price is not None:
                result['loaded_count'] += 1
            continue

        key = _resolve_key(raw_key, raw_name)
        if not key:
            if raw_name or raw_key:
                result['warnings'].append(
                    f'Row {i}: could not match item "{raw_name or raw_key}" — skipped.'
                )
            continue
        if price is None:
            result['warnings'].append(f'Row {i}: "{raw_name or key}" has no price — skipped.')
            continue
        result['prices'][key] = price
        result['loaded_count'] += 1


def parse_pricing_upload(file_storage):
    """
    Parse uploaded pricing CSV or Excel (.xlsx).
    Returns dict with keys: prices (key->float), library (optional per-sq rows), warnings, loaded_count.
    """
    result = {'prices': {}, 'library': [], 'warnings': [], 'loaded_count': 0}
    if not file_storage:
        return result

    raw = file_storage.read()
    if hasattr(file_storage, 'seek'):
        file_storage.seek(0)

    if not raw:
        result['warnings'].append('Pricing file was empty.')
        return result

    if raw[:2] == b'PK':
        fieldnames, rows = _rows_from_xlsx(raw)
        if not fieldnames:
            result['warnings'].append('Could not read pricing from Excel file — check the first sheet.')
            return result
        key_col, name_col, price_col, factor_col = _column_map(fieldnames)
        if not price_col:
            result['warnings'].append('No price column found (use Item Price or unit_price).')
            return result
        _ingest_rows(rows, key_col, name_col, price_col, factor_col, result, start_row=2)
    else:
        text = _decode_text(raw)
        if not text.strip():
            result['warnings'].append('Pricing file was empty.')
            return result
        fieldnames, rows = _dict_rows_from_text(text)
        if not fieldnames:
            result['warnings'].append('Could not read column headers — save as CSV or upload .xlsx.')
            return result
        key_col, name_col, price_col, factor_col = _column_map(fieldnames)
        if not price_col:
            result['warnings'].append('No price column found (use unit_price or Item Price).')
            return result
        _ingest_rows(rows, key_col, name_col, price_col, factor_col, result, start_row=2)

    if result['loaded_count'] == 0 and not result['warnings']:
        result['warnings'].append(
            'No prices were loaded — use Item + Item Price columns, or download the hub pricing template.'
        )
    return result