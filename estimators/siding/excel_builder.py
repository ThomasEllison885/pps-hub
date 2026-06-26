"""PPS Siding Estimator — worksheet aligned with PPS estimator template."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .calculator import calculate_quantities
from .material_catalog import (
    TAKEOFF_LINES, LIBRARY_LINES, DETAIL_LINES, measurement_values,
)

DARK_BLUE = '004C8C'
BLUE = '0096D6'
LIGHT_BLUE = 'EBF6FC'
WHITE = 'FFFFFF'
GRAY_HDR = 'F2F7FB'
WARNING = 'FFF3CD'

SOURCE_LABELS = {
    'eagleview': 'EagleView',
    'aerial_other': 'Other Aerial Report',
    'field': 'Field Measurements',
}

# Fixed assumption cells on Tab 1
SUMMARY_WASTE_ROW = 14
SUMMARY_LABOR_RATE_ROW = 21
SUMMARY_HAUL_RATE_ROW = 22
SUMMARY_TAX_PCT_ROW = 23
SUMMARY_DELIVERY_ROW = 24
SUMMARY_WASTE_CELL = f"'1 – Job Summary'!$C${SUMMARY_WASTE_ROW}"
SUMMARY_LABOR_CELL = f"'1 – Job Summary'!$C${SUMMARY_LABOR_RATE_ROW}"
SUMMARY_HAUL_CELL = f"'1 – Job Summary'!$C${SUMMARY_HAUL_RATE_ROW}"
SUMMARY_TAX_CELL = f"'1 – Job Summary'!$C${SUMMARY_TAX_PCT_ROW}"
SUMMARY_DELIVERY_CELL = f"'1 – Job Summary'!$C${SUMMARY_DELIVERY_ROW}"

SHEET_SUMMARY = '1 – Job Summary'
SHEET_TAKEOFF = '2 – Takeoff'
SHEET_LIBRARY = '3 – Material Library'
SHEET_MATERIALS = '4 – Materials'
SHEET_LABOR = '5 – Labor'
SHEET_TOTAL = '6 – Estimate Total'


def build_estimate_excel(job, buildings, inputs, pricing, library_rows=None):
    pricing = pricing or {}
    library_rows = library_rows or []

    building_results = []
    for b in buildings:
        qty = max(int(b.get('qty') or 1), 1)
        q = calculate_quantities(b.get('measurements') or {}, inputs, qty=qty)
        building_results.append({
            'label': b.get('label') or 'Building',
            'building_type': b.get('building_type') or 'Building',
            'qty': qty,
            'source': b.get('source') or 'field',
            'measurements': b.get('measurements') or {},
            'quantities': q,
            'takeoff': measurement_values(b.get('measurements') or {}, q),
        })

    wb = Workbook()
    wb.remove(wb.active)

    _build_summary(wb, job, building_results, inputs)
    takeoff_refs = _build_takeoff(wb, building_results)
    library_total_row = _build_library(wb, pricing, library_rows)
    mat_subtotal_row = _build_materials(wb, building_results, pricing, takeoff_refs, inputs)
    labor_subtotal_row = _build_labor(wb, building_results)
    _build_totals(wb, mat_subtotal_row, labor_subtotal_row, library_total_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _thin_border():
    s = Side(style='thin', color='D0DCE8')
    return Border(left=s, right=s, top=s, bottom=s)


def _section_head(ws, row, col, text, width=6):
    from openpyxl.utils import get_column_letter
    end = get_column_letter(col + width - 1)
    start = get_column_letter(col)
    ws.merge_cells(f'{start}{row}:{end}{row}')
    c = ws[f'{start}{row}']
    c.value = text
    c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
    c.fill = PatternFill('solid', start_color=DARK_BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 20


def _money_cols(ws, row, cols='GHI'):
    for col in cols:
        ws[f'{col}{row}'].number_format = '_($* #,##0.00_)'
        ws[f'{col}{row}'].alignment = Alignment(horizontal='right')


def _build_summary(wb, job, building_results, inputs):
    ws = wb.create_sheet(SHEET_SUMMARY)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    for col, w in [('B', 24), ('C', 16), ('D', 14), ('E', 14), ('F', 14), ('G', 12)]:
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:G2')
    ws['B2'] = 'PURE PROPERTY SOLUTIONS — Siding Estimate'
    ws['B2'].font = Font(name='Arial', bold=True, size=14, color=WHITE)
    ws['B2'].fill = PatternFill('solid', start_color=DARK_BLUE)

    _section_head(ws, 4, 2, 'INVOICE SUMMARY (pulls from Tab 6)', 6)
    for col, txt in [('B', 'Item'), ('C', 'Cost'), ('D', 'Markup'), ('E', 'Overhead'), ('F', 'Invoice'), ('G', 'Margin %')]:
        ws[f'{col}6'] = txt
        ws[f'{col}6'].font = Font(name='Arial', bold=True, size=9, color=WHITE)
        ws[f'{col}6'].fill = PatternFill('solid', start_color=BLUE)
    ws['B7'] = 'Siding & Metal Work'
    ws['C7'] = f"='{SHEET_TOTAL}'!C9"
    ws['D7'] = f"='{SHEET_TOTAL}'!C11"
    ws['E7'] = f"='{SHEET_TOTAL}'!C12"
    ws['F7'] = f"='{SHEET_TOTAL}'!C13"
    ws['G7'] = '=IF(F7=0,"",D7/F7)'

    row = 9
    info = [
        ('Property:', job.get('property_name', '')),
        ('Address:', job.get('address', '')),
        ('Estimator:', job.get('estimator', '')),
        ('Date:', job.get('date', '')),
    ]
    for label, val in info:
        ws[f'B{row}'] = label
        ws[f'B{row}'].font = Font(bold=True, size=10, color=DARK_BLUE)
        ws[f'C{row}'] = val
        row += 1

    _section_head(ws, 11, 2, 'JOB ASSUMPTIONS — yellow cells drive formulas', 6)
    assumptions = [
        (12, 'Siding Type', inputs.get('siding_type', 'Vinyl Lap'), None),
        (13, 'Vinyl Exposure', f"{inputs.get('exposure_in', 'N/A')}\"" if 'vinyl' in inputs.get('siding_type', '').lower() else 'N/A', None),
        (14, 'Waste Factor (%)', inputs.get('waste_pct', 14), '0'),
        (15, 'Corner Post Length', f"{inputs.get('post_length', 12)} ft", None),
        (16, 'Stories', str(inputs.get('stories', 2)), None),
        (17, 'House Wrap', 'Yes' if inputs.get('include_housewrap') else 'No', None),
        (18, 'Fan Fold', 'Yes' if inputs.get('include_fanfold') else 'No', None),
        (19, 'Soffit', 'Yes' if inputs.get('include_soffit') else 'No', None),
        (20, 'Fascia Metal Wrap', 'Yes' if inputs.get('include_fascia_wrap') else 'No', None),
        (21, 'Labor $ / sq (net)', inputs.get('labor_per_sq', 180), '_($* #,##0_)'),
        (22, 'Haul Off $ / sq', inputs.get('haul_per_sq', 25), '_($* #,##0_)'),
        (23, 'Material Tax (%)', inputs.get('tax_pct', 7), '0'),
        (24, 'Delivery ($)', inputs.get('delivery', 15), '_($* #,##0_)'),
    ]
    border = _thin_border()
    for rr, lbl, val, fmt in assumptions:
        ws[f'B{rr}'] = lbl
        ws[f'C{rr}'] = val
        for col in ('B', 'C'):
            ws[f'{col}{rr}'].font = Font(name='Arial', size=10, bold=(col == 'B'))
            ws[f'{col}{rr}'].fill = PatternFill('solid', start_color=GRAY_HDR if rr % 2 == 0 else WHITE)
            ws[f'{col}{rr}'].border = border
        if fmt:
            ws[f'C{rr}'].number_format = fmt
            ws[f'C{rr}'].fill = PatternFill('solid', start_color=WARNING)

    row = 27
    _section_head(ws, row, 2, 'BUILDINGS ON THIS JOB', 6)
    row += 1
    for col, txt in [('B', 'Building'), ('C', 'Type'), ('D', 'Qty'), ('E', 'Source'), ('F', 'Net Sq'), ('G', 'Order Sq')]:
        ws[f'{col}{row}'] = txt
        ws[f'{col}{row}'].font = Font(bold=True, size=9, color=DARK_BLUE)
        ws[f'{col}{row}'].fill = PatternFill('solid', start_color=LIGHT_BLUE)
    row += 1
    for i, b in enumerate(building_results):
        q = b['quantities']
        vals = [b['label'], b['building_type'], b['qty'], SOURCE_LABELS.get(b['source'], b['source']),
                q['siding_squares_net'], q['siding_squares']]
        for j, col in enumerate('BCDEFG'):
            ws[f'{col}{row}'] = vals[j]
            ws[f'{col}{row}'].fill = PatternFill('solid', start_color=GRAY_HDR if i % 2 else WHITE)
            ws[f'{col}{row}'].border = border
        row += 1


def _build_takeoff(wb, building_results):
    """Returns {building_index: {takeoff_key: row_number}} referencing column D (with waste)."""
    ws = wb.create_sheet(SHEET_TAKEOFF)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    refs = {}
    row = 2
    border = _thin_border()

    for bi, b in enumerate(building_results):
        _section_head(ws, row, 2, f"{b['label'].upper()} — {b['building_type']} TAKEOFF", 5)
        row += 1
        for col, txt in [('B', 'Item'), ('C', 'Qty'), ('D', 'With Waste'), ('E', 'Used For')]:
            ws[f'{col}{row}'] = txt
            ws[f'{col}{row}'].font = Font(bold=True, size=9, color=WHITE)
            ws[f'{col}{row}'].fill = PatternFill('solid', start_color=BLUE)
        row += 1
        refs[bi] = {}
        takeoff = b['takeoff']
        for key, label, field in TAKEOFF_LINES:
            qty = takeoff.get(field) or 0
            ws[f'B{row}'] = label
            ws[f'C{row}'] = qty
            ws[f'D{row}'] = f'=C{row}*(1+{SUMMARY_WASTE_CELL}/100)'
            ws[f'E{row}'] = ''
            refs[bi][key] = row
            for col in 'BCDE':
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].font = Font(name='Arial', size=10)
            row += 1
        row += 2
    return refs


def _build_library(wb, pricing, library_rows):
    ws = wb.create_sheet(SHEET_LIBRARY)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.merge_cells('B2:F2')
    ws['B2'] = 'MATERIAL PRICE LIBRARY (per square) — optional deep layer'
    ws['B2'].font = Font(name='Arial', bold=True, size=12, color=WHITE)
    ws['B2'].fill = PatternFill('solid', start_color=DARK_BLUE)
    ws.merge_cells('B3:F3')
    ws['B3'] = 'Edit unit prices here to build a blended $/sq rate. Tab 4 can use these prices.'
    ws['B3'].font = Font(name='Arial', size=9, italic=True, color='666666')

    row = 5
    for col, txt in [('B', 'Item'), ('C', 'Qty per Sq'), ('D', 'Unit Price'), ('E', 'Extended'), ('F', 'Notes')]:
        ws[f'{col}{row}'] = txt
        ws[f'{col}{row}'].font = Font(bold=True, size=9, color=WHITE)
        ws[f'{col}{row}'].fill = PatternFill('solid', start_color=BLUE)
    row += 1
    first_data = row
    border = _thin_border()

    uploaded = {r['name'].lower(): r for r in library_rows} if library_rows else {}

    for key, label, factor in LIBRARY_LINES:
        ws[f'B{row}'] = label
        up = uploaded.get(label.lower())
        ws[f'C{row}'] = up['qty_per_sq'] if up else factor
        price = pricing.get(key)
        if price is None and up and up.get('unit_price') is not None:
            price = up['unit_price']
        ws[f'D{row}'] = price if price else ''
        ws[f'E{row}'] = f'=IF(D{row}="","",C{row}*D{row})'
        ws[f'F{row}'] = key
        for col in 'BCDEF':
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].font = Font(name='Arial', size=10)
            if col == 'D' and not price:
                ws[f'D{row}'].fill = PatternFill('solid', start_color=WARNING)
        _money_cols(ws, row, 'DE')
        row += 1

    total_row = row + 1
    ws[f'B{total_row}'] = 'Material $ / Sq (library total)'
    ws[f'B{total_row}'].font = Font(bold=True, color=DARK_BLUE)
    ws[f'E{total_row}'] = f'=SUM(E{first_data}:E{row - 1})'
    ws[f'E{total_row}'].font = Font(bold=True)
    _money_cols(ws, total_row, 'E')
    return total_row


def _qty_formula(takeoff_refs, bi, takeoff_key, divisor, fixed_qty, count_mult, row):
    """Build qty formula for a material line."""
    if fixed_qty is not None:
        return fixed_qty
    if takeoff_key == 'housewrap_rolls':
        wall_row = takeoff_refs[bi].get('wall_area')
        if wall_row:
            return f"=ROUND('{SHEET_TAKEOFF}'!D{wall_row}/{divisor},2)"
        return 0
    if takeoff_key == 'siding_sq_order':
        wall_row = takeoff_refs[bi].get('wall_area')
        if wall_row:
            return f"=ROUND('{SHEET_TAKEOFF}'!D{wall_row}/100,2)"
        return 0
    if takeoff_key == 'jchannel_total':
        # sum two takeoff rows with waste
        wd = takeoff_refs[bi].get('window_door_perimeter')
        top = takeoff_refs[bi].get('top_walls')
        if wd and top:
            return f"=ROUND(('{SHEET_TAKEOFF}'!D{wd}+'{SHEET_TAKEOFF}'!D{top})/{divisor},2)"
        return 0
    trow = takeoff_refs[bi].get(takeoff_key)
    if not trow:
        return 0
    base = f"=ROUND('{SHEET_TAKEOFF}'!D{trow}/{divisor},2)" if divisor else f"='{SHEET_TAKEOFF}'!D{trow}"
    if count_mult:
        return f'=ROUND(({base[1:]})*{count_mult},2)' if base.startswith('=') else base * count_mult
    return base


def _material_lines_for_inputs(inputs):
    """Filter detail lines based on job accessory toggles."""
    inputs = inputs or {}
    lines = []
    for entry in DETAIL_LINES:
        key = entry[0]
        if key in ('housewrap_roll', 'housewrap_tape') and not inputs.get('include_housewrap'):
            continue
        if key == 'fanfold_sq' and not inputs.get('include_fanfold'):
            continue
        if key == 'fascia_piece' and not inputs.get('include_fascia_wrap'):
            continue
        lines.append(entry)
    return lines


def _build_materials(wb, building_results, pricing, takeoff_refs, inputs):
    ws = wb.create_sheet(SHEET_MATERIALS)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    for col, w in [('B', 34), ('C', 12), ('D', 14), ('E', 14), ('F', 8)]:
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:E2')
    ws['B2'] = 'MATERIALS BY BUILDING — Item | Qty | Unit Price | Total'
    ws['B2'].font = Font(name='Arial', bold=True, size=12, color=WHITE)
    ws['B2'].fill = PatternFill('solid', start_color=DARK_BLUE)

    row = 4
    price_rows = []
    border = _thin_border()
    building_total_rows = []

    for bi, b in enumerate(building_results):
        _section_head(ws, row, 2, f"{b['label'].upper()} — {b['building_type']}", 4)
        row += 1
        for col, txt in [('B', 'Item'), ('C', 'Qty'), ('D', 'Unit Price'), ('E', 'Total')]:
            ws[f'{col}{row}'] = txt
            ws[f'{col}{row}'].font = Font(bold=True, size=9, color=WHITE)
            ws[f'{col}{row}'].fill = PatternFill('solid', start_color=BLUE)
        row += 1
        first_line = row

        for key, label, tkey, divisor, unit, fixed_qty, count_mult in _material_lines_for_inputs(inputs):
            ws[f'B{row}'] = label
            ws[f'F{row}'] = unit
            qty = _qty_formula(takeoff_refs, bi, tkey, divisor, fixed_qty, count_mult, row)
            ws[f'C{row}'] = qty
            price = pricing.get(key)
            if key == 'haul_off_building' and price is None:
                price = f'={SUMMARY_HAUL_CELL}'
            ws[f'D{row}'] = price if price is not None and price != '' else ''
            ws[f'E{row}'] = f'=IF(D{row}="","",C{row}*D{row})'
            price_rows.append(row)
            for col in 'BCDE':
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].font = Font(name='Arial', size=10)
            if (price is None or price == '') and key != 'haul_off_building':
                ws[f'D{row}'].fill = PatternFill('solid', start_color=WARNING)
            elif key == 'haul_off_building' and not pricing.get(key):
                ws[f'D{row}'].fill = PatternFill('solid', start_color=GRAY_HDR)
            _money_cols(ws, row, 'DE')
            row += 1

        total_row = row
        ws[f'B{total_row}'] = 'Building Material Total'
        ws[f'B{total_row}'].font = Font(bold=True)
        ws[f'E{total_row}'] = f'=SUM(E{first_line}:E{total_row - 1})'
        _money_cols(ws, total_row, 'E')
        row += 1

        ws[f'B{row}'] = 'Building Type Count'
        ws[f'C{row}'] = b['qty']
        row += 1
        expanded_row = row
        ws[f'B{row}'] = 'Expanded Material Cost'
        ws[f'B{row}'].font = Font(bold=True, color=DARK_BLUE)
        ws[f'E{row}'] = f'=E{total_row}*C{row - 1}'
        building_total_rows.append(row)
        _money_cols(ws, row, 'E')
        row += 2

    subtotal_row = row
    ws[f'B{subtotal_row}'] = 'JOB MATERIAL SUBTOTAL'
    ws[f'B{subtotal_row}'].font = Font(bold=True, size=11, color=WHITE)
    ws[f'B{subtotal_row}'].fill = PatternFill('solid', start_color=DARK_BLUE)
    refs = ','.join(f'E{r}' for r in building_total_rows)
    ws[f'E{subtotal_row}'] = f'=SUM({refs})'
    ws[f'E{subtotal_row}'].font = Font(bold=True, size=11, color=WHITE)
    ws[f'E{subtotal_row}'].fill = PatternFill('solid', start_color=DARK_BLUE)
    _money_cols(ws, subtotal_row, 'E')

    note = subtotal_row + 2
    ws.merge_cells(f'B{note}:E{note}')
    ws[f'B{note}'] = (
        'Yellow cells = enter unit price. Totals = Qty × Unit Price. '
        f'Waste % on Tab 1 (C{SUMMARY_WASTE_ROW}) flows through Tab 2 takeoff.'
    )
    ws[f'B{note}'].font = Font(size=9, italic=True, color='7A5C00')
    ws[f'B{note}'].fill = PatternFill('solid', start_color=WARNING)
    return subtotal_row


def _build_labor(wb, building_results):
    ws = wb.create_sheet(SHEET_LABOR)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2

    ws.merge_cells('B2:F2')
    ws['B2'] = 'SIDING & METAL WORK LABOR'
    ws['B2'].font = Font(name='Arial', bold=True, size=12, color=WHITE)
    ws['B2'].fill = PatternFill('solid', start_color=DARK_BLUE)
    ws['B3'] = 'Labor $/sq (net — no material waste):'
    ws['C3'] = f'={SUMMARY_LABOR_CELL}'
    ws['C3'].number_format = '_($* #,##0_)'

    row = 5
    for col, txt in [('B', 'Building'), ('C', 'Type'), ('D', 'Qty'), ('E', 'Net Sq'), ('F', 'Expanded')]:
        ws[f'{col}{row}'] = txt
        ws[f'{col}{row}'].font = Font(bold=True, size=9, color=WHITE)
        ws[f'{col}{row}'].fill = PatternFill('solid', start_color=BLUE)
    row += 1
    first = row
    border = _thin_border()

    for b in building_results:
        q = b['quantities']
        net_sq = q['siding_squares_net']
        ws[f'B{row}'] = b['label']
        ws[f'C{row}'] = b['building_type']
        ws[f'D{row}'] = b['qty']
        ws[f'E{row}'] = net_sq
        ws[f'F{row}'] = f'={SUMMARY_LABOR_CELL}*E{row}*D{row}'
        for col in 'BCDEF':
            ws[f'{col}{row}'].border = border
        _money_cols(ws, row, 'F')
        row += 1

    labor_total_row = row + 1
    ws[f'B{labor_total_row}'] = 'Labor Total'
    ws[f'B{labor_total_row}'].font = Font(bold=True)
    ws[f'F{labor_total_row}'] = f'=SUM(F{first}:F{row - 1})'
    _money_cols(ws, labor_total_row, 'F')

    haul_row = labor_total_row + 2
    ws[f'B{haul_row}'] = 'Haul Off & Dump (job net squares × $/sq)'
    total_net = round(sum(b['quantities']['siding_squares_net'] for b in building_results), 2)
    ws[f'C{haul_row}'] = total_net
    ws[f'D{haul_row}'] = f'={SUMMARY_HAUL_CELL}'
    ws[f'E{haul_row}'] = f'=C{haul_row}*D{haul_row}'
    _money_cols(ws, haul_row, 'DE')

    subtotal_row = haul_row + 2
    ws[f'B{subtotal_row}'] = 'LABOR + HAUL SUBTOTAL'
    ws[f'B{subtotal_row}'].font = Font(bold=True, color=WHITE)
    ws[f'B{subtotal_row}'].fill = PatternFill('solid', start_color=DARK_BLUE)
    ws[f'F{subtotal_row}'] = f'=F{labor_total_row}+E{haul_row}'
    ws[f'F{subtotal_row}'].font = Font(bold=True, color=WHITE)
    ws[f'F{subtotal_row}'].fill = PatternFill('solid', start_color=DARK_BLUE)
    _money_cols(ws, subtotal_row, 'F')
    return subtotal_row


def _build_totals(wb, mat_subtotal_row, labor_subtotal_row, library_total_row):
    ws = wb.create_sheet(SHEET_TOTAL)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 18
    border = _thin_border()

    ws.merge_cells('B2:C2')
    ws['B2'] = 'ESTIMATE TOTAL'
    ws['B2'].font = Font(name='Arial', bold=True, size=14, color=WHITE)
    ws['B2'].fill = PatternFill('solid', start_color=DARK_BLUE)

    def line(r, label, formula, bg=GRAY_HDR, bold=False):
        ws[f'B{r}'] = label
        ws[f'C{r}'] = formula
        for col in 'BC':
            ws[f'{col}{r}'].font = Font(bold=bold, color=WHITE if bg == DARK_BLUE else '000000')
            ws[f'{col}{r}'].fill = PatternFill('solid', start_color=bg)
            ws[f'{col}{r}'].border = border
        ws[f'C{r}'].number_format = '_($* #,##0.00_)'
        ws[f'C{r}'].alignment = Alignment(horizontal='right')

    line(4, 'Material Subtotal (Tab 4)', f"='{SHEET_MATERIALS}'!E{mat_subtotal_row}")
    line(5, 'Material Tax', f"=C4*({SUMMARY_TAX_CELL}/100)")
    line(6, 'Delivery', f'={SUMMARY_DELIVERY_CELL}')
    line(7, 'Material Grand Total', '=C4+C5+C6')
    line(8, 'Labor + Haul Subtotal (Tab 5)', f"='{SHEET_LABOR}'!F{labor_subtotal_row}")
    line(9, 'Combined Cost', '=C7+C8', DARK_BLUE, bold=True)
    line(11, 'Markup ($)', '', WARNING)
    ws['C11'] = 0
    line(12, 'Overhead ($)', '', WARNING)
    ws['C12'] = 0
    line(13, 'Invoice Amount', '=C9+C11+C12', DARK_BLUE, bold=True)
    line(15, 'Material $/Sq (Library Tab 3)', f"='{SHEET_LIBRARY}'!E{library_total_row}")
    line(16, 'Margin % (Markup / Invoice)', '=IF(C13=0,"",C11/C13)', GRAY_HDR)
    ws['C16'].number_format = '0.0%'

    line(18, 'FINAL ESTIMATE', '=C13', DARK_BLUE, bold=True)

    ws.merge_cells('B20:C20')
    ws['B20'] = (
        'Edit markup/overhead on this tab. Waste %, labor $/sq, haul $/sq, tax, and delivery on Tab 1. '
        'Optional per-square detail on Tab 3.'
    )
    ws['B20'].font = Font(size=9, italic=True, color='666666')