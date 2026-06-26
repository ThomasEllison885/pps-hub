"""PPS Gutter Estimator — Excel workbook."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .calculator import calculate_gutter_estimate

DARK_BLUE = '004C8C'
BLUE = '0096D6'
WARNING = 'FFF3CD'
WHITE = 'FFFFFF'
GRAY = 'F2F7FB'


def build_estimate_excel(job, measurements, inputs, confidence=None):
    calc = calculate_gutter_estimate(measurements, inputs)
    wb = Workbook()
    wb.remove(wb.active)

    bdr = Border(
        left=Side(style='thin', color='D0DCE8'),
        right=Side(style='thin', color='D0DCE8'),
        top=Side(style='thin', color='D0DCE8'),
        bottom=Side(style='thin', color='D0DCE8'),
    )

    ws1 = wb.create_sheet('1 – Job Summary')
    ws2 = wb.create_sheet('2 – Takeoff')
    ws3 = wb.create_sheet('3 – Estimate')
    ws4 = wb.create_sheet('4 – Bid Total')

    # Summary
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 16
    ws1.merge_cells('B2:C2')
    ws1['B2'] = 'PPS — Gutter & Downspout Estimate'
    ws1['B2'].font = Font(bold=True, size=14, color=WHITE)
    ws1['B2'].fill = PatternFill('solid', start_color=DARK_BLUE)

    r = 4
    for lbl, val in [
        ('Property:', job.get('property_name', '')),
        ('Address:', job.get('address', '')),
        ('Estimator:', job.get('estimator', '')),
        ('Date:', job.get('date', '')),
        ('Report #:', measurements.get('report_number', '')),
    ]:
        ws1[f'B{r}'] = lbl
        ws1[f'B{r}'].font = Font(bold=True, color=DARK_BLUE)
        ws1[f'C{r}'] = val
        r += 1

    if confidence:
        from estimators.reliability import reliability_excel_lines
        for lbl, val in reliability_excel_lines(confidence):
            ws1[f'B{r}'] = lbl
            ws1[f'B{r}'].font = Font(bold=True, color=DARK_BLUE, size=10)
            ws1[f'C{r}'] = val
            ws1[f'C{r}'].font = Font(size=10)
            r += 1
        r += 1

    assump_row = r
    ws1.merge_cells(f'B{assump_row}:C{assump_row}')
    ws1[f'B{assump_row}'] = 'ASSUMPTIONS'
    ws1[f'B{assump_row}'].font = Font(bold=True, color=WHITE)
    ws1[f'B{assump_row}'].fill = PatternFill('solid', start_color=DARK_BLUE)

    assumption_lines = [
        ('Material', 'K-Style Aluminum (5")', None),
        ('Waste / cuts (%)', inputs.get('waste_pct', 10), '0'),
        ('Gutter + downspout $/LF', inputs.get('gutter_price_per_lf', 7), '_($* #,##0.00_)'),
        ('Gutter guard $/LF', inputs.get('guard_price_per_lf', 2), '_($* #,##0.00_)'),
        ('Extra labor $/LF', inputs.get('labor_per_lf', 0), '_($* #,##0.00_)'),
        ('Tax (%)', inputs.get('tax_pct', 7.5), '0.0%'),
        ('Target margin (%)', inputs.get('margin_pct', 25), '0.0%'),
    ]
    rr = assump_row + 1
    tax_row = margin_row = rr
    for lbl, val, fmt in assumption_lines:
        ws1[f'B{rr}'] = lbl
        ws1[f'C{rr}'] = val
        ws1[f'B{rr}'].border = bdr
        ws1[f'C{rr}'].border = bdr
        if lbl == 'Tax (%)':
            tax_row = rr
        if lbl == 'Target margin (%)':
            margin_row = rr
        if fmt:
            ws1[f'C{rr}'].number_format = fmt
            ws1[f'C{rr}'].fill = PatternFill('solid', start_color=WARNING)
        rr += 1

    invoice_row = rr + 1
    ws1[f'B{invoice_row}'] = 'Invoice Total'
    ws1[f'B{invoice_row}'].font = Font(bold=True)
    ws1[f'C{invoice_row}'] = "='4 – Bid Total'!C8"
    ws1[f'C{invoice_row}'].number_format = '_($* #,##0.00_)'

    # Takeoff
    ws2.sheet_view.showGridLines = False
    ws2['B2'] = 'GUTTER TAKEOFF'
    ws2['B2'].font = Font(bold=True, color=DARK_BLUE)
    takeoff = [
        ('Gutter run (eaves)', calc['gutter_lf_raw'], 'ft', 'From report eaves or manual'),
        ('Downspout count', calc['downspout_count'], 'ea', f"~1 per {inputs.get('downspout_spacing_ft', 35)} LF"),
        ('Downspout vertical LF', calc['downspout_lf'], 'ft', f"{inputs.get('downspout_lf_each', 10)} ft each"),
        ('Total LF (gutter + DS)', calc['total_lf_raw'], 'ft', 'Before waste'),
        ('Order LF (with waste)', calc['total_lf_order'], 'ft', f"{calc['waste_pct']}% waste"),
    ]
    if calc['include_guards']:
        takeoff.append(('Guard coverage LF', calc['guard_lf'], 'ft', 'Gutter run + waste'))
    takeoff += [
        ('Hangers (est.)', calc['hangers'], 'ea', 'Every ~2.5 LF'),
        ('Elbows', calc['elbows'], 'ea', '2 per downspout'),
        ('End caps', calc['end_caps'], 'ea', 'Per run'),
    ]
    row = 4
    for col, h in [('B', 'Item'), ('C', 'Qty'), ('D', 'Unit'), ('E', 'Notes')]:
        ws2[f'{col}{row}'] = h
        ws2[f'{col}{row}'].font = Font(bold=True, color=WHITE)
        ws2[f'{col}{row}'].fill = PatternFill('solid', start_color=BLUE)
    row += 1
    for lbl, qty, unit, note in takeoff:
        ws2[f'B{row}'] = lbl
        ws2[f'C{row}'] = qty
        ws2[f'D{row}'] = unit
        ws2[f'E{row}'] = note
        row += 1

    # Line items
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['B'].width = 36
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 14
    ws3['B2'] = 'PRICING (yellow = editable)'
    ws3['B2'].font = Font(bold=True, color=DARK_BLUE)
    lines = [
        ('Gutter + downspout install', calc['total_lf_order'], inputs.get('gutter_price_per_lf', 7)),
    ]
    if calc['include_guards']:
        lines.append(('Gutter guards', calc['guard_lf'], inputs.get('guard_price_per_lf', 2)))
    if calc['labor_cost']:
        lines.append(('Additional labor', calc['total_lf_order'], inputs.get('labor_per_lf', 0)))

    row = 4
    for col, h in [('B', 'Item'), ('C', 'Qty'), ('D', 'Unit Price'), ('E', 'Extended')]:
        ws3[f'{col}{row}'] = h
        ws3[f'{col}{row}'].font = Font(bold=True, color=WHITE)
        ws3[f'{col}{row}'].fill = PatternFill('solid', start_color=BLUE)
    row += 1
    first = row
    for lbl, qty, price in lines:
        ws3[f'B{row}'] = lbl
        ws3[f'C{row}'] = qty
        ws3[f'D{row}'] = price
        ws3[f'E{row}'] = f'=C{row}*D{row}'
        ws3[f'D{row}'].fill = PatternFill('solid', start_color=WARNING)
        for c in 'BCDE':
            ws3[f'{c}{row}'].border = bdr
            ws3[f'{c}{row}'].number_format = '_($* #,##0.00_)' if c in 'DE' else 'General'
        row += 1
    sub_row = row
    ws3[f'B{sub_row}'] = 'Subtotal'
    ws3[f'B{sub_row}'].font = Font(bold=True)
    ws3[f'E{sub_row}'] = f'=SUM(E{first}:E{sub_row - 1})'
    ws3[f'E{sub_row}'].font = Font(bold=True)

    # Totals
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['B'].width = 30
    ws4.column_dimensions['C'].width = 16
    items = [
        (4, 'Line items subtotal', f"='3 – Estimate'!E{sub_row}"),
        (5, 'Tax', f"=C4*('1 – Job Summary'!C{tax_row}/100)"),
        (6, 'Cost before margin', '=C4+C5'),
        (8, 'Invoice (with margin)', f"=C6/(1-'1 – Job Summary'!C{margin_row}/100)"),
    ]
    for rr, lbl, formula in items:
        ws4[f'B{rr}'] = lbl
        ws4[f'C{rr}'] = formula
        ws4[f'C{rr}'].number_format = '_($* #,##0.00_)'
        if rr == 8:
            ws4[f'B{rr}'].font = Font(bold=True, color=WHITE)
            ws4[f'C{rr}'].font = Font(bold=True, color=WHITE)
            ws4[f'B{rr}'].fill = PatternFill('solid', start_color=DARK_BLUE)
            ws4[f'C{rr}'].fill = PatternFill('solid', start_color=DARK_BLUE)

    from estimators.excel_branding import brand_estimate_workbook
    brand_estimate_workbook(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf