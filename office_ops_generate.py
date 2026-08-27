"""Generate Monday Numbers Excel from QB exports + frozen 2026 goals template.

Owner decisions 2026-08-04:
  - Goals stay from Monthly Outlook template (update for 2027 later).
  - Sales = invoice send date × Amount; Sales Rep 50/50 on multi-rep.
  - Draws/downpayments count here (in QB); bonuses are separate (closed jobs).
  - No A/R-by-rep; company AR + Bridges breakdown when possible.
  - Output Excel + email path; Stephanie notes via past-due prompt.
"""

from __future__ import annotations

import io
import os
import re
from collections import defaultdict
from calendar import month_abbr
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from office_ops import _parse_sales_reps

TEMPLATE_PATH = Path(__file__).resolve().parent / 'static' / 'office_ops' / 'monthly_outlook_goals_template_2026.xlsx'

# Monthly Sales: label → Actual $ row
SALES_ACTUAL_ROWS = {
    'thomas ellison': 4,
    'thomas': 4,
    'tony cumella': 10,
    'tony': 10,
    'adam cupito': 16,
    'adam': 16,
    'andy potts': 22,
    'andy': 22,
    'rachel farler': 28,
    'rachel': 28,
}
SALES_SHARED_ROWS = {
    'thomas ellison': 34,
    'tony cumella': 35,
    'adam cupito': 36,
    'andy potts': 37,
    'rachel farler': 38,
}

FILL_POS = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')
FILL_NEG = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
FILL_POS_CF = PatternFill(start_color='FFB7E1CD', end_color='FFB7E1CD', fill_type='solid')
FILL_NEG_CF = PatternFill(start_color='FFF4C7C3', end_color='FFF4C7C3', fill_type='solid')
FONT_POS = Font(color='FF006100')
FONT_NEG = Font(color='FF9C0006')
FILL_YELLOW = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
FILL_HEADER = PatternFill(start_color='FF1A5276', end_color='FF1A5276', fill_type='solid')
FONT_HEADER = Font(color='FFFFFFFF', bold=True, name='Calibri', size=12)
FONT_BODY = Font(name='Calibri', size=11)
FONT_BOLD = Font(name='Calibri', size=11, bold=True)

MONTH_COLS = {1: 2, 2: 3, 3: 4, 4: 5, 5: 6, 6: 7, 7: 8, 8: 9, 9: 10, 10: 11, 11: 12, 12: 13}


def _parse_invoice_date(val):
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val
    from datetime import date as date_cls
    if isinstance(val, date_cls):
        return datetime(val.year, val.month, val.day)
    if isinstance(val, (int, float)) and 20000 < float(val) < 80000:
        try:
            from datetime import timedelta
            d = date_cls(1899, 12, 30) + timedelta(days=int(val))
            return datetime(d.year, d.month, d.day)
        except (OverflowError, ValueError):
            return None
    s = str(val).strip()
    for fmt in (
        '%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y', '%Y-%m-%d %H:%M:%S',
        '%B %d, %Y', '%b %d, %Y', '%m-%d-%Y',
    ):
        try:
            return datetime.strptime(s[:19] if '%Y-%m-%d %H:%M:%S' == fmt else s, fmt)
        except ValueError:
            continue
    return None


def aggregate_sales_from_invoice_list(invoice_list, year=2026):
    """invoice_list items: date, amount, sales_reps[], is_50_50_style.

    Returns:
      by_rep_month: {rep_full_name: {month: dollars}}
      by_rep_shared_month: multi-rep only share by rep/month
      team_month: {month: dollars}
      meta
    """
    by_rep = defaultdict(lambda: defaultdict(float))
    by_shared = defaultdict(lambda: defaultdict(float))
    team = defaultdict(float)
    count_by_month = defaultdict(int)
    count = 0
    split_count = 0
    for inv in invoice_list or []:
        dt = _parse_invoice_date(inv.get('date'))
        if not dt or dt.year != year:
            continue
        # Count full invoice amount for sales (draws count for this report)
        amt = float(inv.get('amount') or 0)
        if amt == 0:
            continue
        reps = inv.get('sales_reps') or _parse_sales_reps(inv.get('sales_rep_raw') or inv.get('salesman'))
        if not reps:
            reps = ['(unassigned)']
        share = amt / len(reps)
        m = dt.month
        team[m] += amt
        count_by_month[m] += 1
        count += 1
        multi = len(reps) >= 2
        if multi:
            split_count += 1
        for r in reps:
            by_rep[r][m] += share
            if multi:
                by_shared[r][m] += share
    return {
        'by_rep_month': {k: dict(v) for k, v in by_rep.items()},
        'by_shared_month': {k: dict(v) for k, v in by_shared.items()},
        'team_month': dict(team),
        'invoice_count_by_month': dict(count_by_month),
        'invoice_count': count,
        'split_invoice_count': split_count,
        'year': year,
    }


def _bridges_ar_breakdown(customers):
    """Split Bridges/BOPC into Pebble / Pine-Meadow / old.

    Owner correction 2026-08: ~$400k+ 'BOPC 2701 / 2025' is **Pebble**, not
    Pine/Meadow. Only explicit pine/meadow names go to Pine/Meadow.
    """
    buckets = {
        'Bridges — Pebble': 0.0,
        'Bridges — Pine/Meadow': 0.0,
        'Bridges — old / other': 0.0,
    }
    parts = {'Bridges — Pebble': [], 'Bridges — Pine/Meadow': [], 'Bridges — old / other': []}
    for c in customers or []:
        name = (c.get('customer') or '')
        low = name.lower()
        # Only Bridges/BOPC family (not unrelated "Meadows" apartments)
        is_bridges = any(
            x in low for x in ('bopc', 'bridges of pine', 'bridges / bopc', 'bridges of pine creek')
        ) or low.startswith('bridges')
        if not is_bridges and 'bopc' not in (c.get('parent') or '').lower():
            # parent field may be Bridges of Pine Creek
            if 'bridges' not in (c.get('parent') or '').lower() and 'bopc' not in (c.get('parent') or '').lower():
                continue
        # Skip parent total-for (double-count + "pine" false positive)
        if c.get('is_parent_total') or (
            not c.get('job') and low.strip() in ('bridges of pine creek', 'bridges / bopc')
        ):
            continue
        job = (c.get('job') or name.split('·')[-1]).strip().lower()
        total = float(c.get('total') or 0)
        # Pebble first (includes BOPC 2701 / 2025 — owner 2026-08)
        if (
            'pebble' in low
            or 'pebble' in job
            or '2701' in low
            or '2701' in job
            or 'phase 3' in job
            or 'phase 3' in low
        ):
            key = 'Bridges — Pebble'
        elif 'meadow' in job or (job.startswith('pine') and 'creek' not in job):
            key = 'Bridges — Pine/Meadow'
        else:
            key = 'Bridges — old / other'
        buckets[key] += total
        parts[key].append(name)
    lines = [
        {'label': k, 'total': v, 'includes': parts[k]}
        for k, v in buckets.items()
        if abs(v) > 0.5
    ]
    if not lines and any(parts.values()):
        lines = [{'label': 'Bridges / BOPC (rolled up)', 'total': sum(buckets.values()), 'includes': []}]
    return lines, [n for ns in parts.values() for n in ns]


def _fill_sales_sheet(ws, sales_agg):
    """Write Actual $ and Shared rows from invoice aggregation; keep Goal/Historial."""
    by_rep = sales_agg['by_rep_month']
    by_shared = sales_agg['by_shared_month']

    # Clear Shared first
    for row in SALES_SHARED_ROWS.values():
        for col in MONTH_COLS.values():
            cell = ws.cell(row, col)
            cell.value = 0
            cell.number_format = '"$"#,##0.00'
            cell.fill = FILL_YELLOW

    # Map rep names in agg to rows
    def row_for(rep_name):
        low = (rep_name or '').lower()
        if low in SALES_ACTUAL_ROWS:
            return SALES_ACTUAL_ROWS[low]
        for k, r in SALES_ACTUAL_ROWS.items():
            if k in low or low in k:
                return r
        return None

    # Zero all actual months then fill
    for rep, months in by_rep.items():
        r = row_for(rep)
        if not r:
            continue
        for m in range(1, 13):
            col = MONTH_COLS[m]
            cell = ws.cell(r, col)
            val = months.get(m, 0.0)
            cell.value = round(val, 2) if val else 0
            cell.number_format = '"$"#,##0.00'
            cell.fill = FILL_YELLOW
        # Full year col N = 14 if used
        if ws.cell(r, 14).value is not None or True:
            total = sum(months.get(m, 0) for m in range(1, 13))
            ws.cell(r, 14).value = round(total, 2)
            ws.cell(r, 14).number_format = '"$"#,##0.00'

    for rep, months in by_shared.items():
        r = SALES_SHARED_ROWS.get(rep) or SALES_SHARED_ROWS.get(
            next((k for k in SALES_SHARED_ROWS if k.split()[0].lower() in rep.lower()), None)
        )
        # direct map
        low = rep.lower()
        r = None
        for k, row in SALES_SHARED_ROWS.items():
            if k in low or low in k or k.split()[0] == low.split()[0]:
                r = row
                break
        if not r:
            continue
        for m, val in months.items():
            col = MONTH_COLS.get(m)
            if not col:
                continue
            cell = ws.cell(r, col)
            cell.value = round(val, 2)
            cell.number_format = '"$"#,##0.00'


def _fill_team_sheet(ws, sales_agg):
    """2026 Actual $ row 13 from team_month; leave 2025 and goals alone."""
    team = sales_agg['team_month']
    # 2026 block starts row 9; Actual $ is row 13
    for m in range(1, 13):
        col = MONTH_COLS[m]
        cell = ws.cell(13, col)
        val = team.get(m, 0.0)
        cell.value = round(val, 2) if val else None
        if val:
            cell.number_format = '"$"#,##0.00'
            cell.fill = FILL_YELLOW
    # Full year sum formula
    ws.cell(13, 14).value = '=SUM(B13:M13)'
    ws.cell(13, 14).number_format = '"$"#,##0.00'


def _note_parts(val):
    """Normalize notes_by_customer values: str or {note, overdue, total}."""
    if isinstance(val, dict):
        return (
            (val.get('note') or '').strip(),
            val.get('overdue'),
            val.get('total'),
        )
    return (str(val).strip() if val is not None else ''), None, None


def _format_money_compact(n):
    try:
        if n is None:
            return None
        return f'${float(n):,.0f}'
    except (TypeError, ValueError):
        return None


# ── "Unusual increases or decreases" ────────────────────────────────────────
#
# Thomas, 2026-08-27. Two thresholds, because either one alone gives a bad
# list. Dollars alone lets a $400k line drifting 2% outrank a $12k line that
# doubled. Percent alone fills the section with $40 accounts tripling. A line
# has to clear both to be worth his Monday morning.
#
# Both are tunable and both are stated in the output, because a filtered list
# that does not say what it filtered reads as "nothing else moved".
PL_MOVE_MIN_DOLLARS = 2500.0
PL_MOVE_MIN_PCT = 0.15
PL_MOVERS_LIMIT = 6  # per section, not overall
# A line that went from nothing to something (or the reverse) has no
# percentage — it is a new or discontinued account, which is worth saying out
# loud rather than dividing by zero over.
PL_APPEARED_MIN_DOLLARS = 2500.0

# Named, because the Insights sheet styles a heading by exact match against
# this list. When the builder had its own string literals, adding a section
# meant remembering to add it here too — and forgetting rendered it as body
# text, indistinguishable from the bullets above it.
SEC_SALES = 'SALES'
SEC_SALES_BY_REP = 'SALES BY REP (YTD; multi-rep invoices split 50/50)'
SEC_MARGIN = 'MARGIN & PROFIT'
SEC_MOVERS = 'UNUSUAL MOVES (P&L, year over year)'
SEC_AR = 'A/R (company total)'
SEC_PAST_DUE = 'PAST-DUE UPDATES'

INSIGHT_SECTIONS = (
    SEC_SALES, SEC_SALES_BY_REP, SEC_MARGIN, SEC_MOVERS, SEC_AR, SEC_PAST_DUE,
)


SECTION_HEADINGS = (
    ('income', 'Income lines'),
    ('cogs', 'Job costs'),
    ('expenses', 'Overhead'),
)


def _dollars(n):
    """$1,200 / -$1,200. A contra line like "Discounts given" is stored
    negative in QB, and "$-21,000" reads as a typo."""
    v = float(n or 0)
    return f'-${_money(abs(v))}' if v < 0 else f'${_money(v)}'


def _mover_sentence(m):
    """One line of plain English for one mover."""
    if m['kind'] == 'new':
        return f'{m["label"]}: new this year — {_dollars(m["ty"])} (nothing last year)'
    if m['kind'] == 'stopped':
        return (f'{m["label"]}: nothing this year — was {_dollars(m["py"])} '
                f'last year')
    arrow = 'up' if m['delta'] > 0 else 'down'
    return (f'{m["label"]}: {_dollars(m["py"])} → {_dollars(m["ty"])} — {arrow} '
            f'${_money(abs(m["delta"]))} ({_pct_delta(m["pct"])})')


def pl_movers(pl_summary, min_dollars=PL_MOVE_MIN_DOLLARS,
              min_pct=PL_MOVE_MIN_PCT, limit=PL_MOVERS_LIMIT):
    """P&L line items whose year-over-year move is both material and large.

    Returns dicts with `kind`: 'up', 'down', 'new' (no prior-year figure) or
    'stopped' (nothing this year). Sorted by dollars moved, biggest first —
    that is the order the question "what changed" is actually asked in.
    """
    movers = []
    for line in (pl_summary or {}).get('lines') or []:
        ty = float(line.get('ty') or 0.0)
        py = float(line.get('py') or 0.0)
        delta = ty - py
        if py == 0 and ty != 0:
            if abs(ty) >= PL_APPEARED_MIN_DOLLARS:
                movers.append({**line, 'ty': ty, 'py': py, 'delta': delta,
                               'pct': None, 'kind': 'new'})
            continue
        if ty == 0 and py != 0:
            if abs(py) >= PL_APPEARED_MIN_DOLLARS:
                movers.append({**line, 'ty': ty, 'py': py, 'delta': delta,
                               'pct': None, 'kind': 'stopped'})
            continue
        if py == 0:
            continue
        pct = delta / abs(py)
        if abs(delta) < min_dollars or abs(pct) < min_pct:
            continue
        movers.append({**line, 'ty': ty, 'py': py, 'delta': delta, 'pct': pct,
                       'kind': 'up' if delta > 0 else 'down'})
    movers.sort(key=lambda m: -abs(m['delta']))
    # The cap is PER SECTION, not overall. Income lines are an order of
    # magnitude bigger than overhead lines, so one global top-8 would be eight
    # revenue rows and the overhead question — the one Thomas asked — would
    # never appear.
    kept = []
    per_section = {}
    for m in movers:
        sec = m.get('section') or 'other'
        per_section[sec] = per_section.get(sec, 0) + 1
        if per_section[sec] <= limit:
            kept.append(m)
    return kept


def _build_insights(sales_agg, ar_summary, notes_by_customer=None, pl_summary=None):
    """Deeper sales / margin / profit / AR narrative for leadership."""
    lines = []
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    lines.append('Monday Numbers · Insights')
    lines.append(f'Generated {now}')
    lines.append('')
    lines.append(
        'Sales = invoiced $ by invoice date (includes draws/downpayments for this pack). '
        'Bonus math still waits for fully closed jobs — not changed here.'
    )
    lines.append('')

    team = sales_agg.get('team_month') or {}
    ytd = sum(team.values()) if team else 0.0
    months = sorted(team.keys()) if team else []
    latest = months[-1] if months else None
    prior = months[-2] if len(months) >= 2 else None

    lines.append(SEC_SALES)
    if team:
        lines.append(f'• Company YTD invoiced: ${_money(ytd)}')
        if latest:
            latest_amt = team[latest]
            lines.append(f'• Latest month ({month_abbr[latest]}): ${_money(latest_amt)}')
            if prior:
                prior_amt = team[prior]
                delta = latest_amt - prior_amt
                pace = (delta / prior_amt) if prior_amt else 0
                direction = 'up' if delta > 0 else 'down' if delta < 0 else 'flat'
                lines.append(
                    f'• vs prior month ({month_abbr[prior]} ${_money(prior_amt)}): '
                    f'{direction} ${_money(abs(delta))} ({_pct_delta(pace)})'
                )
            # Day-based linear run-rate (not /12 months): avoids overweighting
            # late-in-month vs early-in-month when "latest month" is partial.
            if ytd:
                sales_year = int(sales_agg.get('year') or datetime.now().year)
                year_start = date(sales_year, 1, 1)
                year_end = date(sales_year, 12, 31)
                days_in_year = (year_end - year_start).days + 1  # 365 or 366
                today = date.today()
                if today.year > sales_year:
                    as_of = year_end
                elif today.year < sales_year:
                    as_of = year_start
                else:
                    as_of = min(today, year_end)
                days_elapsed = max(1, (as_of - year_start).days + 1)
                if days_elapsed < days_in_year:
                    daily = ytd / days_elapsed
                    run_rate = daily * days_in_year
                    weeks_elapsed = days_elapsed / 7.0
                    lines.append(
                        f'• Simple full-year run-rate from YTD: ~${_money(run_rate)} '
                        f'(linear by day: ${_money(ytd)} over {days_elapsed} of '
                        f'{days_in_year} days ≈ ${_money(daily)}/day · '
                        f'~${_money(daily * 7)}/wk; not a forecast)'
                    )
                    gap_10m = 10000000 - ytd
                    if gap_10m > 0:
                        days_left = days_in_year - days_elapsed
                        need_day = gap_10m / days_left if days_left else gap_10m
                        need_wk = need_day * 7
                        lines.append(
                            f'• To hit $10M goal: ~${_money(need_day)}/day '
                            f'(~${_money(need_wk)}/wk) for remaining {days_left} day(s) '
                            f'(${_money(gap_10m)} still needed)'
                        )
                    else:
                        lines.append('• YTD already at/above $10M full-year goal.')
                elif days_elapsed >= days_in_year:
                    lines.append(
                        f'• Full-year invoiced (year complete): ${_money(ytd)}'
                    )
                    if ytd >= 10000000:
                        lines.append('• YTD already at/above $10M full-year goal.')
        lines.append(
            f'• Invoice volume: {sales_agg.get("invoice_count", 0)} · '
            f'50/50-style splits: {sales_agg.get("split_invoice_count", 0)}'
        )
        # Stephanie request: largest invoice month by $ and by # of invoices sent
        peak_dollar_m = None
        peak_count_m = None
        if team:
            peak_dollar_m = max(team.keys(), key=lambda m: team[m])
            lines.append(
                f'• Largest invoice month by $: {month_abbr[peak_dollar_m]} '
                f'(${_money(team[peak_dollar_m])})'
            )
        count_by_m = sales_agg.get('invoice_count_by_month') or {}
        if count_by_m:
            peak_count_m = max(count_by_m.keys(), key=lambda m: count_by_m[m])
            lines.append(
                f'• Largest invoice month by # of invoices: {month_abbr[peak_count_m]} '
                f'({count_by_m[peak_count_m]} invoices)'
            )
        if (
            peak_dollar_m is not None
            and peak_count_m is not None
            and peak_dollar_m != peak_count_m
        ):
            lines.append(
                f'• Note: peak $ month ({month_abbr[peak_dollar_m]}) differs from '
                f'peak volume month ({month_abbr[peak_count_m]}) — bigger average ticket vs more invoices.'
            )
    else:
        lines.append('• Upload Invoice List by Date (Sales Rep) to fill sales.')
    lines.append('')

    by_rep = sales_agg.get('by_rep_month') or {}
    if by_rep:
        lines.append(SEC_SALES_BY_REP)
        ranked = sorted(
            ((r, sum(m.values())) for r, m in by_rep.items()),
            key=lambda x: -x[1],
        )
        ranked = [(r, t) for r, t in ranked if r != '(unassigned)']
        total_rep = sum(t for _, t in ranked) or 1.0
        for r, tot in ranked:
            share = tot / total_rep
            lm = ''
            if latest:
                lm_amt = (by_rep.get(r) or {}).get(latest, 0)
                lm = f' · {month_abbr[latest]} ${_money(lm_amt)}'
            lines.append(f'• {r}: ${_money(tot)} ({share:.0%} of credited sales){lm}')
        if ranked:
            top_name, top_tot = ranked[0]
            lines.append(
                f'• Concentration: {top_name} is {top_tot / total_rep:.0%} of rep-credited YTD.'
            )
        lines.append('')

    lines.append(SEC_MARGIN)
    if pl_summary:
        inc = pl_summary.get('income_ty')
        inc_py = pl_summary.get('income_py')
        gp = pl_summary.get('gross_profit_ty')
        gp_py = pl_summary.get('gross_profit_py')
        ni = pl_summary.get('net_income_ty')
        ni_py = pl_summary.get('net_income_py')
        period = pl_summary.get('period_label') or 'YTD'
        lines.append(f'• P&L period: {period} (vs same period prior year)')
        if inc is not None:
            line = f'• Income (TY): ${_money(inc)}'
            if inc_py is not None:
                line += f' · PY ${_money(inc_py)} ({_pct_change(inc, inc_py)})'
            lines.append(line)
        if gp is not None and inc:
            gm = gp / inc if inc else 0
            line = f'• Gross profit: ${_money(gp)} · gross margin {gm:.1%}'
            if gp_py is not None:
                line += f' · PY ${_money(gp_py)} ({_pct_change(gp, gp_py)})'
            lines.append(line)
            if gp_py is not None and inc_py:
                gm_py = gp_py / inc_py if inc_py else 0
                gm_pts = (gm - gm_py) * 100
                lines.append(
                    f'• Gross margin vs PY: {gm:.1%} vs {gm_py:.1%} ({gm_pts:+.1f} pts)'
                )
        if ni is not None:
            nm = (ni / inc) if inc else 0
            line = f'• Net income: ${_money(ni)} · net margin {nm:.1%}'
            if ni_py is not None:
                line += f' · PY ${_money(ni_py)} ({_pct_change(ni, ni_py)})'
            lines.append(line)
            if ni_py is not None:
                if ni > ni_py:
                    lines.append(
                        '• Profit is ahead of last year YTD — still watch AR for cash conversion.'
                    )
                elif ni < ni_py:
                    lines.append(
                        '• Profit is behind last year YTD — focus on margin and volume, not only invoices.'
                    )
        if pl_summary.get('cogs_ty') is not None and inc:
            cogs_ratio = pl_summary['cogs_ty'] / inc if inc else 0
            line = f'• COGS as % of income: {cogs_ratio:.1%}'
            if pl_summary.get('cogs_py') is not None and inc_py:
                cogs_py_ratio = pl_summary['cogs_py'] / inc_py
                line += (f' · PY {cogs_py_ratio:.1%} '
                         f'({(cogs_ratio - cogs_py_ratio) * 100:+.1f} pts)')
            lines.append(line)
        if inc is not None and inc_py:
            # Operating expense = what is left between gross profit and net.
            opex = (gp - ni) if (gp is not None and ni is not None) else None
            opex_py = (gp_py - ni_py) if (
                gp_py is not None and ni_py is not None) else None
            if opex is not None:
                line = f'• Operating expense (gross profit − net income): ${_money(opex)}'
                if opex_py is not None:
                    line += f' · PY ${_money(opex_py)} ({_pct_change(opex, opex_py)})'
                    if inc:
                        line += (f' · {opex / inc:.1%} of income vs '
                                 f'{opex_py / inc_py:.1%}')
                lines.append(line)
        pl_lines = pl_summary.get('lines') or []
        if pl_lines:
            lines.append(
                f'• Read {len(pl_lines)} line item(s) from the P&L for the '
                f'year-over-year comparison below.'
            )
        withheld = pl_summary.get('withheld_comp_lines') or 0
        if withheld:
            lines.append(
                f'• {withheld} compensation line(s) are deliberately not in '
                f'this pack — payroll and owner comp stay out of Hub reports.'
            )
    else:
        lines.append(
            '• Upload P&L (compare this year vs last year) on Office Ops for margin & profit depth.'
        )
    lines.append('')

    movers = pl_movers(pl_summary) if pl_summary else []
    if pl_summary and (pl_summary.get('lines') or []):
        lines.append(SEC_MOVERS)
        if movers:
            # Grouped, because "Services is up 21%" is the sales story already
            # told above, while "Vehicle Repairs is up 171%" is a question for
            # Monday. Reading them in one flat list buries the second under
            # the first, which is always bigger in dollars.
            for section, heading in SECTION_HEADINGS:
                group = [m for m in movers if m.get('section') == section]
                if not group:
                    continue
                lines.append(f'• {heading}')
                for m in group:
                    lines.append('    – ' + _mover_sentence(m))
            ungrouped = [m for m in movers
                         if m.get('section') not in dict(SECTION_HEADINGS)]
            for m in ungrouped:
                lines.append('• ' + _mover_sentence(m))
            lines.append(
                f'• Shown when a line moved at least ${_money(PL_MOVE_MIN_DOLLARS)} '
                f'AND {PL_MOVE_MIN_PCT:.0%} year over year. Everything else held '
                f'steady enough not to ask about.'
            )
        else:
            lines.append(
                f'• Nothing moved more than ${_money(PL_MOVE_MIN_DOLLARS)} and '
                f'{PL_MOVE_MIN_PCT:.0%} year over year. That is the finding, not '
                f'a gap in the data.'
            )
        lines.append('')

    lines.append(SEC_AR)
    if ar_summary and ar_summary.get('grand_total'):
        g = ar_summary['grand_total']
        tot = float(g.get('total') or 0)
        cur = float(g.get('current') or 0)
        d91 = float(g.get('91_and_over') or 0)
        lines.append(f'• Total AR: ${_money(tot)}')
        lines.append(
            f'• Aging: Current ${_money(cur)} · 1–30 ${_money(g.get("1_30"))} · '
            f'31–60 ${_money(g.get("31_60"))} · 61–90 ${_money(g.get("61_90"))} · '
            f'91+ ${_money(d91)}'
        )
        overdue = tot - cur
        if tot:
            lines.append(f'• Mix: Current {cur / tot:.0%} · 91+ {d91 / tot:.0%} of total AR')
            lines.append(
                f'• Past due (anything out of Current): ${_money(overdue)} '
                f'({overdue / tot:.0%} of AR)'
            )
        # Operating AR is the number Stephanie and Thomas actually steer on —
        # it was in the AR draft and nowhere in this pack.
        op = ar_summary.get('operating_ex_bopc') or {}
        if op.get('total') is not None and op.get('total') != tot:
            op_tot = float(op.get('total') or 0)
            op_overdue = op_tot - float(op.get('current') or 0)
            lines.append(
                f'• Operating AR (excluding Bridges/BOPC): ${_money(op_tot)} '
                f'· past due ${_money(op_overdue)} '
                f'· 91+ ${_money(op.get("91_and_over"))}'
            )
        customers = (ar_summary.get('all_customers')
                     or ar_summary.get('top_customers_by_balance') or [])
        if customers:
            with_91 = [c for c in customers
                       if float(c.get('91_and_over') or 0) > 0]
            lines.append(
                f'• {len(customers)} customer(s) carrying a balance · '
                f'{len(with_91)} of them have money in 91+'
            )
            ranked = sorted(customers,
                            key=lambda c: -float(c.get('total') or 0))
            if tot and ranked:
                top5 = sum(float(c.get('total') or 0) for c in ranked[:5])
                biggest = ranked[0]
                lines.append(
                    f'• Concentration: top 5 customers are {top5 / tot:.0%} of AR '
                    f'· largest is {biggest.get("customer")} at '
                    f'${_money(biggest.get("total"))}'
                )
        chase = ar_summary.get('chase_list') or []
        if chase:
            lines.append('• Most overdue by weight (91+ counts heaviest):')
            for c in chase[:5]:
                tag = ' — legacy/BOPC, handle with care' if c.get(
                    'is_legacy_or_bopc') else ''
                lines.append(
                    f'    – {c.get("customer")}: ${_money(c.get("overdue"))} past due '
                    f'of ${_money(c.get("total"))} (91+ ${_money(c.get("91_and_over"))})'
                    f'{tag}'
                )
        bridges = ar_summary.get('bridges_lines') or []
        if bridges:
            lines.append('• Bridges breakdown:')
            for b in bridges:
                lines.append(f'    – {b["label"]}: ${_money(b["total"])}')
        elif ar_summary.get('bopc'):
            lines.append(
                f'• Bridges / BOPC (rolled up): ${_money(ar_summary["bopc"].get("total"))}'
            )
        if ytd and tot:
            lines.append(
                f'• AR / YTD invoices: {tot / ytd:.0%} '
                f'(higher = more cash still in receivables)'
            )
    else:
        lines.append('• Upload A/R Aging Summary to fill AR.')
    lines.append('')

    if notes_by_customer:
        lines.append(SEC_PAST_DUE)
        # Sort largest past-due $ first so importance is obvious
        sorted_notes = sorted(
            notes_by_customer.items(),
            key=lambda kv: -float((_note_parts(kv[1])[1] or 0) or 0),
        )
        for cust, val in sorted_notes:
            note, overdue, _total = _note_parts(val)
            if not note or not str(note).strip():
                continue
            amt = _format_money_compact(overdue)
            if amt:
                lines.append(f'• {cust} ({amt} past due): {note.strip()}')
            else:
                lines.append(f'• {cust}: {note.strip()}')
        lines.append('')

    lines.append('—')
    lines.append(
        'Thursday pack · goals from 2026 template · actuals from Invoice List · optional P&L YoY'
    )
    return '\n'.join(lines)


def _pct_delta(n):
    try:
        return f'{float(n):+.0%}'
    except (TypeError, ValueError):
        return 'n/a'


def _pct_change(ty, py):
    try:
        ty, py = float(ty), float(py)
        if py == 0:
            return 'n/a'
        return f'{(ty - py) / abs(py):+.0%} YoY'
    except (TypeError, ValueError):
        return 'n/a'


def _money(n):
    try:
        return f'{float(n or 0):,.0f}'
    except (TypeError, ValueError):
        return '0'


def generate_from_qb(invoice_list, ar_summary=None, notes_by_customer=None, pl_summary=None, year=2026, template_path=None):
    """Build Monday Excel from Invoice List + AR + optional P&L YoY + notes."""
    path = Path(template_path or TEMPLATE_PATH)
    if not path.exists():
        raise FileNotFoundError(f'Goals template missing: {path}')

    sales_agg = aggregate_sales_from_invoice_list(invoice_list, year=year)

    # Enrich AR summary with bridges breakdown
    ar = dict(ar_summary or {})
    if ar.get('top_customers_by_balance') or ar.get('chase_list'):
        customers = ar.get('top_customers_by_balance') or []
        # also use all customers if present
        if ar.get('chase_list'):
            # chase is subset; prefer full list if we stored it
            pass
        lines, parts = _bridges_ar_breakdown(
            ar.get('all_customers') or ar.get('top_customers_by_balance') or ar.get('chase_list')
        )
        ar['bridges_lines'] = lines
        ar['bridges_source_names'] = parts

    wb = load_workbook(path)

    # Drop Profit and Margin unless we later merge P&L
    if 'Profit and Margin' in wb.sheetnames:
        del wb['Profit and Margin']

    if 'Monthly Sales' in wb.sheetnames:
        _fill_sales_sheet(wb['Monthly Sales'], sales_agg)
        _ensure_cf(wb['Monthly Sales'], 'B5:M6', le=True)
        _ensure_cf(wb['Monthly Sales'], 'B11:M12', le=True)
        _ensure_cf(wb['Monthly Sales'], 'B17:M18', le=True)
        _ensure_cf(wb['Monthly Sales'], 'B23:M24', le=True)
        _ensure_cf(wb['Monthly Sales'], 'C29:M30', le=True)

    if 'Monthly Team' in wb.sheetnames:
        _fill_team_sheet(wb['Monthly Team'], sales_agg)
        _ensure_cf(wb['Monthly Team'], 'B6:M7')
        _ensure_cf(wb['Monthly Team'], 'B14:M15', le=True)

    if 'Quarterly Breakdowns' in wb.sheetnames:
        _ensure_cf(wb['Quarterly Breakdowns'], 'B4:E5')
        _ensure_cf(wb['Quarterly Breakdowns'], 'D8:D19')
        _ensure_cf(wb['Quarterly Breakdowns'], 'I8:I19')

    # Re-open data_only won't work without Excel cache for new values —
    # paint Difference rows by computing from filled Actual vs Goal where possible
    _paint_computed_diffs(wb)

    insights = _build_insights(sales_agg, ar, notes_by_customer=notes_by_customer, pl_summary=pl_summary)

    if 'Insights' in wb.sheetnames:
        del wb['Insights']
    ws_i = wb.create_sheet('Insights', 0)
    # Wide layout: merge A–H so email/viewer doesn't clip long lines
    ws_i.column_dimensions['A'].width = 100
    for col in range(2, 9):
        ws_i.column_dimensions[get_column_letter(col)].width = 18
    ws_i['A1'] = 'Monday Numbers · Insights'
    ws_i['A1'].font = Font(name='Calibri', size=20, bold=True, color='FFFFFFFF')
    ws_i['A1'].fill = FILL_HEADER
    ws_i['A1'].alignment = Alignment(vertical='center', horizontal='left')
    ws_i.merge_cells('A1:H1')
    ws_i.row_dimensions[1].height = 36
    for i, line in enumerate(insights.splitlines(), start=3):
        cell = ws_i.cell(i, 1, line)
        # Merge each insight line across A–H so viewers show full width
        ws_i.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        # One list, shared with the builder. The renderer used to carry its own
        # copy of the section titles plus a prefix fallback, so adding a
        # section meant styling it in a second place or watching it render as
        # body text.
        is_head = line in INSIGHT_SECTIONS
        if is_head:
            cell.font = Font(name='Calibri', size=16, bold=True, color='FF1A5276')
            ws_i.row_dimensions[i].height = 24
        elif line.startswith('Monday Numbers') or line.startswith('Generated'):
            cell.font = Font(name='Calibri', size=14, bold=True, color='FF333333')
            ws_i.row_dimensions[i].height = 20
        else:
            cell.font = Font(name='Calibri', size=14, color='FF222222')
            # Grow row for long wrapped lines (~90 chars per visual line at this width)
            text_len = len(line or '')
            wraps = max(1, (text_len // 95) + 1)
            ws_i.row_dimensions[i].height = max(20, 16 * wraps + 4)
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    # Keep Insights from being shrunk by autosize later
    ws_i.sheet_view.showGridLines = False

    if 'AR Totals' in wb.sheetnames:
        del wb['AR Totals']
    ws_ar = wb.create_sheet('AR Totals', 1)
    ws_ar['A1'] = 'Company A/R'
    ws_ar['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFFFF')
    ws_ar['A1'].fill = FILL_HEADER
    ws_ar.merge_cells('A1:C1')
    ws_ar.row_dimensions[1].height = 26
    ws_ar['A3'] = 'Bucket'
    ws_ar['B3'] = 'Amount'
    for col in (1, 2):
        ws_ar.cell(3, col).fill = FILL_YELLOW
        ws_ar.cell(3, col).font = FONT_BOLD
    if ar.get('grand_total'):
        g = ar['grand_total']
        rows = [
            ('Current', g.get('current')),
            ('1–30', g.get('1_30')),
            ('31–60', g.get('31_60')),
            ('61–90', g.get('61_90')),
            ('91+', g.get('91_and_over')),
            ('TOTAL AR', g.get('total')),
        ]
        for i, (lab, amt) in enumerate(rows, 4):
            ws_ar.cell(i, 1, lab).font = Font(name='Calibri', size=12)
            c = ws_ar.cell(i, 2, float(amt or 0))
            c.number_format = '"$"#,##0'
            c.font = Font(name='Calibri', size=12)
            if lab == 'TOTAL AR':
                ws_ar.cell(i, 1).font = Font(name='Calibri', size=12, bold=True)
                c.font = Font(name='Calibri', size=12, bold=True)
        r = 11
        ws_ar.cell(r, 1, 'Bridges').font = FONT_BOLD
        r += 1
        for b in ar.get('bridges_lines') or []:
            ws_ar.cell(r, 1, b['label']).font = Font(name='Calibri', size=12)
            c = ws_ar.cell(r, 2, float(b['total'] or 0))
            c.number_format = '"$"#,##0'
            c.font = Font(name='Calibri', size=12)
            r += 1
        if notes_by_customer:
            r += 1
            ws_ar.cell(r, 1, 'Past-due notes').font = FONT_BOLD
            r += 1
            ws_ar.cell(r, 1, 'Customer').font = FONT_BOLD
            ws_ar.cell(r, 2, 'Past due $').font = FONT_BOLD
            ws_ar.cell(r, 3, 'Comment').font = FONT_BOLD
            for col in (1, 2, 3):
                ws_ar.cell(r, col).fill = FILL_YELLOW
            r += 1
            sorted_notes = sorted(
                notes_by_customer.items(),
                key=lambda kv: -float((_note_parts(kv[1])[1] or 0) or 0),
            )
            for cust, val in sorted_notes:
                note, overdue, _total = _note_parts(val)
                if not note or not str(note).strip():
                    continue
                ws_ar.cell(r, 1, cust).font = Font(name='Calibri', size=11)
                if overdue is not None:
                    c = ws_ar.cell(r, 2, float(overdue or 0))
                    c.number_format = '"$"#,##0'
                    c.font = Font(name='Calibri', size=11, bold=True)
                else:
                    ws_ar.cell(r, 2, '—').font = Font(name='Calibri', size=11)
                ws_ar.cell(r, 3, str(note).strip()).font = Font(name='Calibri', size=11)
                r += 1
    else:
        ws_ar['A4'] = 'Upload A/R Aging Summary to populate.'

    # Optional P&L summary sheet (sales / margin / profit only)
    if pl_summary:
        if 'P&L Snapshot' in wb.sheetnames:
            del wb['P&L Snapshot']
        ws_pl = wb.create_sheet('P&L Snapshot', 2)
        ws_pl['A1'] = 'P&L snapshot — sales, margin, profit and what moved (YoY)'
        ws_pl['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFFFF')
        ws_pl['A1'].fill = FILL_HEADER
        ws_pl.merge_cells('A1:D1')
        ws_pl['A2'] = pl_summary.get('period_label') or ''
        headers = ['Metric', 'This year', 'Prior year', 'YoY']
        for i, h in enumerate(headers, 1):
            cell = ws_pl.cell(4, i, h)
            cell.fill = FILL_YELLOW
            cell.font = FONT_BOLD
        metrics = [
            ('Income', 'income_ty', 'income_py'),
            ('COGS', 'cogs_ty', 'cogs_py'),
            ('Gross profit', 'gross_profit_ty', 'gross_profit_py'),
            ('Net income', 'net_income_ty', 'net_income_py'),
        ]
        for ri, (lab, k_ty, k_py) in enumerate(metrics, 5):
            ws_pl.cell(ri, 1, lab)
            ty = pl_summary.get(k_ty)
            py = pl_summary.get(k_py)
            if ty is not None:
                c = ws_pl.cell(ri, 2, float(ty))
                c.number_format = '"$"#,##0'
            if py is not None:
                c = ws_pl.cell(ri, 3, float(py))
                c.number_format = '"$"#,##0'
            if ty is not None and py not in (None, 0):
                yoy = (float(ty) - float(py)) / abs(float(py))
                c = ws_pl.cell(ri, 4, yoy)
                c.number_format = '0.0%'
                if yoy > 0:
                    c.fill = FILL_POS
                elif yoy < 0:
                    c.fill = FILL_NEG
        # Margins
        ws_pl.cell(10, 1, 'Gross margin')
        ws_pl.cell(11, 1, 'Net margin')
        if pl_summary.get('income_ty'):
            if pl_summary.get('gross_profit_ty') is not None:
                c = ws_pl.cell(10, 2, pl_summary['gross_profit_ty'] / pl_summary['income_ty'])
                c.number_format = '0.0%'
            if pl_summary.get('net_income_ty') is not None:
                c = ws_pl.cell(11, 2, pl_summary['net_income_ty'] / pl_summary['income_ty'])
                c.number_format = '0.0%'
        if pl_summary.get('income_py'):
            if pl_summary.get('gross_profit_py') is not None:
                c = ws_pl.cell(10, 3, pl_summary['gross_profit_py'] / pl_summary['income_py'])
                c.number_format = '0.0%'
            if pl_summary.get('net_income_py') is not None:
                c = ws_pl.cell(11, 3, pl_summary['net_income_py'] / pl_summary['income_py'])
                c.number_format = '0.0%'

        # ── the detail, and what moved in it ────────────────────────────
        row = 13
        movers = pl_movers(pl_summary)
        if movers:
            ws_pl.cell(row, 1, 'Unusual moves year over year').font = Font(
                name='Calibri', size=13, bold=True, color='FF1A5276')
            row += 1
            for i, h in enumerate(
                    ['Line', 'This year', 'Prior year', 'Change', 'YoY'], 1):
                c = ws_pl.cell(row, i, h)
                c.fill = FILL_YELLOW
                c.font = FONT_BOLD
            row += 1
            for m in movers:
                ws_pl.cell(row, 1, m['label'])
                ws_pl.cell(row, 2, m['ty']).number_format = '"$"#,##0'
                ws_pl.cell(row, 3, m['py']).number_format = '"$"#,##0'
                c = ws_pl.cell(row, 4, m['delta'])
                c.number_format = '"$"#,##0;[Red]-"$"#,##0'
                c.fill = FILL_POS if m['delta'] > 0 else FILL_NEG
                if m['pct'] is None:
                    ws_pl.cell(row, 5, 'new' if m['kind'] == 'new' else 'stopped')
                else:
                    ws_pl.cell(row, 5, m['pct']).number_format = '0.0%'
                row += 1
            ws_pl.cell(
                row, 1,
                f'Listed when a line moved at least ${PL_MOVE_MIN_DOLLARS:,.0f} '
                f'and {PL_MOVE_MIN_PCT:.0%} year over year.'
            ).font = Font(name='Calibri', size=10, italic=True, color='FF666666')
            row += 2

        detail = pl_summary.get('lines') or []
        if detail:
            ws_pl.cell(row, 1, 'All line items').font = Font(
                name='Calibri', size=13, bold=True, color='FF1A5276')
            row += 1
            for i, h in enumerate(['Line', 'This year', 'Prior year', 'Change'], 1):
                c = ws_pl.cell(row, i, h)
                c.fill = FILL_YELLOW
                c.font = FONT_BOLD
            row += 1
            for line in sorted(detail, key=lambda d: -abs(float(d.get('ty') or 0))):
                ty = float(line.get('ty') or 0)
                py = float(line.get('py') or 0)
                ws_pl.cell(row, 1, line.get('label'))
                ws_pl.cell(row, 2, ty).number_format = '"$"#,##0'
                ws_pl.cell(row, 3, py).number_format = '"$"#,##0'
                ws_pl.cell(row, 4, ty - py).number_format = '"$"#,##0;[Red]-"$"#,##0'
                row += 1
        withheld = pl_summary.get('withheld_comp_lines') or 0
        if withheld:
            # Say it on the sheet, not only in the insights. A list that is
            # silently incomplete is worse than a shorter list.
            ws_pl.cell(
                row, 1,
                f'{withheld} compensation line(s) withheld — payroll and owner '
                f'comp are kept out of Hub reports.'
            ).font = Font(name='Calibri', size=10, italic=True, color='FF666666')

    # Widen columns so currency doesn't show as ####
    _autosize_workbook(wb)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), insights, {
        'sales_agg': {
            'invoice_count': sales_agg['invoice_count'],
            'split_invoice_count': sales_agg['split_invoice_count'],
            'team_ytd': sum(sales_agg['team_month'].values()),
            'by_rep_ytd': {r: sum(m.values()) for r, m in sales_agg['by_rep_month'].items()},
        },
        'pl_included': bool(pl_summary),
        'generated_at': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
    }


def _autosize_workbook(wb):
    """Resize columns so currency values fit (avoid ####).

    Insights is handled separately (wide merged layout) — do not shrink it.
    """
    from openpyxl.utils import get_column_letter
    for ws in wb.worksheets:
        if ws.title == 'Insights':
            # Preserve wide Insights layout for email/viewers
            ws.column_dimensions['A'].width = 100
            for col_idx in range(2, 9):
                ws.column_dimensions[get_column_letter(col_idx)].width = 18
            continue
        max_row = min(ws.max_row or 1, 80)
        max_col = min(ws.max_column or 1, 20)
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            max_len = 8
            for row_idx in range(1, max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is None:
                    continue
                if isinstance(cell.value, (int, float)):
                    max_len = max(max_len, 14)
                else:
                    max_len = max(max_len, min(60, len(str(cell.value)) + 2))
            if ws.title in ('Monthly Team', 'Monthly Sales', 'Quarterly Breakdowns') and col_idx >= 2:
                max_len = max(max_len, 13)
            if ws.title == 'AR Totals':
                max_len = max(max_len, 18 if col_idx == 1 else 16)
            if ws.title == 'P&L Snapshot':
                max_len = max(max_len, 14)
            ws.column_dimensions[letter].width = min(55, max(10, max_len))


def _paint_computed_diffs(wb):
    """After writing Actual values, compute Difference/% hit cells as values with colors."""
    if 'Monthly Sales' not in wb.sheetnames:
        return
    ws = wb['Monthly Sales']
    # For each rep block: Actual row R, Goal R-1, Diff R+1, Pct R+2
    blocks = [(4, 3, 5, 6), (10, 9, 11, 12), (16, 15, 17, 18), (22, 21, 23, 24), (28, 27, 29, 30)]
    for actual_r, goal_r, diff_r, pct_r in blocks:
        for m, col in MONTH_COLS.items():
            try:
                actual = float(ws.cell(actual_r, col).value or 0)
            except (TypeError, ValueError):
                actual = 0.0
            goal_raw = ws.cell(goal_r, col).value
            try:
                if isinstance(goal_raw, str) and goal_raw.startswith('='):
                    # leave formula for goal; read cached or skip paint for formula-only
                    goal = None
                else:
                    goal = float(goal_raw or 0) if goal_raw not in ('$ -', '', None) else 0.0
            except (TypeError, ValueError):
                goal = 0.0
            if goal is None:
                continue
            diff = actual - goal
            dcell = ws.cell(diff_r, col, round(diff, 2))
            dcell.number_format = '"$"#,##0.00'
            if diff > 0:
                dcell.fill = FILL_POS
                dcell.font = Font(color='FF006100', name='Calibri', size=11)
            elif diff < 0:
                dcell.fill = FILL_NEG
                dcell.font = Font(color='FF9C0006', name='Calibri', size=11)
            if goal:
                pct = diff / goal
                pcell = ws.cell(pct_r, col, pct)
                pcell.number_format = '0%'
                if pct > 0:
                    pcell.fill = FILL_POS
                    pcell.font = Font(color='FF006100', name='Calibri', size=11)
                elif pct < 0:
                    pcell.fill = FILL_NEG
                    pcell.font = Font(color='FF9C0006', name='Calibri', size=11)

    if 'Monthly Team' in wb.sheetnames:
        ws = wb['Monthly Team']
        # 2026 Actual row 13, Goal row 11 (formulas from %), Diff 14, Pct 15
        for m, col in MONTH_COLS.items():
            actual = ws.cell(13, col).value
            if actual is None:
                continue
            try:
                actual = float(actual)
            except (TypeError, ValueError):
                continue
            # Goal $ may be formula — use Goal % * 10M from N11
            try:
                annual = float(ws.cell(11, 14).value or 10000000)
            except (TypeError, ValueError):
                annual = 10000000.0
            goal_pct = ws.cell(10, col).value
            try:
                goal = float(goal_pct) * annual if goal_pct is not None else 0.0
            except (TypeError, ValueError):
                goal = 0.0
            diff = actual - goal
            dcell = ws.cell(14, col, round(diff, 2))
            dcell.number_format = '"$"#,##0.00'
            if diff > 0:
                dcell.fill = FILL_POS
                dcell.font = Font(color='FF006100', name='Calibri', size=11)
            elif diff < 0:
                dcell.fill = FILL_NEG
                dcell.font = Font(color='FF9C0006', name='Calibri', size=11)
            if goal:
                pct = diff / goal
                pcell = ws.cell(15, col, pct)
                pcell.number_format = '0%'
                if pct > 0:
                    pcell.fill = FILL_POS
                    pcell.font = Font(color='FF006100', name='Calibri', size=11)
                elif pct < 0:
                    pcell.fill = FILL_NEG
                    pcell.font = Font(color='FF9C0006', name='Calibri', size=11)


def _ensure_cf(ws, rng, le=False):
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator='greaterThan', formula=['0'], fill=FILL_POS_CF, font=FONT_POS),
    )
    op = 'lessThanOrEqual' if le else 'lessThan'
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator=op, formula=['0'], fill=FILL_NEG_CF, font=FONT_NEG),
    )
