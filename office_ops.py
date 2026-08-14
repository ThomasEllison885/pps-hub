"""Office Ops — Stephanie + Thomas workspace for AR digests and weekly Numbers.

Owner request 2026-08-03 / build-out 2026-08:
  - Access: office_manager (Stephanie) + admin (Thomas) only.
  - Files land via Hub upload (Postgres), not a shared team vault dump.
  - AR Aging Summary → totals / chase list / Numbers draft skeleton.
  - AR Aging Detail → open invoices by age bucket (invoice-level chase).
  - Later: sticky notes, salesman/50-50 from invoice custom field, Sub Info,
    Outlook sheet, Gmail drafts.

Does NOT write to QuickBooks. Does NOT auto-send email.
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import date, datetime
from decimal import Decimal

from psycopg2.extras import RealDictCursor

# Who may open /office-ops and upload. Admin always; Stephanie by key + role.
OFFICE_OPS_USER_KEYS = frozenset({'stephanie_whetstone', 'thomas_ellison'})
OFFICE_OPS_ROLES = frozenset({'office_manager', 'admin'})

MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB

# Each vision-pass row is a real network round-trip (plus a pdftoppm render
# for scanned PDFs) -- a board-wide run over dozens of rows in one request
# can outlast gunicorn's 120s worker timeout and just look like "Network
# error" in the browser. The route enforces this cap server-side regardless
# of what the client asks for, so the run-in-batches property can't be
# bypassed by a client bug.
VISION_PASS_BATCH_SIZE = 6
# ar_aging kept for early uploads; new kinds are explicit.
ALLOWED_KINDS = frozenset({
    'ar_aging',
    'ar_aging_summary',
    'ar_aging_detail',
    'invoice_list',
    'monthly_outlook',
    'monday_report',
    'profit_loss',
})
KIND_SUMMARY = 'ar_aging_summary'
KIND_DETAIL = 'ar_aging_detail'
KIND_INVOICE_LIST = 'invoice_list'
KIND_OUTLOOK = 'monthly_outlook'
KIND_MONDAY = 'monday_report'
KIND_PL = 'profit_loss'
PACK_KIND = 'ar_aging'  # combined AR pack

# Customers always called out in the brief (ops judgment, not accounting).
LEGACY_CUSTOMER_HINTS = (
    'bridges of pine creek',
    'bopc',
    '3800 cornell',
)

BOPC_ALIASES = ('bridges of pine creek', 'bopc')


def can_access_office_ops(users, user_key):
    if not user_key:
        return False
    if user_key in OFFICE_OPS_USER_KEYS:
        return True
    role = (users.get(user_key) or {}).get('role')
    return role in OFFICE_OPS_ROLES


def init_tables(cur):
    cur.execute('''
        CREATE TABLE IF NOT EXISTS office_ops_files (
            id SERIAL PRIMARY KEY,
            kind VARCHAR(50) NOT NULL,
            filename VARCHAR(255) NOT NULL,
            mime_type VARCHAR(100),
            size_bytes INTEGER NOT NULL,
            file_data BYTEA NOT NULL,
            uploaded_by VARCHAR(100),
            uploaded_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    cur.execute(
        'CREATE INDEX IF NOT EXISTS idx_office_ops_files_kind_time '
        'ON office_ops_files(kind, uploaded_at DESC)'
    )
    cur.execute('''
        CREATE TABLE IF NOT EXISTS office_ops_packs (
            id SERIAL PRIMARY KEY,
            pack_date DATE NOT NULL,
            source_file_id INTEGER REFERENCES office_ops_files(id) ON DELETE SET NULL,
            kind VARCHAR(50) NOT NULL DEFAULT 'ar_aging',
            summary_json JSONB NOT NULL,
            numbers_draft_md TEXT,
            created_by VARCHAR(100),
            created_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    cur.execute(
        'CREATE INDEX IF NOT EXISTS idx_office_ops_packs_created '
        'ON office_ops_packs(created_at DESC)'
    )
    # Stephanie's past-due status notes (survive re-uploads; keyed by customer name)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS office_ops_ar_notes (
            customer_key VARCHAR(255) PRIMARY KEY,
            customer_display VARCHAR(255) NOT NULL,
            note TEXT NOT NULL DEFAULT '',
            updated_by VARCHAR(100),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    ''')


def _money(val):
    if val is None or val == '':
        return 0.0
    if isinstance(val, (int, float, Decimal)):
        return float(val)
    s = str(val).strip().replace(',', '').replace('$', '').replace('(', '-').replace(')', '')
    if not s or s == '-':
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _norm_header(h):
    if h is None:
        return ''
    return re.sub(r'\s+', ' ', str(h).strip().lower())


def _is_summary_header_row(cells):
    """QB A/R Aging Summary: CURRENT | 1-30 | … | Total (customer in col 0)."""
    joined = ' '.join(_norm_header(c) for c in cells if c is not None)
    has_current = 'current' in joined
    has_1_30 = bool(re.search(r'1\s*[-–—]\s*30', joined) or '1-30' in joined)
    has_total = 'total' in joined
    # Some exports use "Not yet due" instead of CURRENT
    has_not_yet = 'not yet due' in joined
    return (has_current or has_not_yet) and has_1_30 and has_total


def _is_detail_header_row(cells):
    """QB A/R Aging Detail: Date | Transaction type | Num | Customer full name | …"""
    joined = _header_joined(cells)
    # Invoice List also has Transaction type + customer + open balance. A Sales
    # Rep column (or an Invoice List title handled elsewhere) means it is not aging.
    if _has_sales_rep_header(joined):
        return False
    return (
        ('customer full name' in joined and 'open balance' in joined)
        or (
            'date' in joined
            and 'customer' in joined
            and 'open balance' in joined
            and 'transaction' in joined
        )
    )


# Back-compat name used by older call sites
def _is_header_row(cells):
    return _is_summary_header_row(cells)


def _header_joined(cells):
    return ' '.join(_norm_header(c) for c in cells if c is not None)


def _has_sales_rep_header(joined):
    """True if a header row names a sales-rep column (QB keeps renaming this)."""
    if not joined:
        return False
    if any(tok in joined for tok in (
        'sales rep', 'salesman', 'salesperson', 'salesrep', 'sales-rep',
        'sales rep name', 'rep name',
    )):
        return True
    # lone "rep" / "reps" column, but not "description" / "open"
    tokens = set(joined.split())
    return 'rep' in tokens or 'reps' in tokens


def _has_invoice_list_identity(joined):
    """Date + invoice # + customer + dollars — the Invoice List core, Sales Rep optional.

    QB's 2026 report redesign dropped or renamed Sales Rep on Invoice List by Date
    and relabeled Date → Transaction date. Identity without Sales Rep still means
    Invoice List, not aging.
    """
    has_num = (
        'num' in joined or 'no.' in joined or 'number' in joined
        or 'invoice #' in joined or 'invoice no' in joined
    )
    has_name = 'name' in joined or 'customer' in joined
    has_amt = 'open balance' in joined or 'amount' in joined
    has_date = 'date' in joined
    return bool(has_num and has_name and has_amt and has_date)


def _is_invoice_list_header_row(cells):
    """QB Invoice List by Date with a Sales Rep column (old / preferred export)."""
    joined = _header_joined(cells)
    return _has_invoice_list_identity(joined) and _has_sales_rep_header(joined)


def _is_invoice_list_header_row_loose(cells):
    """Same identity without requiring Sales Rep (QB 2026 dropped that column)."""
    return _has_invoice_list_identity(_header_joined(cells))


def _title_blob(rows, limit=8):
    parts = []
    for row in (rows or [])[:limit]:
        for cell in row:
            if cell and isinstance(cell, str) and cell.strip():
                parts.append(cell.strip().lower())
    return ' '.join(parts)


def detect_ar_report_type(filename, raw_bytes):
    """Return 'summary' | 'detail' | 'invoice_list' | None from filename + content peek."""
    name = (filename or '').lower().replace('+', ' ')
    if 'sales by customer type' in name:
        return None
    if 'invoice list' in name or 'invoice_list' in name:
        return 'invoice_list'
    if 'detail' in name and 'aging' in name:
        return 'detail'
    if 'summary' in name and 'aging' in name:
        return 'summary'
    # Bare "detail" / "summary" in the filename is last resort — title wins below
    # because QB's new Invoice List has been saved as "Sheet1.xlsx".

    # Peek rows without full parse
    rows = _load_tabular_rows(filename, raw_bytes, max_rows=40)
    title = _title_blob(rows)
    if 'sales by customer type' in title:
        return None
    if 'invoice list' in title:
        return 'invoice_list'
    if 'aging detail' in title:
        return 'detail'
    if 'aging summary' in title:
        return 'summary'

    for row in rows:
        if _is_invoice_list_header_row(row):
            return 'invoice_list'
        if _is_detail_header_row(row):
            return 'detail'
        if _is_summary_header_row(row):
            return 'summary'

    # Filename fallback after title (so "Report Detail.xlsx" that is really
    # Invoice List by Date still classifies from the title above).
    if 'detail' in name:
        return 'detail'
    if 'summary' in name:
        return 'summary'
    return None


def _load_tabular_rows(filename, raw_bytes, max_rows=None):
    name = (filename or '').lower()
    if name.endswith('.csv') or name.endswith('.txt'):
        text = raw_bytes.decode('utf-8-sig', errors='replace')
        reader = csv.reader(io.StringIO(text))
        rows = []
        for i, row in enumerate(reader):
            rows.append(row)
            if max_rows and i + 1 >= max_rows:
                break
        return rows
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        rows.append(list(row))
        if max_rows and i >= max_rows:
            break
    wb.close()
    return rows


def _map_aging_cols(header_cells):
    """Map column index → bucket key from QB-style header row."""
    colmap = {}
    for i, h in enumerate(header_cells):
        n = _norm_header(h)
        if not n:
            continue
        if n in ('', 'customer', 'name') or i == 0 and 'current' not in n:
            if i == 0:
                colmap['customer'] = i
            continue
        if n == 'current':
            colmap['current'] = i
        elif '1' in n and '30' in n:
            colmap['1_30'] = i
        elif '31' in n and '60' in n:
            colmap['31_60'] = i
        elif '61' in n and '90' in n:
            colmap['61_90'] = i
        elif '91' in n or 'over' in n:
            colmap['91_and_over'] = i
        elif n == 'total':
            colmap['total'] = i
    if 'customer' not in colmap:
        colmap['customer'] = 0
    return colmap


def _row_bucket(cells, colmap):
    def g(key):
        idx = colmap.get(key)
        if idx is None or idx >= len(cells):
            return 0.0
        return _money(cells[idx])

    total = g('total')
    if total == 0.0:
        total = g('current') + g('1_30') + g('31_60') + g('61_90') + g('91_and_over')
    return {
        'current': g('current'),
        '1_30': g('1_30'),
        '31_60': g('31_60'),
        '61_90': g('61_90'),
        '91_and_over': g('91_and_over'),
        'total': total,
    }


def parse_pl_bytes(filename, raw_bytes):
    """Parse QB Profit & Loss export with this year / prior year / % change columns.

    Returns sales / gross profit / net income only (for margin insights — not expense detail).
    """
    rows = _load_tabular_rows(filename, raw_bytes)
    if not rows:
        raise ValueError('P&L file is empty.')
    period = None
    for row in rows[:8]:
        for cell in row:
            if cell and isinstance(cell, str) and (
                'january' in cell.lower() or '–' in cell or '-' in cell
            ):
                if any(ch.isdigit() for ch in cell):
                    period = cell.strip()
                    break

    def find_row(*labels):
        for row in rows:
            if not row or row[0] is None:
                continue
            lab = str(row[0]).strip().lower()
            for want in labels:
                if lab == want.lower() or lab.startswith(want.lower()):
                    def num(i):
                        if i >= len(row) or row[i] in (None, ''):
                            return None
                        try:
                            return float(row[i])
                        except (TypeError, ValueError):
                            return None
                    return {
                        'label': str(row[0]).strip(),
                        'ty': num(1),
                        'py': num(2),
                        'pct': num(3),
                    }
        return None

    income = find_row('Total for Income', 'Total Income')
    # Prefer Services as sales proxy if present, else total income
    services = find_row('Services')
    cogs = find_row('Total for Cost of Goods Sold', 'Total Cost of Goods Sold')
    gp = find_row('Gross Profit')
    ni = find_row('Net Income')
    if not income and not services:
        raise ValueError(
            'Could not find Income/Services on P&L. Export Profit and Loss with '
            'columns for this year, prior year, and % change.'
        )
    inc_ty = (services or income or {}).get('ty')
    # Prefer Total for Income for top line if Services is subset
    if income and income.get('ty') is not None:
        inc_ty = income['ty']
        inc_py = income.get('py')
    else:
        inc_py = (services or {}).get('py')

    return {
        'report': 'Profit and Loss',
        'report_kind': 'profit_loss',
        'period_label': period,
        'income_ty': inc_ty,
        'income_py': inc_py if income else (services or {}).get('py'),
        'services_ty': (services or {}).get('ty'),
        'services_py': (services or {}).get('py'),
        'cogs_ty': (cogs or {}).get('ty'),
        'cogs_py': (cogs or {}).get('py'),
        'gross_profit_ty': (gp or {}).get('ty'),
        'gross_profit_py': (gp or {}).get('py'),
        'net_income_ty': (ni or {}).get('ty'),
        'net_income_py': (ni or {}).get('py'),
        'parsed_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
    }


def parse_ar_aging_bytes(filename, raw_bytes, expect=None):
    """Parse AR/invoice export. expect: summary|detail|invoice_list|None.

    Returns a pack fragment; kinds merge in process_ar_file.
    """
    detected = detect_ar_report_type(filename, raw_bytes)
    kind = expect or detected
    if expect and detected and expect != detected:
        title = _title_blob(_load_tabular_rows(filename, raw_bytes, max_rows=8))
        fname = (filename or '').lower()
        # New Invoice List (no Sales Rep) header-detects as aging detail.
        if expect == 'invoice_list' and detected == 'detail' and 'aging' not in title and not (
            'aging' in fname and 'detail' in fname
        ):
            detected = 'invoice_list'
            kind = 'invoice_list'
        else:
            pretty = {
                'summary': 'A/R Aging Summary',
                'detail': 'A/R Aging Detail',
                'invoice_list': 'Invoice List by Date',
            }
            raise ValueError(
                f'This file looks like {pretty.get(detected, detected)}, '
                f'but it was uploaded as {pretty.get(expect, expect)}. '
                f'Use the matching drop zone (or upload again — we auto-detect).'
            )
    if kind == 'invoice_list':
        return _parse_invoice_list(filename, raw_bytes)
    if kind == 'detail':
        return _parse_ar_detail(filename, raw_bytes)
    if kind == 'summary':
        name = (filename or '').lower()
        if name.endswith('.csv') or name.endswith('.txt'):
            return _parse_ar_csv(raw_bytes)
        return _parse_ar_xlsx(raw_bytes)
    # Unknown
    peek = _title_blob(_load_tabular_rows(filename, raw_bytes, max_rows=8))
    if 'sales by customer type' in peek or 'sales by customer type' in (filename or '').lower():
        raise ValueError(
            'This is QuickBooks “Sales by Customer Type Detail” — line items, no customer '
            'and no Sales Rep. That will not fill the Numbers board.\n'
            'Export Reports → Sales → Invoice List by Date (Excel). Sales Rep is helpful '
            'but no longer required. Do not use Sales by Customer Type Detail.'
        )
    raise ValueError(
        'Could not identify the export. From QuickBooks use one of:\n'
        '• A/R Aging Summary — CURRENT / 1-30 / … / Total\n'
        '• A/R Aging Detail — Date / Transaction type / Customer / Open balance\n'
        '• Invoice List by Date — one row per invoice (Sales Rep if QB still offers it)\n'
        'Then use the matching upload box on Office Ops.'
    )


# Short first names used in 50/50 Sales Rep fields → full QB sales rep names.
_SALES_REP_ALIASES = {
    'adam': 'Adam Cupito',
    'andy': 'Andy Potts',
    'tony': 'Tony Cumella',
    'rachel': 'Rachel Farler',
    'thomas': 'Thomas Ellison',
    'tom': 'Thomas Ellison',
}


def _normalize_sales_rep_name(part):
    p = (part or '').strip()
    if not p:
        return p
    # Already a full known name
    low = p.lower()
    for full in _SALES_REP_ALIASES.values():
        if low == full.lower():
            return full
    # First-name only in multi-rep strings
    if low in _SALES_REP_ALIASES:
        return _SALES_REP_ALIASES[low]
    first = low.split()[0] if low.split() else low
    if first in _SALES_REP_ALIASES and len(p.split()) == 1:
        return _SALES_REP_ALIASES[first]
    return p


def _parse_sales_reps(raw):
    """Split 'Adam / Andy' → ['Adam Cupito', 'Andy Potts']; single name unchanged."""
    if not raw:
        return []
    s = str(raw).strip()
    if not s:
        return []
    parts = re.split(r'\s*/\s*|\s+and\s+|&', s, flags=re.I)
    out = []
    for p in parts:
        n = _normalize_sales_rep_name(p)
        if n and n not in out:
            out.append(n)
    return out


def _looks_like_aging_bucket(text):
    t = (text or '').strip().lower()
    if not t:
        return False
    return (
        'past due' in t
        or t in ('current', 'not yet due')
        or 'or more' in t
        or bool(re.search(r'\d+\s*[-–—]\s*\d+', t))
    )


def _is_unspecified_rep_label(text):
    t = (text or '').strip().lower()
    return t in (
        'not specified', 'unspecified', 'unassigned', 'none',
        'no sales rep', 'no salesman', '(unassigned)',
    )


def _is_plausible_rep_group_label(text):
    """QB group-header text: 'Adam / Andy', 'Tony', 'Adam Cupito'."""
    t = (text or '').strip()
    if not t or t.lower().startswith('total'):
        return False
    if _looks_like_aging_bucket(t) or _is_unspecified_rep_label(t):
        return False
    firsts = [
        p.strip().split()[0].lower()
        for p in re.split(r'\s*/\s*|\s+and\s+|&', t)
        if p.strip()
    ]
    if firsts and all(f in _SALES_REP_ALIASES for f in firsts):
        return True
    # Full known name
    if t.lower() in {v.lower() for v in _SALES_REP_ALIASES.values()}:
        return True
    # First Last — future reps, not a sentence
    parts = t.split()
    if (
        len(parts) == 2
        and parts[0][0].isalpha()
        and parts[1][0].isalpha()
        and not any(ch.isdigit() for ch in t)
        and len(t) < 40
    ):
        return True
    return False


def _invoice_date_str(val):
    """Normalize QB / Sheets / Excel dates to YYYY-MM-DD for the Numbers pack."""
    if val is None or val == '':
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)) and 20000 < float(val) < 80000:
        try:
            from datetime import timedelta
            d = date(1899, 12, 30) + timedelta(days=int(val))
            return d.strftime('%Y-%m-%d')
        except (OverflowError, ValueError):
            pass
    return str(val).strip()


def _collapse_invoice_list_line_items(invoices):
    """If QB exported line items (same Num more than once), roll up to one invoice."""
    if not invoices:
        return invoices
    counts = {}
    for inv in invoices:
        counts[inv['num']] = counts.get(inv['num'], 0) + 1
    if all(c == 1 for c in counts.values()):
        return invoices
    rolled = {}
    order = []
    for inv in invoices:
        key = (inv['num'], (inv.get('type') or 'Invoice').lower())
        if key not in rolled:
            rolled[key] = dict(inv)
            order.append(key)
            continue
        slot = rolled[key]
        slot['amount'] = (slot.get('amount') or 0) + (inv.get('amount') or 0)
        # Keep the remaining open balance (same on each line) — do not sum it
        if inv.get('customer') and not slot.get('customer'):
            slot['customer'] = inv['customer']
        if inv.get('sales_reps') and not slot.get('sales_reps'):
            slot['sales_reps'] = inv['sales_reps']
            slot['sales_rep_raw'] = inv.get('sales_rep_raw')
            slot['is_50_50_style'] = inv.get('is_50_50_style', False)
    return [rolled[k] for k in order]


def _map_invoice_list_columns(header):
    """Map Invoice List header cells → keys. Tolerates QB 2026 label changes."""
    col = {}
    for i, h in enumerate(header):
        n = _norm_header(h)
        if not n:
            continue
        if n in ('date', 'transaction date', 'invoice date') or (
            n.endswith(' date') and 'due' not in n
        ):
            col.setdefault('date', i)
        elif 'transaction' in n or n == 'type':
            col.setdefault('type', i)
        elif n in ('num', 'no.', 'no', 'number', '#', 'invoice #', 'invoice no', 'invoice no.', 'invoice number'):
            col.setdefault('num', i)
        elif n in ('name', 'customer', 'customer name', 'customer full name'):
            col.setdefault('name', i)
        elif 'memo' in n or n == 'description':
            col.setdefault('memo', i)
        elif 'due' in n:
            col.setdefault('due', i)
        elif n == 'amount':
            col.setdefault('amount', i)
        elif 'open' in n and 'balance' in n:
            col.setdefault('open', i)
        elif _has_sales_rep_header(n):
            col.setdefault('sales_rep', i)
        elif n in ('rep', 'reps'):
            col.setdefault('sales_rep', i)
    return col


def _parse_invoice_list(filename, raw_bytes):
    """Parse QB Invoice List by Date (Sales Rep when present — 50/50 uses 'A / B')."""
    rows = _load_tabular_rows(filename, raw_bytes)
    if not rows:
        raise ValueError('Invoice List is empty.')

    title = _title_blob(rows)
    if 'sales by customer type' in title or 'sales by customer type' in (filename or '').lower():
        raise ValueError(
            'This is QuickBooks “Sales by Customer Type Detail” — line items, no customer '
            'and no Sales Rep. Export Invoice List by Date instead.'
        )

    date_range = None
    for row in rows[:8]:
        for cell in row:
            if cell and isinstance(cell, str) and re.search(r'\d{4}', cell) and ('-' in cell or 'to' in cell.lower()):
                date_range = cell.strip()
                break

    header_idx = None
    for i, row in enumerate(rows[:40]):
        if _is_invoice_list_header_row(row) or _is_invoice_list_header_row_loose(row):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            'Could not find Invoice List header (expected Date, Num, Name/Customer, Amount). '
            'In QuickBooks: Reports → Sales → Invoice List by Date, Export Excel. '
            'Sales Rep is optional if QB no longer lets you add it.'
        )

    header = rows[header_idx]
    col = _map_invoice_list_columns(header)

    if 'num' not in col or 'name' not in col:
        raise ValueError('Invoice List missing Num or Name/Customer column.')
    missing_sales_col = 'sales_rep' not in col

    def cell(row, key):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    invoices = []
    current_group_raw = None
    used_group_headers = False
    for row in rows[header_idx + 1:]:
        if not row or all(v in (None, '') for v in row):
            continue
        lead = next((str(c).strip() for c in row if c not in (None, '')), '')
        if lead.lower().startswith('total for'):
            continue

        txn = cell(row, 'type')
        num = cell(row, 'num')
        name = cell(row, 'name')
        date_v = cell(row, 'date')

        # QB now groups Invoice List by sales rep: a row that is only
        # "Adam / Andy" (no invoice #). On a flattened Excel the name sits
        # in the Date column. Carry that onto the rows below.
        filled = [c for c in row if c not in (None, '')]
        is_group = (
            num in (None, '')
            and (txn in (None, '') or str(txn).strip().lower() not in (
                'invoice', 'credit memo', 'sales receipt',
            ))
            and len(filled) == 1
            and (
                _is_plausible_rep_group_label(lead)
                or _is_unspecified_rep_label(lead)
            )
        )
        if is_group:
            current_group_raw = None if _is_unspecified_rep_label(lead) else lead
            used_group_headers = True
            continue

        if not num and not name:
            continue
        if txn is not None and str(txn).strip().lower() not in (
            'invoice', 'credit memo', 'sales receipt', ''
        ):
            if not num:
                continue
        if num is None:
            continue

        open_bal = _money(cell(row, 'open'))
        amount = _money(cell(row, 'amount'))
        rep_raw = cell(row, 'sales_rep')
        if not (rep_raw not in (None, '') and str(rep_raw).strip()):
            rep_raw = current_group_raw
        reps = _parse_sales_reps(rep_raw)
        is_split = len(reps) >= 2
        inv = {
            'date': _invoice_date_str(date_v),
            'type': str(txn or 'Invoice').strip(),
            'num': str(int(num)) if isinstance(num, float) and num == int(num) else str(num).strip(),
            'customer': str(name or '').strip(),
            'memo': str(cell(row, 'memo') or '').strip() or None,
            'due_date': _invoice_date_str(cell(row, 'due')),
            'amount': amount,
            'open_balance': open_bal,
            'sales_rep_raw': str(rep_raw).strip() if rep_raw else None,
            'sales_reps': reps,
            'is_50_50_style': is_split,
            'open_share_each': (open_bal / len(reps)) if reps and open_bal else 0.0,
        }
        invoices.append(inv)

    invoices = _collapse_invoice_list_line_items(invoices)

    # Open-AR attribution by rep (equal split for multi-rep)
    by_rep = {}
    open_invoices = [i for i in invoices if (i.get('open_balance') or 0) > 0]
    for inv in open_invoices:
        reps = inv['sales_reps'] or ['(unassigned)']
        share = inv['open_balance'] / len(reps)
        for r in reps:
            slot = by_rep.setdefault(r, {
                'sales_rep': r,
                'open_ar': 0.0,
                'invoice_count': 0,
                'split_invoice_count': 0,
            })
            slot['open_ar'] += share
            slot['invoice_count'] += 1
            if inv['is_50_50_style']:
                slot['split_invoice_count'] += 1

    rep_table = sorted(by_rep.values(), key=lambda x: -x['open_ar'])
    split_count = sum(1 for i in invoices if i['is_50_50_style'])
    open_total = sum(i['open_balance'] for i in open_invoices)

    return {
        'report': 'Invoice List by Date',
        'report_kind': 'invoice_list',
        'source_format': 'invoice_list_xlsx',
        'date_range': date_range,
        'parsed_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        'invoice_list': invoices,
        'invoice_list_count': len(invoices),
        'open_invoice_count': len(open_invoices),
        'open_ar_from_list': open_total,
        'split_invoice_count': split_count,
        'sales_rep_open_ar': rep_table,
        'salesman_field_present': (not missing_sales_col) or used_group_headers,
        'notes_for_humans': (
            [
                'Sales Rep taken from QuickBooks group headers (the “Adam / Andy” section rows). '
                'Multi-name values count as 50/50.',
            ] if used_group_headers and missing_sales_col else [
                'This Invoice List has no Sales Rep column and no group headers we recognized. '
                'Team invoiced totals still fill. Per-rep / 50/50 will show as unassigned.',
            ] if missing_sales_col else [
                'Sales Rep from Invoice List: multi-name values like “Adam / Andy” count as 50/50-style '
                '(open AR attributed equally across listed reps until you tell us a different split rule).',
                'Match open AR aging to salesmen by invoice number when Detail + Invoice List are both uploaded.',
            ]
        ),
    }


def _parse_ar_csv(raw_bytes):
    text = raw_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError('CSV is empty.')
    header_idx = 0
    for i, row in enumerate(rows[:20]):
        if _is_summary_header_row(row) or any(_norm_header(c) == 'customer' for c in row):
            header_idx = i
            break
    colmap = _map_aging_cols(rows[header_idx])
    # Flat CSV from our prior exports: customer,current,1_30,...
    if 'current' not in colmap:
        headers = [_norm_header(h) for h in rows[header_idx]]
        aliases = {
            'customer': 'customer',
            'current': 'current',
            '1_30': '1_30',
            '31_60': '31_60',
            '61_90': '61_90',
            '91_and_over': '91_and_over',
            'total': 'total',
        }
        colmap = {}
        for i, h in enumerate(headers):
            key = aliases.get(h.replace(' ', '_').replace('-', '_'))
            if key:
                colmap[key] = i
        if 'customer' not in colmap:
            colmap['customer'] = 0

    customers = []
    for row in rows[header_idx + 1:]:
        if not row or all(c in (None, '') for c in row):
            continue
        cust = str(row[colmap.get('customer', 0)] or '').strip()
        if not cust or cust.upper() == 'TOTAL':
            continue
        if cust.lower().startswith('total for '):
            cust = cust[10:].strip()
        buckets = _row_bucket(row, colmap)
        if buckets['total'] == 0 and not any(
            buckets[k] for k in ('current', '1_30', '31_60', '61_90', '91_and_over')
        ):
            continue
        customers.append({'customer': cust, **buckets})
    return _build_summary(customers, as_of_label=None, source='csv')


def _parse_ar_xlsx(raw_bytes):
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        raise ValueError('Spreadsheet is empty.')

    as_of_label = None
    for row in rows[:8]:
        for cell in row:
            if cell and isinstance(cell, str) and cell.strip().lower().startswith('as of'):
                as_of_label = cell.strip()
                break

    header_idx = None
    for i, row in enumerate(rows[:25]):
        if _is_summary_header_row(row):
            header_idx = i
            break
    if header_idx is None:
        # Common mistake: Detail file in Summary slot
        for row in rows[:15]:
            if _is_detail_header_row(row):
                raise ValueError(
                    'This is an A/R Aging Detail export, not Summary. '
                    'Use the “Aging Detail” upload box (invoice-level), '
                    'or export Reports → A/R Aging Summary for totals.'
                )
        raise ValueError(
            'Could not find an A/R Aging Summary header row '
            '(expected CURRENT / 1-30 / Total). '
            'In QuickBooks: Reports → Who owes you → A/R Aging Summary → Export to Excel.'
        )

    colmap = _map_aging_cols(rows[header_idx])
    # Walk QB hierarchy so job codes keep parent names:
    #   Chestnut Park / 7163 / Total for Chestnut Park
    # → "Chestnut Park · 7163" (never bare "7163")
    # Standalone "29th Street Living" (has $ but no Total for) does NOT nest
    # the following customers under it.
    customers = []
    parent_stack = []
    body = rows[header_idx + 1:]

    def _has_money(b):
        return bool(
            b.get('total')
            or any(b.get(k) for k in ('current', '1_30', '31_60', '61_90', '91_and_over'))
        )

    def _has_total_for_later(name, from_idx):
        target = name.lower()
        for later in body[from_idx + 1:]:
            if not later or later[0] is None:
                continue
            lab = str(later[0]).strip()
            if lab.lower().startswith('total for '):
                if lab[10:].strip().lower() == target:
                    return True
        return False

    def _display_name(job_or_cust, parents):
        job = (job_or_cust or '').strip()
        if not parents:
            return job
        parent = parents[0]
        if job.lower() == parent.lower():
            return parent
        return f'{parent} · {job}'

    for idx, row in enumerate(body):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if not name:
            continue
        low = name.lower()
        if name.upper() == 'TOTAL' or (
            low.startswith('total') and not low.startswith('total for ')
        ):
            continue
        buckets = _row_bucket(row, colmap)

        if low.startswith('total for '):
            closed = name[10:].strip()
            while parent_stack and parent_stack[-1].lower() != closed.lower():
                parent_stack.pop()
            if parent_stack and parent_stack[-1].lower() == closed.lower():
                parent_stack.pop()
            if _has_money(buckets):
                customers.append({
                    'customer': closed,
                    'parent': closed,
                    'job': None,
                    'is_parent_total': True,
                    **buckets,
                })
            continue

        parents_before = list(parent_stack)
        money = _has_money(buckets)
        will_nest = _has_total_for_later(name, idx)

        if money:
            display = _display_name(name, parents_before)
            customers.append({
                'customer': display,
                'parent': parents_before[0] if parents_before else name,
                'job': name if parents_before else None,
                'is_parent_total': False,
                **buckets,
            })
            # Only open nesting if QB later has "Total for <this name>"
            if will_nest:
                parent_stack.append(name)
        else:
            # Header with no $ — only nest if a matching Total-for closes it
            # (avoids orphan headers like "KCG Adjustment" swallowing Morgan)
            if will_nest:
                parent_stack.append(name)

    by_key = {}
    for c in customers:
        if c.get('job') and c.get('parent'):
            key = f"{c['parent'].lower()}::{c['job'].lower()}"
        else:
            key = f"total::{c['customer'].strip().lower()}"
        prev = by_key.get(key)
        if not prev or c['total'] >= prev['total']:
            by_key[key] = c

    parents_with_jobs = {
        c['parent'].lower()
        for c in by_key.values()
        if c.get('job') and c.get('parent')
    }
    filtered = []
    for c in by_key.values():
        if c.get('is_parent_total') and c['customer'].lower() in parents_with_jobs:
            continue
        if (
            not c.get('job')
            and not c.get('is_parent_total')
            and c.get('parent')
            and c['parent'].lower() in parents_with_jobs
            and c['customer'].lower() == c['parent'].lower()
        ):
            continue
        filtered.append(c)
    customers = sorted(filtered, key=lambda x: -x['total'])
    out = _build_summary(customers, as_of_label=as_of_label, source='xlsx')
    out['report'] = 'A/R Aging Summary'
    out['report_kind'] = 'summary'
    return out


def _bucket_from_section_label(label):
    n = _norm_header(label)
    if not n:
        return None
    if n.startswith('total for'):
        return None
    if '91' in n or 'or more' in n:
        return '91_and_over'
    if re.search(r'61\s*[-–—]\s*90', n):
        return '61_90'
    if re.search(r'31\s*[-–—]\s*60', n):
        return '31_60'
    if re.search(r'1\s*[-–—]\s*30', n):
        return '1_30'
    if n == 'current' or 'not yet due' in n:
        return 'current'
    return None


def _parse_ar_detail(filename, raw_bytes):
    """Parse QB A/R Aging Detail into customer rollup + invoice list."""
    rows = _load_tabular_rows(filename, raw_bytes)
    if not rows:
        raise ValueError('Detail spreadsheet is empty.')

    as_of_label = None
    for row in rows[:8]:
        for cell in row:
            if cell and isinstance(cell, str) and cell.strip().lower().startswith('as of'):
                as_of_label = cell.strip()
                break

    header_idx = None
    for i, row in enumerate(rows[:25]):
        if _is_detail_header_row(row):
            header_idx = i
            break
    if header_idx is None:
        for row in rows[:15]:
            if _is_summary_header_row(row):
                raise ValueError(
                    'This is an A/R Aging Summary export, not Detail. '
                    'Use the “Aging Summary” upload box for totals.'
                )
        raise ValueError(
            'Could not find an A/R Aging Detail header '
            '(expected Date / Transaction type / Customer / Open balance). '
            'In QuickBooks: Reports → Who owes you → A/R Aging Detail → Export to Excel.'
        )

    header = rows[header_idx]
    # Map columns (Detail often has a blank leading column)
    col = {}
    for i, h in enumerate(header):
        n = _norm_header(h)
        if not n:
            continue
        if n == 'date' or n.endswith(' date') and 'due' not in n:
            col.setdefault('date', i)
        elif 'transaction' in n or n == 'type':
            col['type'] = i
        elif n in ('num', 'no.', 'number', '#') or n == 'num':
            col['num'] = i
        elif 'customer' in n:
            col['customer'] = i
        elif 'due' in n:
            col['due'] = i
        elif n == 'amount':
            col['amount'] = i
        elif 'open' in n and 'balance' in n:
            col['open'] = i
        elif 'sales' in n or 'rep' in n or n in ('salesman', 'salesperson', 'sales rep'):
            col['salesman'] = i
        elif 'memo' in n or 'description' in n or 'product' in n or 'service' in n:
            col.setdefault('memo', i)

    if 'customer' not in col:
        raise ValueError('Detail export missing Customer column.')
    if 'open' not in col and 'amount' not in col:
        raise ValueError('Detail export missing Open balance / Amount column.')

    def cell(row, key):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    current_bucket = 'current'
    invoices = []
    by_customer = {}

    for row in rows[header_idx + 1:]:
        if not row or all(v in (None, '') for v in row):
            continue
        # Section labels live in first non-empty cell
        first = next((str(v).strip() for v in row if v not in (None, '')), '')
        if first.upper() == 'TOTAL' or first.lower().startswith('total for'):
            # Prefer TOTAL open-balance if present
            continue
        bucket_hit = _bucket_from_section_label(first)
        if bucket_hit and cell(row, 'type') in (None, ''):
            current_bucket = bucket_hit
            continue

        txn = cell(row, 'type')
        if txn is None:
            continue
        txn_s = str(txn).strip()
        if txn_s.lower() not in ('invoice', 'credit memo', 'payment', 'journal entry', 'deposit', 'cheque', 'check', 'sales receipt'):
            # Still allow rows that look like invoices via open balance + customer
            if not cell(row, 'customer'):
                continue
        cust_raw = cell(row, 'customer')
        if not cust_raw:
            continue
        cust = str(cust_raw).strip()
        open_bal = _money(cell(row, 'open') if col.get('open') is not None else cell(row, 'amount'))
        if open_bal == 0 and str(txn_s).lower() != 'invoice':
            continue
        # Credit memos reduce AR
        if str(txn_s).lower() == 'credit memo' and open_bal > 0:
            open_bal = -open_bal

        inv = {
            'date': str(cell(row, 'date') or ''),
            'type': txn_s,
            'num': str(cell(row, 'num') or ''),
            'customer': cust,
            'due_date': str(cell(row, 'due') or ''),
            'amount': _money(cell(row, 'amount')),
            'open_balance': open_bal,
            'age_bucket': current_bucket,
            'salesman': str(cell(row, 'salesman') or '').strip() or None,
            'memo': str(cell(row, 'memo') or '').strip() or None,
        }
        # Heuristic: salesman on a free-text line (e.g. "Sales: Andy / Adam")
        if not inv['salesman'] and inv['memo']:
            m = re.search(
                r'(?:sales(?:man|person| rep)?|psc|consultant)\s*[:\-]\s*(.+)$',
                inv['memo'],
                re.I,
            )
            if m:
                inv['salesman'] = m.group(1).strip()[:120]

        if str(txn_s).lower() != 'invoice' and open_bal == 0:
            continue
        invoices.append(inv)

        # Parent customer before ":" for rollups
        parent = cust.split(':', 1)[0].strip()
        slot = by_customer.setdefault(parent, {
            'customer': parent,
            'current': 0.0,
            '1_30': 0.0,
            '31_60': 0.0,
            '61_90': 0.0,
            '91_and_over': 0.0,
            'total': 0.0,
            'invoice_count': 0,
        })
        b = current_bucket if current_bucket in slot else 'current'
        slot[b] = slot.get(b, 0.0) + open_bal
        slot['total'] += open_bal
        slot['invoice_count'] += 1

    customers = sorted(by_customer.values(), key=lambda x: -x['total'])
    out = _build_summary(customers, as_of_label=as_of_label, source='detail_xlsx')
    out['report'] = 'A/R Aging Detail'
    out['report_kind'] = 'detail'
    out['invoices'] = invoices[:500]  # cap payload
    out['invoice_count'] = len(invoices)
    out['invoices_truncated'] = max(0, len(invoices) - 500)
    # Salesman coverage
    with_sales = sum(1 for i in invoices if i.get('salesman'))
    out['salesman_field_present'] = with_sales > 0
    out['salesman_invoice_count'] = with_sales
    if with_sales == 0:
        out['notes_for_humans'] = list(out.get('notes_for_humans') or []) + [
            'No salesman/rep column found on this Detail export. '
            'If salesmen are stored as a custom field or invoice line in QB, '
            'export that field (or we parse a “Sales: Name” description line once the export includes it).',
        ]
    return out


def _build_summary(customers, as_of_label, source):
    grand = {
        'current': 0.0,
        '1_30': 0.0,
        '31_60': 0.0,
        '61_90': 0.0,
        '91_and_over': 0.0,
        'total': 0.0,
    }
    for c in customers:
        for k in grand:
            grand[k] += c.get(k, 0.0)

    def _is_bopc_name(name):
        low = (name or '').lower()
        if any(a in low for a in BOPC_ALIASES):
            return True
        # QB often splits Bridges into "BOPC Turns", "BOPC Capital", etc.
        return low.startswith('bopc') or low.startswith('bridges of pine')

    # Roll up all BOPC / Bridges lines (job-level Total-for rows in QB exports).
    bopc_parts = [c for c in customers if _is_bopc_name(c['customer'])]
    bopc = None
    if bopc_parts:
        bopc = {
            'customer': 'BOPC / Bridges of Pine Creek (rolled up)',
            'parts': [p['customer'] for p in bopc_parts],
            'current': sum(p['current'] for p in bopc_parts),
            '1_30': sum(p['1_30'] for p in bopc_parts),
            '31_60': sum(p['31_60'] for p in bopc_parts),
            '61_90': sum(p['61_90'] for p in bopc_parts),
            '91_and_over': sum(p['91_and_over'] for p in bopc_parts),
            'total': sum(p['total'] for p in bopc_parts),
        }

    # Operating AR: exclude BOPC full balance (common Stephanie/Thomas view)
    operating = dict(grand)
    if bopc:
        for k in ('current', '1_30', '31_60', '61_90', '91_and_over', 'total'):
            operating[k] = max(0.0, operating[k] - bopc.get(k, 0.0))

    # Chase list: prioritize 61+ and 91+, then 31-60, then 1-30 — exclude pure current
    chase = []
    for c in customers:
        overdue = c['1_30'] + c['31_60'] + c['61_90'] + c['91_and_over']
        if overdue <= 0:
            continue
        weight = (
            c['91_and_over'] * 4
            + c['61_90'] * 3
            + c['31_60'] * 2
            + c['1_30']
        )
        is_legacy = _is_bopc_name(c['customer']) or any(
            h in c['customer'].lower() for h in LEGACY_CUSTOMER_HINTS
        )
        chase.append({
            **c,
            'overdue': overdue,
            'weight': weight,
            'is_legacy_or_bopc': is_legacy,
        })
    chase.sort(key=lambda x: (-x['weight'], -x['overdue']))

    top_by_balance = customers[:15]
    top_chase = chase[:20]

    numbers_draft = _numbers_draft_md(
        as_of_label=as_of_label,
        grand=grand,
        operating=operating,
        bopc=bopc,
        top_chase=top_chase,
    )

    return {
        'report': 'A/R Aging Summary',
        'as_of_label': as_of_label,
        'source_format': source,
        'parsed_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        'customer_count': len(customers),
        'grand_total': grand,
        'operating_ex_bopc': operating,
        'bopc': bopc,
        'all_customers': customers,
        'top_customers_by_balance': top_by_balance,
        'chase_list': top_chase,
        'notes_for_humans': [
            'BOPC / Bridges is split out for operating AR — confirm with Stephanie before quoting totals externally.',
            '50/50 split jobs are not auto-detected yet — flag those in Notes when you know them; we will learn the pattern.',
            'Draft is for edit/send by Stephanie — Hub does not email this automatically.',
        ],
        'numbers_draft_md': numbers_draft,
    }


def _fmt_money(n):
    try:
        v = float(n or 0)
    except (TypeError, ValueError):
        v = 0.0
    return f'${v:,.0f}'


def _numbers_draft_md(as_of_label, grand, operating, bopc, top_chase, sales_rep_open_ar=None):
    as_of = as_of_label or 'this week'
    lines = [
        f'**Numbers draft** — AR snapshot ({as_of})',
        '',
        '_Edit freely. Sales vs goal still comes from your Monthly Outlook sheet — paste that block above or below._',
        '',
        '### AR snapshot',
        f'- **Total AR (all customers):** {_fmt_money(grand["total"])}',
        f'- **Current:** {_fmt_money(grand["current"])} · **1–30:** {_fmt_money(grand["1_30"])} · '
        f'**31–60:** {_fmt_money(grand["31_60"])} · **61–90:** {_fmt_money(grand["61_90"])} · '
        f'**91+:** {_fmt_money(grand["91_and_over"])}',
    ]
    if bopc:
        lines.append(
            f'- **BOPC / Bridges (included above):** {_fmt_money(bopc["total"])} '
            f'(91+: {_fmt_money(bopc["91_and_over"])})'
        )
        lines.append(
            f'- **Operating AR (ex-BOPC, rough):** {_fmt_money(operating["total"])}'
        )
    # A/R is company total only — Stephanie owns collections (not by rep).
    lines.extend(['', '### Collection focus (overdue-weighted)', ''])
    if not top_chase:
        lines.append('_No overdue balances detected in this export._')
    else:
        for i, c in enumerate(top_chase[:12], 1):
            tag = ' _(legacy/BOPC — handle with care)_' if c.get('is_legacy_or_bopc') else ''
            lines.append(
                f'{i}. **{c["customer"]}** — total {_fmt_money(c["total"])}, '
                f'overdue {_fmt_money(c["overdue"])} '
                f'(91+ {_fmt_money(c["91_and_over"])}){tag}'
            )
    lines.extend([
        '',
        '### Notes / 50–50 splits',
        '- Multi-rep Sales Rep values (e.g. Adam / Andy) are treated as equal shares of open balance.',
        '- _Override here if a job is not equal 50/50._',
        '',
        '—',
        'Generated from Office Ops · not sent automatically',
    ])
    return '\n'.join(lines)


def save_upload(get_db_fn, kind, filename, mime_type, raw_bytes, user_key):
    if kind not in ALLOWED_KINDS:
        return {'success': False, 'error': f'Unknown upload type: {kind}'}
    if not raw_bytes:
        return {'success': False, 'error': 'Empty file.'}
    if len(raw_bytes) > MAX_UPLOAD_BYTES:
        return {'success': False, 'error': f'File too large (max {MAX_UPLOAD_BYTES // (1024*1024)}MB).'}
    conn = get_db_fn()
    if not conn:
        return {'success': False, 'error': 'Database unavailable.'}
    try:
        cur = conn.cursor()
        cur.execute(
            '''
            INSERT INTO office_ops_files
                (kind, filename, mime_type, size_bytes, file_data, uploaded_by)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id, uploaded_at
            ''',
            (
                kind,
                (filename or 'upload')[:255],
                (mime_type or '')[:100],
                len(raw_bytes),
                psycopg2_binary(raw_bytes),
                user_key,
            ),
        )
        row = cur.fetchone()
        file_id = row[0]
        uploaded_at = row[1]
        conn.commit()
        cur.close()
        conn.close()
        return {
            'success': True,
            'file_id': file_id,
            'uploaded_at': uploaded_at.isoformat() if uploaded_at else None,
            'filename': filename,
            'size_bytes': len(raw_bytes),
        }
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        print(f'Office Ops upload error: {e}')
        return {'success': False, 'error': 'Could not save upload.'}


def psycopg2_binary(raw_bytes):
    """Wrap bytes for BYTEA insert without importing Binary at module top if unused."""
    from psycopg2 import Binary
    return Binary(raw_bytes)


def _attach_invoice_samples(out):
    """Attach sample open invoices onto chase rows (match parent or job name)."""
    inv_by_key = {}
    for inv in out.get('invoices') or []:
        full = (inv.get('customer') or '').strip()
        parent = full.split(':', 1)[0].strip()
        job = full.split(':', 1)[1].strip() if ':' in full else ''
        for key in filter(None, (full.lower(), parent.lower(), job.lower())):
            inv_by_key.setdefault(key, []).append(inv)
    for c in out.get('chase_list') or []:
        key = (c.get('customer') or '').strip().lower()
        samples = inv_by_key.get(key, [])[:5]
        if not samples:
            samples = [
                i for i in (out.get('invoices') or [])
                if key and key in (i.get('customer') or '').lower()
            ][:5]
        c['sample_invoices'] = [
            {
                'num': i.get('num'),
                'open_balance': i.get('open_balance'),
                'age_bucket': i.get('age_bucket'),
                'due_date': i.get('due_date'),
                'salesman': i.get('salesman') or i.get('sales_rep_raw'),
            }
            for i in samples
        ]


def _apply_sales_rep_map(out):
    """Stamp sales_rep from Invoice List onto Aging Detail invoices by invoice #."""
    ilist = out.get('invoice_list') or []
    if not ilist:
        return
    by_num = {}
    for inv in ilist:
        num = str(inv.get('num') or '').strip()
        if num:
            by_num[num] = inv
    matched = 0
    for inv in out.get('invoices') or []:
        num = str(inv.get('num') or '').strip()
        src = by_num.get(num)
        if not src:
            continue
        inv['salesman'] = src.get('sales_rep_raw')
        inv['sales_reps'] = src.get('sales_reps') or []
        inv['is_50_50_style'] = src.get('is_50_50_style', False)
        matched += 1
    out['sales_rep_matched_invoices'] = matched
    out['salesman_field_present'] = True
    # Rebuild rep open-AR from aging open balances when possible
    if out.get('invoices'):
        by_rep = {}
        for inv in out['invoices']:
            open_bal = inv.get('open_balance') or 0
            if open_bal <= 0:
                continue
            reps = inv.get('sales_reps') or (
                _parse_sales_reps(inv.get('salesman')) if inv.get('salesman') else ['(unassigned)']
            )
            if not reps:
                reps = ['(unassigned)']
            share = open_bal / len(reps)
            for r in reps:
                slot = by_rep.setdefault(r, {
                    'sales_rep': r,
                    'open_ar': 0.0,
                    'invoice_count': 0,
                    'split_invoice_count': 0,
                })
                slot['open_ar'] += share
                slot['invoice_count'] += 1
                if inv.get('is_50_50_style') or len(reps) >= 2:
                    slot['split_invoice_count'] += 1
        out['sales_rep_open_ar'] = sorted(by_rep.values(), key=lambda x: -x['open_ar'])


def _merge_pack_payload(existing, incoming):
    """Combine Summary + Detail + Invoice List fragments into one pack."""
    if not existing:
        out = dict(incoming)
        out['sources'] = {
            'summary': out.get('summary_source_filename') or (
                out.get('source_filename') if out.get('report_kind') == 'summary' else None
            ),
            'detail': out.get('detail_source_filename') or (
                out.get('source_filename') if out.get('report_kind') == 'detail' else None
            ),
            'invoice_list': out.get('invoice_list_source_filename') or (
                out.get('source_filename') if out.get('report_kind') == 'invoice_list' else None
            ),
        }
        if out.get('report_kind') == 'summary':
            out['summary_source_filename'] = out.get('source_filename')
        if out.get('report_kind') == 'detail':
            out['detail_source_filename'] = out.get('source_filename')
        if out.get('report_kind') == 'invoice_list':
            out['invoice_list_source_filename'] = out.get('source_filename')
        if out.get('chase_list') and out.get('invoices'):
            _attach_invoice_samples(out)
        if out.get('invoice_list'):
            _apply_sales_rep_map(out)
        out['parsed_at'] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
        return out
    if not incoming:
        return existing

    out = dict(existing)
    inc_kind = incoming.get('report_kind')

    if inc_kind == 'summary' or (incoming.get('report') or '').lower().find('summary') >= 0:
        for k in (
            'grand_total', 'operating_ex_bopc', 'bopc', 'top_customers_by_balance',
            'chase_list', 'customer_count', 'as_of_label', 'numbers_draft_md',
            'notes_for_humans', 'source_format', 'report',
        ):
            if k in incoming:
                out[k] = incoming[k]
        out['report_kind'] = 'summary'
        out['summary_source_filename'] = incoming.get('source_filename')
        out['summary_source_file_id'] = incoming.get('source_file_id')

    if inc_kind == 'detail' or (
        incoming.get('invoices') is not None and inc_kind != 'invoice_list'
    ):
        # Only replace aging invoices from Detail (not from invoice list)
        if inc_kind == 'detail' or incoming.get('report') == 'A/R Aging Detail':
            out['invoices'] = incoming.get('invoices') or []
            out['invoice_count'] = incoming.get('invoice_count', len(out['invoices']))
            out['invoices_truncated'] = incoming.get('invoices_truncated', 0)
            out['detail_source_filename'] = incoming.get('source_filename')
            out['detail_source_file_id'] = incoming.get('source_file_id')
            out['detail_as_of_label'] = incoming.get('as_of_label')
            if not out.get('grand_total'):
                for k in (
                    'grand_total', 'operating_ex_bopc', 'bopc', 'top_customers_by_balance',
                    'chase_list', 'customer_count', 'as_of_label', 'numbers_draft_md',
                    'notes_for_humans', 'report',
                ):
                    if k in incoming:
                        out[k] = incoming[k]
                out['report_kind'] = 'detail'

    if inc_kind == 'invoice_list':
        out['invoice_list'] = incoming.get('invoice_list') or []
        out['invoice_list_count'] = incoming.get('invoice_list_count', len(out['invoice_list']))
        out['open_invoice_count'] = incoming.get('open_invoice_count')
        out['open_ar_from_list'] = incoming.get('open_ar_from_list')
        out['split_invoice_count'] = incoming.get('split_invoice_count')
        out['sales_rep_open_ar'] = incoming.get('sales_rep_open_ar')
        out['salesman_field_present'] = True
        out['invoice_list_source_filename'] = incoming.get('source_filename')
        out['invoice_list_source_file_id'] = incoming.get('source_file_id')
        out['invoice_list_date_range'] = incoming.get('date_range')
        notes = list(out.get('notes_for_humans') or [])
        for n in incoming.get('notes_for_humans') or []:
            if n not in notes:
                notes.append(n)
        out['notes_for_humans'] = notes

    if out.get('chase_list') and out.get('invoices'):
        _attach_invoice_samples(out)
    if out.get('invoice_list'):
        _apply_sales_rep_map(out)
        # After map, re-attach samples so salesman shows on chase
        if out.get('chase_list') and out.get('invoices'):
            _attach_invoice_samples(out)

    if out.get('grand_total'):
        out['numbers_draft_md'] = _numbers_draft_md(
            as_of_label=out.get('as_of_label'),
            grand=out['grand_total'],
            operating=out.get('operating_ex_bopc') or out['grand_total'],
            bopc=out.get('bopc'),
            top_chase=out.get('chase_list') or [],
            sales_rep_open_ar=out.get('sales_rep_open_ar'),
        )

    out['parsed_at'] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
    out['sources'] = {
        'summary': out.get('summary_source_filename'),
        'detail': out.get('detail_source_filename'),
        'invoice_list': out.get('invoice_list_source_filename'),
    }
    return out


def process_ar_file(get_db_fn, file_id, user_key, expect=None):
    conn = get_db_fn()
    if not conn:
        return {'success': False, 'error': 'Database unavailable.'}
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            'SELECT id, kind, filename, file_data FROM office_ops_files WHERE id = %s',
            (file_id,),
        )
        frow = cur.fetchone()
        if not frow:
            cur.close()
            conn.close()
            return {'success': False, 'error': 'Upload not found.'}
        if frow['kind'] not in ALLOWED_KINDS:
            cur.close()
            conn.close()
            return {'success': False, 'error': 'Not an AR aging file.'}

        # Prefer explicit kind from upload slot
        kind_hint = frow['kind']
        if expect is None:
            if kind_hint == KIND_DETAIL:
                expect = 'detail'
            elif kind_hint == KIND_INVOICE_LIST:
                expect = 'invoice_list'
            elif kind_hint in (KIND_SUMMARY, 'ar_aging'):
                expect = 'summary'

        raw = bytes(frow['file_data'])
        # Auto-correct kind if content disagrees (user used wrong box).
        # Invoice List without Sales Rep header-detects as aging detail — keep
        # the Invoice List slot unless the file is clearly an aging report.
        detected = detect_ar_report_type(frow['filename'], raw)
        if detected and expect and detected != expect:
            title = _title_blob(_load_tabular_rows(frow['filename'], raw, max_rows=8))
            fname = (frow['filename'] or '').lower()
            clearly_aging = 'aging' in title or ('aging' in fname and 'detail' in fname)
            if expect == 'invoice_list' and detected == 'detail' and not clearly_aging:
                detected = 'invoice_list'
            else:
                expect = detected  # trust content; still process successfully

        parsed = parse_ar_aging_bytes(frow['filename'], raw, expect=expect)
        parsed['source_filename'] = frow['filename']
        parsed['source_file_id'] = file_id

        # Merge with latest combined pack if present
        cur.execute(
            '''
            SELECT summary_json FROM office_ops_packs
            WHERE kind = %s
            ORDER BY created_at DESC LIMIT 1
            ''',
            (PACK_KIND,),
        )
        prev = cur.fetchone()
        existing = None
        if prev and prev.get('summary_json'):
            existing = prev['summary_json']
            if isinstance(existing, str):
                existing = json.loads(existing)
        merged = _merge_pack_payload(existing, parsed)

        pack_date = date.today()
        cur.execute(
            '''
            INSERT INTO office_ops_packs
                (pack_date, source_file_id, kind, summary_json, numbers_draft_md, created_by)
            VALUES (%s, %s, %s, %s::jsonb, %s, %s)
            RETURNING id, created_at
            ''',
            (
                pack_date,
                file_id,
                PACK_KIND,
                json.dumps(merged),
                merged.get('numbers_draft_md'),
                user_key,
            ),
        )
        prow = cur.fetchone()
        # Fix stored file kind if auto-detected differently
        if detected:
            fixed = {
                'detail': KIND_DETAIL,
                'summary': KIND_SUMMARY,
                'invoice_list': KIND_INVOICE_LIST,
            }.get(detected, KIND_SUMMARY)
            if frow['kind'] != fixed:
                cur.execute(
                    'UPDATE office_ops_files SET kind = %s WHERE id = %s',
                    (fixed, file_id),
                )
        conn.commit()
        cur.close()
        conn.close()
        return {
            'success': True,
            'pack_id': prow['id'],
            'created_at': prow['created_at'].isoformat() if prow['created_at'] else None,
            'summary': merged,
            'detected': detected or expect,
        }
    except ValueError as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        return {'success': False, 'error': str(e)}
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        print(f'Office Ops process error: {e}')
        return {
            'success': False,
            'error': (
                'Could not process that AR file. '
                'Use A/R Aging Summary for totals and A/R Aging Detail for invoices.'
            ),
        }


def get_latest_pack(get_db_fn, kind=None):
    kind = kind or PACK_KIND
    conn = get_db_fn()
    if not conn:
        return None
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            '''
            SELECT p.id, p.pack_date, p.kind, p.summary_json, p.numbers_draft_md,
                   p.created_by, p.created_at, p.source_file_id,
                   f.filename AS source_filename
            FROM office_ops_packs p
            LEFT JOIN office_ops_files f ON f.id = p.source_file_id
            WHERE p.kind = %s
            ORDER BY p.created_at DESC
            LIMIT 1
            ''',
            (kind,),
        )
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return None
        summary = row['summary_json']
        if isinstance(summary, str):
            summary = json.loads(summary)
        return {
            'id': row['id'],
            'pack_date': row['pack_date'].isoformat() if row['pack_date'] else None,
            'kind': row['kind'],
            'summary': summary,
            'numbers_draft_md': row['numbers_draft_md'] or (summary or {}).get('numbers_draft_md'),
            'created_by': row['created_by'],
            'created_at': row['created_at'].isoformat() if row['created_at'] else None,
            'source_file_id': row['source_file_id'],
            'source_filename': row['source_filename'],
        }
    except Exception as e:
        print(f'Office Ops latest pack error: {e}')
        try:
            conn.close()
        except Exception:
            pass
        return None


def _is_bridges_customer_name(name):
    low = (name or '').lower()
    return any(
        x in low
        for x in (
            'bopc', 'bridges', 'pine creek', 'pebble', 'meadow',
            'bridges of pine',
        )
    ) or low.startswith('bopc')


def build_past_due_prompt_rows(ar_summary, notes_dict=None, limit=15):
    """Customer-level past-due rows for the notes modal — NOT per invoice.

    - One modal, up to `limit` rows (default 15).
    - All Bridges/BOPC/Pebble/Pine-Meadow job lines collapse to a single
      "Bridges / BOPC" row (Stephanie's email style), not one box per sub-job.
    - Source is overdue-weighted chase list / customers with overdue > 0.
    """
    notes_dict = notes_dict or {}
    chase = list(ar_summary.get('chase_list') or [])
    if not chase and ar_summary.get('all_customers'):
        # Build minimal overdue rows from all_customers
        for c in ar_summary['all_customers']:
            overdue = (
                float(c.get('1_30') or 0)
                + float(c.get('31_60') or 0)
                + float(c.get('61_90') or 0)
                + float(c.get('91_and_over') or 0)
            )
            if overdue > 0:
                chase.append({**c, 'overdue': overdue})
        chase.sort(key=lambda x: -float(x.get('overdue') or 0))

    bridges_items = []
    other = []
    for c in chase:
        name = c.get('customer') or ''
        if _is_bridges_customer_name(name) or _is_bridges_customer_name(c.get('parent') or ''):
            bridges_items.append(c)
        else:
            other.append(c)

    rows = []
    if bridges_items:
        # Prefer a single parent total (Bridges of Pine Creek) when present so we
        # do not double-count parent Total-for + job lines.
        parent_totals = [
            c for c in bridges_items
            if not c.get('job') and 'bridges of pine' in (c.get('customer') or '').lower()
        ]
        if parent_totals:
            bridges_total = max(float(c.get('total') or 0) for c in parent_totals)
            bridges_overdue = max(float(c.get('overdue') or 0) for c in parent_totals)
            bridges_parts = [c.get('customer') for c in bridges_items if c.get('job')]
        else:
            bridges_total = sum(float(c.get('total') or 0) for c in bridges_items)
            bridges_overdue = sum(float(c.get('overdue') or 0) for c in bridges_items)
            bridges_parts = [c.get('customer') for c in bridges_items]
        note = ''
        for key in ('bridges / bopc', 'bridges of pine creek', 'bopc'):
            if notes_dict.get(key) and notes_dict[key].get('note'):
                note = notes_dict[key]['note']
                break
        if not note:
            for k, v in notes_dict.items():
                if _is_bridges_customer_name(k) and v.get('note'):
                    note = v['note']
                    break
        rows.append({
            'customer': 'Bridges / BOPC',
            'total': bridges_total,
            'overdue': bridges_overdue,
            'note': note,
            'rolled_up_from': bridges_parts,
            'is_bridges_rollup': True,
        })

    for c in other:
        if len(rows) >= limit:
            break
        display = c.get('customer') or ''
        key = display.lower()
        rows.append({
            'customer': display,
            'total': float(c.get('total') or 0),
            'overdue': float(c.get('overdue') or 0),
            'note': (notes_dict.get(key) or {}).get('note', ''),
            'is_bridges_rollup': False,
        })

    return rows[:limit]


def get_ar_notes(get_db_fn):
    conn = get_db_fn()
    if not conn:
        return {}
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            'SELECT customer_key, customer_display, note, updated_by, updated_at '
            'FROM office_ops_ar_notes ORDER BY customer_display'
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
        out = {}
        for r in rows:
            out[r['customer_key']] = {
                'customer': r['customer_display'],
                'note': r['note'] or '',
                'updated_by': r['updated_by'],
                'updated_at': r['updated_at'].isoformat() if r['updated_at'] else None,
            }
        return out
    except Exception as e:
        print(f'Office Ops notes load error: {e}')
        try:
            conn.close()
        except Exception:
            pass
        return {}


def save_ar_notes(get_db_fn, notes_map, user_key):
    """notes_map: {customer_display: note_text}"""
    conn = get_db_fn()
    if not conn:
        return {'success': False, 'error': 'Database unavailable.'}
    try:
        cur = conn.cursor()
        for display, note in (notes_map or {}).items():
            display = (display or '').strip()
            if not display:
                continue
            key = display.lower()
            note = (note or '').strip()
            if not note:
                cur.execute('DELETE FROM office_ops_ar_notes WHERE customer_key = %s', (key,))
            else:
                cur.execute(
                    '''
                    INSERT INTO office_ops_ar_notes
                        (customer_key, customer_display, note, updated_by, updated_at)
                    VALUES (%s, %s, %s, %s, NOW())
                    ON CONFLICT (customer_key) DO UPDATE SET
                        customer_display = EXCLUDED.customer_display,
                        note = EXCLUDED.note,
                        updated_by = EXCLUDED.updated_by,
                        updated_at = NOW()
                    ''',
                    (key, display, note, user_key),
                )
        conn.commit()
        cur.close()
        conn.close()
        return {'success': True}
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        print(f'Office Ops notes save error: {e}')
        return {'success': False, 'error': str(e)}


def generate_thursday_pack(get_db_fn, user_key):
    """Build Monday/Thursday Excel from latest Invoice List + AR Summary + notes.

    Does NOT require Stephanie to maintain Monthly Outlook — goals from template.
    """
    from office_ops_generate import generate_from_qb

    conn = get_db_fn()
    if not conn:
        return {'success': False, 'error': 'Database unavailable.'}
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        # Latest invoice list file
        cur.execute(
            '''
            SELECT id, filename, file_data FROM office_ops_files
            WHERE kind = %s ORDER BY uploaded_at DESC LIMIT 1
            ''',
            (KIND_INVOICE_LIST,),
        )
        inv_row = cur.fetchone()
        if not inv_row:
            cur.close()
            conn.close()
            return {
                'success': False,
                'error': 'Upload Invoice List by Date first.',
            }
        inv_raw = bytes(inv_row['file_data'])
        inv_parsed = parse_ar_aging_bytes(inv_row['filename'], inv_raw, expect='invoice_list')

        # Latest AR pack for totals
        cur.execute(
            '''
            SELECT summary_json FROM office_ops_packs
            WHERE kind = %s ORDER BY created_at DESC LIMIT 1
            ''',
            (PACK_KIND,),
        )
        ar_row = cur.fetchone()
        ar_summary = None
        if ar_row and ar_row.get('summary_json'):
            ar_summary = ar_row['summary_json']
            if isinstance(ar_summary, str):
                ar_summary = json.loads(ar_summary)

        # Optional P&L for margin/profit YoY
        cur.execute(
            '''
            SELECT filename, file_data FROM office_ops_files
            WHERE kind = %s ORDER BY uploaded_at DESC LIMIT 1
            ''',
            (KIND_PL,),
        )
        pl_row = cur.fetchone()
        pl_summary = None
        if pl_row:
            try:
                pl_summary = parse_pl_bytes(pl_row['filename'], bytes(pl_row['file_data']))
            except Exception as e:
                print(f'P&L parse skipped: {e}')

        cur.close()
        conn.close()

        notes = get_ar_notes(get_db_fn)
        # Attach past-due $ to each saved comment so Insights/AR show importance
        past_rows = build_past_due_prompt_rows(ar_summary or {}, notes, limit=50)
        overdue_by_name = {
            (r.get('customer') or '').lower().strip(): r for r in past_rows
        }
        notes_by_customer = {}
        for v in notes.values():
            if not v.get('note'):
                continue
            name = v.get('customer') or ''
            key = name.lower().strip()
            row = overdue_by_name.get(key)
            if not row:
                # fuzzy: Bridges rollup or partial name match
                for ok, r in overdue_by_name.items():
                    if key == ok or key in ok or ok in key:
                        row = r
                        break
                    if _is_bridges_customer_name(name) and r.get('is_bridges_rollup'):
                        row = r
                        break
            notes_by_customer[name] = {
                'note': v['note'],
                'overdue': float(row['overdue']) if row and row.get('overdue') is not None else None,
                'total': float(row['total']) if row and row.get('total') is not None else None,
            }

        report_bytes, insights, meta = generate_from_qb(
            inv_parsed.get('invoice_list') or [],
            ar_summary=ar_summary,
            notes_by_customer=notes_by_customer,
            pl_summary=pl_summary,
            year=2026,
        )

        # Store generated file
        conn = get_db_fn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        fname = f"PPS_Monday_Numbers_{date.today().isoformat()}.xlsx"
        cur.execute(
            '''
            INSERT INTO office_ops_files
                (kind, filename, mime_type, size_bytes, file_data, uploaded_by)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id
            ''',
            (
                KIND_MONDAY,
                fname,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                len(report_bytes),
                psycopg2_binary(report_bytes),
                user_key,
            ),
        )
        monday_id = cur.fetchone()['id']
        pack_payload = {
            'report': 'Monday Numbers (generated from QB)',
            'report_kind': 'monday_report',
            'insights_md': insights,
            'source_invoice_list': inv_row['filename'],
            'monday_file_id': monday_id,
            'ar_total': (ar_summary or {}).get('grand_total'),
            'bopc': (ar_summary or {}).get('bopc'),
            'bridges_lines': (ar_summary or {}).get('bridges_lines'),
            'meta': meta,
            'notes_count': len(notes_by_customer),
            'parsed_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
            'recipients_note': (
                'Email to Thomas + Tony; if sent from system/admin, include Stephanie.'
            ),
        }
        cur.execute(
            '''
            INSERT INTO office_ops_packs
                (pack_date, source_file_id, kind, summary_json, numbers_draft_md, created_by)
            VALUES (%s, %s, %s, %s::jsonb, %s, %s)
            RETURNING id
            ''',
            (
                date.today(),
                monday_id,
                KIND_MONDAY,
                json.dumps(pack_payload),
                insights,
                user_key,
            ),
        )
        pack_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        return {
            'success': True,
            'monday_file_id': monday_id,
            'pack_id': pack_id,
            'insights_md': insights,
            'download_url': f'/api/office-ops/files/{monday_id}/download',
            'ar_included': bool(ar_summary and ar_summary.get('grand_total')),
            'invoice_source': inv_row['filename'],
        }
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        print(f'Thursday pack error: {e}')
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': f'Could not generate report: {e}'}


def process_monthly_outlook(get_db_fn, file_id, user_key):
    """Build Monday Excel from uploaded Monthly Outlook + latest AR totals."""
    from office_ops_monday import generate_monday_report

    conn = get_db_fn()
    if not conn:
        return {'success': False, 'error': 'Database unavailable.'}
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            'SELECT id, kind, filename, file_data FROM office_ops_files WHERE id = %s',
            (file_id,),
        )
        frow = cur.fetchone()
        if not frow:
            cur.close()
            conn.close()
            return {'success': False, 'error': 'Upload not found.'}
        raw = bytes(frow['file_data'])
        # Pull company AR totals from latest pack if present
        ar_summary = None
        cur.execute(
            '''
            SELECT summary_json FROM office_ops_packs
            WHERE kind = %s ORDER BY created_at DESC LIMIT 1
            ''',
            (PACK_KIND,),
        )
        prow = cur.fetchone()
        if prow and prow.get('summary_json'):
            ar_summary = prow['summary_json']
            if isinstance(ar_summary, str):
                ar_summary = json.loads(ar_summary)

        report_bytes, insights, meta = generate_monday_report(raw, ar_summary=ar_summary)
        # Store generated report as a file for download
        cur.execute(
            '''
            INSERT INTO office_ops_files
                (kind, filename, mime_type, size_bytes, file_data, uploaded_by)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id, uploaded_at
            ''',
            (
                KIND_MONDAY,
                f"PPS_Monday_Numbers_{date.today().isoformat()}.xlsx",
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                len(report_bytes),
                psycopg2_binary(report_bytes),
                user_key,
            ),
        )
        row = cur.fetchone()
        monday_id = row[0]
        # Light pack record for UI (insights text)
        pack_payload = {
            'report': 'Monday Numbers',
            'report_kind': 'monday_report',
            'insights_md': insights,
            'source_outlook': frow['filename'],
            'source_outlook_file_id': file_id,
            'monday_file_id': monday_id,
            'ar_total': (ar_summary or {}).get('grand_total'),
            'bopc': (ar_summary or {}).get('bopc'),
            'operating_ex_bopc': (ar_summary or {}).get('operating_ex_bopc'),
            'meta': meta,
            'parsed_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        }
        cur.execute(
            '''
            INSERT INTO office_ops_packs
                (pack_date, source_file_id, kind, summary_json, numbers_draft_md, created_by)
            VALUES (%s, %s, %s, %s::jsonb, %s, %s)
            RETURNING id, created_at
            ''',
            (
                date.today(),
                monday_id,
                KIND_MONDAY,
                json.dumps(pack_payload),
                insights,
                user_key,
            ),
        )
        pack_row = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return {
            'success': True,
            'monday_file_id': monday_id,
            'pack_id': pack_row['id'],
            'insights_md': insights,
            'download_url': f'/api/office-ops/files/{monday_id}/download',
            'ar_included': bool(ar_summary and ar_summary.get('grand_total')),
        }
    except Exception as e:
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        print(f'Office Ops Monday report error: {e}')
        return {'success': False, 'error': f'Could not build Monday report: {e}'}


def get_latest_monday_pack(get_db_fn):
    return get_latest_pack(get_db_fn, kind=KIND_MONDAY)


def get_file_bytes(get_db_fn, file_id):
    conn = get_db_fn()
    if not conn:
        return None
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            'SELECT id, kind, filename, mime_type, file_data FROM office_ops_files WHERE id = %s',
            (file_id,),
        )
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return None
        return {
            'id': row['id'],
            'kind': row['kind'],
            'filename': row['filename'],
            'mime_type': row['mime_type'] or 'application/octet-stream',
            'data': bytes(row['file_data']),
        }
    except Exception as e:
        print(f'Office Ops get file error: {e}')
        try:
            conn.close()
        except Exception:
            pass
        return None


def list_recent_files(get_db_fn, kind=None, limit=12):
    conn = get_db_fn()
    if not conn:
        return []
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        if kind:
            cur.execute(
                '''
                SELECT id, kind, filename, size_bytes, uploaded_by, uploaded_at
                FROM office_ops_files
                WHERE kind = %s
                ORDER BY uploaded_at DESC
                LIMIT %s
                ''',
                (kind, limit),
            )
        else:
            cur.execute(
                '''
                SELECT id, kind, filename, size_bytes, uploaded_by, uploaded_at
                FROM office_ops_files
                WHERE kind IN %s
                ORDER BY uploaded_at DESC
                LIMIT %s
                ''',
                (tuple(ALLOWED_KINDS), limit),
            )
        rows = cur.fetchall()
        cur.close()
        conn.close()
        out = []
        for r in rows:
            out.append({
                'id': r['id'],
                'kind': r['kind'],
                'filename': r['filename'],
                'size_bytes': r['size_bytes'],
                'uploaded_by': r['uploaded_by'],
                'uploaded_at': r['uploaded_at'].isoformat() if r['uploaded_at'] else None,
            })
        return out
    except Exception as e:
        print(f'Office Ops list files error: {e}')
        try:
            conn.close()
        except Exception:
            pass
        return []


def register_routes(app, get_db_fn, users, require_login, send_email_fn=None,
                     claude_api_key=None, claude_model=None):
    from io import BytesIO

    from flask import jsonify, redirect, render_template, request, send_file, session, url_for

    import insurance_compliance

    def _gate():
        user_key = session.get('user_key')
        if not can_access_office_ops(users, user_key):
            return redirect(url_for('dashboard'))
        return None

    @app.route('/office-ops')
    @require_login
    def office_ops_page():
        blocked = _gate()
        if blocked:
            return blocked
        return render_template(
            'office_ops_landing.html',
            user_display=(users.get(session.get('user_key')) or {}).get('display', session.get('user_key')),
        )

    @app.route('/office-ops/numbers')
    @require_login
    def office_ops_numbers_page():
        blocked = _gate()
        if blocked:
            return blocked
        user_key = session.get('user_key')
        from hub_usage import record_usage
        record_usage(get_db_fn, user_key, 'office_ops', 'open', 'Numbers page')
        pack = get_latest_pack(get_db_fn)
        monday = get_latest_monday_pack(get_db_fn)
        files = list_recent_files(get_db_fn)
        notes = get_ar_notes(get_db_fn)
        # ONE modal, customer-level rows only (never per invoice). Bridges/BOPC
        # rolled into a single line so Stephanie isn't prompted 5+ times.
        past_due = build_past_due_prompt_rows(
            (pack or {}).get('summary') or {},
            notes,
            limit=15,
        )
        return render_template(
            'office_ops.html',
            user_key=user_key,
            user_display=(users.get(user_key) or {}).get('display', user_key),
            pack=pack,
            monday=monday,
            recent_files=files,
            past_due=past_due,
            ar_notes=notes,
        )

    @app.route('/office-ops/compliance')
    @require_login
    def office_ops_compliance_page():
        blocked = _gate()
        if blocked:
            return blocked
        from hub_usage import record_usage
        record_usage(get_db_fn, session.get('user_key'), 'compliance', 'open', 'Compliance page')
        rows, last_run_at = insurance_compliance.get_latest_snapshot_rows(get_db_fn)
        from datetime import date as date_type
        cats = insurance_compliance.categorize_rows(rows, date_type.today())
        try:
            pay_request_result = insurance_compliance.check_pay_requests(get_db_fn)
        except Exception as e:
            print(f'Compliance page: Pay Request cross-check failed: {e}')
            pay_request_result = {'flagged': [], 'unmatched': []}
        return render_template(
            'office_ops_compliance.html',
            user_display=(users.get(session.get('user_key')) or {}).get('display', session.get('user_key')),
            checked=len(rows),
            last_run_at=last_run_at.isoformat() if last_run_at else None,
            expired=cats['expired'],
            soon=cats['soon'],
            later=cats['later'],
            new_subs=cats['new_subs'],
            mismatches=cats['mismatches'],
            needs_manual=cats['needs_manual'],
            pay_flagged=pay_request_result['flagged'],
            pay_unmatched=pay_request_result['unmatched'],
        )

    @app.route('/office-ops/compliance/refresh', methods=['POST'])
    @require_login
    def office_ops_compliance_refresh():
        blocked = _gate()
        if blocked:
            return jsonify({'success': False, 'error': 'Not allowed.'}), 403
        if send_email_fn is None:
            return jsonify({'success': False, 'error': 'Email sending not configured.'}), 500
        user_key = session.get('user_key')
        try:
            recipients = [users['stephanie_whetstone']['email'], users['thomas_ellison']['email']]
            result = insurance_compliance.run_weekly_compliance_check(get_db_fn, send_email_fn, recipients)
            from hub_usage import record_usage
            record_usage(
                get_db_fn, user_key, 'compliance', 'refresh',
                f"{result.get('checked', 0)} subs",
            )
            return jsonify({'success': True, **result})
        except Exception as e:
            print(f'Office Ops compliance refresh error ({user_key}): {e}')
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/office-ops/compliance/override', methods=['POST'])
    @require_login
    def office_ops_compliance_override():
        blocked = _gate()
        if blocked:
            return jsonify({'success': False, 'error': 'Not allowed.'}), 403
        user_key = session.get('user_key')
        data = request.get_json(silent=True) or {}
        item_id = (data.get('item_id') or '').strip()
        date_str = (data.get('date') or '').strip()
        if not item_id or not date_str:
            return jsonify({'success': False, 'error': 'item_id and date are required.'}), 400
        from datetime import date as date_type
        try:
            override_date = date_type.fromisoformat(date_str)
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid date (use YYYY-MM-DD).'}), 400
        try:
            ok = insurance_compliance.save_override(get_db_fn, item_id, override_date, user_key)
            if not ok:
                return jsonify({'success': False, 'error': 'Sub not found — run a compliance check first.'}), 404
            return jsonify({'success': True})
        except Exception as e:
            print(f'Office Ops compliance override error ({user_key}): {e}')
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/office-ops/compliance/vision-pass', methods=['POST'])
    @require_login
    def office_ops_compliance_vision_pass():
        """On-demand: try Claude vision on every 'needs manual entry' sub
        whose COI is a photo. Not automatic — Thomas/Stephanie trigger this
        from the page. Same Stephanie+Thomas gate as the rest of Office Ops."""
        blocked = _gate()
        if blocked:
            return jsonify({'success': False, 'error': 'Not allowed.'}), 403
        user_key = session.get('user_key')
        data = request.get_json(silent=True) or {}
        try:
            limit = int(data.get('limit') or VISION_PASS_BATCH_SIZE)
        except (TypeError, ValueError):
            limit = VISION_PASS_BATCH_SIZE
        limit = max(1, min(limit, VISION_PASS_BATCH_SIZE))
        try:
            result = insurance_compliance.run_vision_pass(get_db_fn, claude_api_key, claude_model, limit=limit)
            if 'error' in result:
                return jsonify({'success': False, 'error': result['error']}), 500
            from hub_usage import record_usage
            record_usage(
                get_db_fn, user_key, 'compliance', 'vision',
                f"{result.get('dated', 0)} dated of {result.get('attempted', 0)}",
            )
            return jsonify({'success': True, **result})
        except Exception as e:
            print(f'Office Ops vision pass error ({user_key}): {e}')
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/office-ops/compliance/coi/<item_id>')
    @require_login
    def office_ops_compliance_coi(item_id):
        """Proxy the Monday COI so Stephanie/Thomas can view a photo or PDF
        in the Hub. Monday's pre-signed URL expires; we re-resolve on each
        click. Stephanie + Thomas only (same gate as the rest of Office Ops)."""
        blocked = _gate()
        if blocked:
            return blocked
        try:
            data, filename, content_type, err = insurance_compliance.load_coi_asset(
                get_db_fn, item_id,
            )
        except Exception as e:
            print(f'Office Ops COI view error ({item_id}): {e}')
            return (f'Could not load COI: {e}', 502)
        if err:
            return (err, 404)
        from hub_usage import record_usage
        record_usage(
            get_db_fn, session.get('user_key'), 'compliance', 'view',
            filename or item_id,
        )
        return send_file(
            BytesIO(data),
            mimetype=content_type or 'application/octet-stream',
            as_attachment=False,
            download_name=filename or 'coi',
            max_age=60,
        )

    @app.route('/api/office-ops/upload', methods=['POST'])
    @require_login
    def office_ops_upload():
        user_key = session.get('user_key')
        if not can_access_office_ops(users, user_key):
            return jsonify({'success': False, 'error': 'Not allowed.'}), 403
        kind = (request.form.get('kind') or KIND_SUMMARY).strip()
        # Map UI kinds
        if kind in ('summary', 'ar_summary'):
            kind = KIND_SUMMARY
        elif kind in ('detail', 'ar_detail'):
            kind = KIND_DETAIL
        elif kind in ('invoice_list', 'invoices', 'sales_rep'):
            kind = KIND_INVOICE_LIST
        elif kind in ('monthly_outlook', 'outlook', 'monday'):
            kind = KIND_OUTLOOK
        elif kind in ('profit_loss', 'pl', 'pnl', 'p&l'):
            kind = KIND_PL
        f = request.files.get('file')
        if not f or not f.filename:
            return jsonify({'success': False, 'error': 'Choose a file to upload.'}), 400
        raw = f.read()
        fname = (f.filename or '').lower().replace('+', ' ')
        # Content wins over wrong drop-zone label (AR types). Do not steal an
        # Invoice List drop into Aging Detail — QB's 2026 Invoice List without
        # Sales Rep looks like detail headers.
        if kind not in (KIND_OUTLOOK, KIND_PL) and 'outlook' not in fname and 'profit' not in fname and 'p&l' not in fname:
            detected = detect_ar_report_type(f.filename, raw)
            if detected == 'invoice_list':
                kind = KIND_INVOICE_LIST
            elif detected == 'detail' and kind != KIND_INVOICE_LIST:
                kind = KIND_DETAIL
            elif detected == 'summary' and kind != KIND_INVOICE_LIST:
                kind = KIND_SUMMARY
        if 'outlook' in fname or 'monthly' in fname:
            kind = KIND_OUTLOOK
        if 'profit' in fname or 'p&l' in fname or 'pnl' in fname or 'loss' in fname:
            # Prefer P&L over generic "loss" in other names
            if 'aging' not in fname and 'invoice' not in fname:
                kind = KIND_PL
        saved = save_upload(
            get_db_fn, kind, f.filename, f.mimetype or '', raw, user_key,
        )
        if not saved.get('success'):
            return jsonify(saved), 400
        from hub_usage import record_usage
        record_usage(get_db_fn, user_key, 'office_ops', 'upload', f.filename or kind, kind)

        # P&L store only — used on next Generate
        if kind == KIND_PL:
            try:
                pl = parse_pl_bytes(f.filename, raw)
            except ValueError as e:
                return jsonify({'success': False, 'error': str(e), 'file_id': saved['file_id']}), 400
            return jsonify({
                'success': True,
                'detected': 'profit_loss',
                'file_id': saved['file_id'],
                'pl_summary': {
                    'period_label': pl.get('period_label'),
                    'income_ty': pl.get('income_ty'),
                    'net_income_ty': pl.get('net_income_ty'),
                    'gross_profit_ty': pl.get('gross_profit_ty'),
                },
            })

        # Monday report path (legacy outlook upload)
        if kind == KIND_OUTLOOK:
            result = process_monthly_outlook(get_db_fn, saved['file_id'], user_key)
            if not result.get('success'):
                return jsonify({
                    'success': False,
                    'error': result.get('error', 'Monday report failed.'),
                    'file_id': saved['file_id'],
                }), 400
            return jsonify({
                'success': True,
                'detected': 'monthly_outlook',
                'file_id': saved['file_id'],
                'monday_file_id': result['monday_file_id'],
                'download_url': result['download_url'],
                'insights_md': result.get('insights_md'),
                'ar_included': result.get('ar_included'),
            })

        expect = {
            KIND_DETAIL: 'detail',
            KIND_INVOICE_LIST: 'invoice_list',
        }.get(kind, 'summary')
        result = process_ar_file(get_db_fn, saved['file_id'], user_key, expect=expect)
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', 'Processed save failed.'),
                'file_id': saved['file_id'],
            }), 400
        return jsonify({
            'success': True,
            'file_id': saved['file_id'],
            'pack_id': result['pack_id'],
            'detected': result.get('detected') or expect,
            'summary': result['summary'],
            'numbers_draft_md': result['summary'].get('numbers_draft_md'),
        })

    @app.route('/api/office-ops/files/<int:file_id>/download')
    @require_login
    def office_ops_file_download(file_id):
        from flask import send_file
        user_key = session.get('user_key')
        if not can_access_office_ops(users, user_key):
            return jsonify({'error': 'Not allowed.'}), 403
        meta = get_file_bytes(get_db_fn, file_id)
        if not meta:
            return jsonify({'error': 'File not found.'}), 404
        return send_file(
            io.BytesIO(meta['data']),
            mimetype=meta['mime_type'],
            as_attachment=True,
            download_name=meta['filename'] or 'office_ops.xlsx',
        )

    @app.route('/api/office-ops/notes', methods=['GET', 'POST'])
    @require_login
    def office_ops_notes():
        user_key = session.get('user_key')
        if not can_access_office_ops(users, user_key):
            return jsonify({'error': 'Not allowed.'}), 403
        if request.method == 'GET':
            return jsonify({'success': True, 'notes': get_ar_notes(get_db_fn)})
        data = request.get_json(silent=True) or {}
        notes_map = data.get('notes') or {}
        result = save_ar_notes(get_db_fn, notes_map, user_key)
        if not result.get('success'):
            return jsonify(result), 400
        return jsonify({'success': True})

    @app.route('/api/office-ops/generate', methods=['POST'])
    @require_login
    def office_ops_generate():
        """Thursday pack: Invoice List + AR Summary + past-due notes → Monday Excel."""
        user_key = session.get('user_key')
        if not can_access_office_ops(users, user_key):
            return jsonify({'error': 'Not allowed.'}), 403
        # Optional notes save in same request (from modal)
        data = request.get_json(silent=True) or {}
        if data.get('notes'):
            save_ar_notes(get_db_fn, data['notes'], user_key)
        result = generate_thursday_pack(get_db_fn, user_key)
        if not result.get('success'):
            return jsonify(result), 400
        from hub_usage import record_usage
        record_usage(get_db_fn, user_key, 'office_ops', 'generate', 'Thursday pack')
        return jsonify(result)

    @app.route('/api/office-ops/latest')
    @require_login
    def office_ops_latest():
        user_key = session.get('user_key')
        if not can_access_office_ops(users, user_key):
            return jsonify({'error': 'Not allowed.'}), 403
        pack = get_latest_pack(get_db_fn)
        if not pack:
            return jsonify({'success': True, 'pack': None})
        return jsonify({'success': True, 'pack': pack})
