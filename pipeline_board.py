"""Pipeline Board — live shared Consultant/PM proposal-tracking sheet.

Replaces the ad hoc shared Google Sheets each Consultant/PM pair has been keeping
(e.g. "who's waiting on sub pricing", "did I log every proposal I sent"). One
board per pair, kept current for everyone viewing it via short client polling.

Deliberately NOT Socket.IO/WebSockets (was, briefly, on 2026-07-29): that
required the whole Hub to run on a single gevent worker, and any one
non-cooperative blocking call anywhere in the app (Claude API, docx/xlsx
generation) then stalled *every* request, not just this module -- a global
Hub slowdown traced back to this module. Plain polling against the existing
REST endpoints needs no special worker class, no monkey-patching, and cannot
regress the rest of the Hub's performance again by construction: this module
now has zero effect on how gunicorn runs.

One board per consultant, keyed by the consultant's user_key (`pair_key`) —
the consultant owns the client relationship; everyone writes the same rows.
Opened past the two-pair pilot 2026-08-13: Adam / Andy / Tony / Rachel.

**Access, as of 2026-08-21: every board is open to everyone on the roster.**
The old per-board `BOARD_ACCESS` roster and the `BOARD_ACCESS_ALL` oversight
set are both deleted — see the note above their former location. What remains
is `PRIMARY_PM_FOR_CONSULTANT`, which is assignment (where you land, whose
name is in the header), not permission. If a board ever needs restricting
again, that belongs in `tiers.py`, not in a new dict in this file.

Status values were chosen by reviewing two real exported sheets (Rachel Farler's
and another consultant's) rather than guessed — see chat history 2026-07-29.
Do NOT add an "IC" status: in the sample data every "IC" row was actually an
"Indian Creek" property tag left in the wrong column, not a real status.
"""

import json
import re
from datetime import datetime, date

import openpyxl
from psycopg2.extras import RealDictCursor

from tiers import is_leadership

# Consultant keys that have a live board. Order is the default-home
# tiebreak for people who can open more than one (James, admin pickers).
BOARD_CONSULTANTS = ('andy_potts', 'adam_cupito', 'tony_cumella', 'rachel_farler')
BOARD_CONSULTANT_SET = frozenset(BOARD_CONSULTANTS)

# Header label: the 1:1 working PM for that consultant's board.
# Deliberately NOT derived from USERS[...]['proposal_access'] -- that field
# is a broader "who can touch this consultant's proposals" grant (James
# floats Andy and Adam; Jordan's Hub grant is still `'all'`), not the same
# thing as "who is this consultant's actual PM pair." Scanning
# proposal_access once picked James for Andy by dict order — confirmed
# wrong 2026-07-30. Same rule still holds after the 2026-08-13 expansion.
PRIMARY_PM_FOR_CONSULTANT = {
    'andy_potts': 'ben_ramsey',
    'adam_cupito': 'jordan_allen',
    'tony_cumella': 'nick_triplett',
    'rachel_farler': 'derek_kidney',
}

# BOARD_ACCESS and BOARD_ACCESS_ALL were deleted 2026-08-21: every board is now
# open to everyone on the roster (Thomas — "everyone can really have the same
# access"). Those two lists were the clearest case for the tier rework: they
# existed only because `proposal_access` could not be trusted to answer "who is
# on this board", so a second roster got built beside it, and then a third set
# on top for oversight. All three are gone.
#
# PRIMARY_PM_FOR_CONSULTANT above survives, but purely as ASSIGNMENT — which
# board opens when you hit /pipeline-board with no ?pair=, and whose name shows
# in the header. It grants nothing. Keeping that distinction visible is the
# point: if a future change needs to restrict a board, it belongs in tiers.py,
# not in a new dict here.

# Back-compat aliases — a couple of comments / older tests used the
# pilot names. Prefer BOARD_* above.
PILOT_PAIR_CONSULTANTS = BOARD_CONSULTANT_SET
PILOT_PM_FOR_CONSULTANT = PRIMARY_PM_FOR_CONSULTANT
PRESENCE_TTL_SECONDS = 12  # no explicit "disconnect" under polling; just expire

# Ordered roughly by pipeline stage so the dropdown reads top-to-bottom as
# the work progresses. Early field stages (new / needs_scope / scoped) let
# Andy & Ben glance at what still needs a walk vs. what's ready to write —
# feedback 2026-08 from Andy Potts. "New" is the cleaner label for what he
# called "Untouched". Terminal-for-our-queue statuses (sent / awarded /
# cancelled) are the ones the board does *not* highlight as in-progress.
STATUSES = [
    {'value': 'new', 'label': 'New'},
    {'value': 'needs_scope', 'label': 'Needs walk/scope'},
    {'value': 'scoped', 'label': 'Walked/scoped'},
    {'value': 'draft', 'label': 'Draft'},
    {'value': 'sent', 'label': 'Sent'},
    {'value': 'awarded', 'label': 'Awarded'},
    {'value': 'on_hold', 'label': 'On Hold'},
    {'value': 'declined', 'label': 'Declined'},
    {'value': 'cancelled', 'label': 'Cancelled'},
]
STATUS_VALUES = frozenset(s['value'] for s in STATUSES)
STATUS_LABELS = {s['value']: s['label'] for s in STATUSES}
# Rows not in this set get a highlight — still in some form of progress.
COMPLETED_STATUSES = frozenset({'sent', 'awarded', 'cancelled'})

MAX_TEXT = 500
MAX_NOTES = 10000

_EDITABLE_TEXT_FIELDS = ('proposal_number', 'property_name', 'address', 'project',
                         'trade_partner', 'client_contact', 'notes')
_EDITABLE_NUMERIC_FIELDS = ('amount', 'sub_pay')

# Presence lives in Postgres (pipeline_board_presence table), refreshed by a
# heartbeat call every poll cycle rather than a socket connection. It used to
# be a module-level dict, but that's per-gunicorn-worker -- with 2 workers, a
# user's heartbeat only reaches whichever worker handled that request, so the
# *other* worker's requests never saw them as present. Fixed 2026-08-04 by
# moving it to the DB, which every worker reads/writes the same rows from.
# Deliberately still no Redis and no worker-count change, per the hard-won
# single-worker-outage lesson in CLAUDE.md -- this is just a small table.
# Entries older than PRESENCE_TTL_SECONDS are treated as gone -- the only way
# to detect someone closed their tab without an explicit disconnect event.


def _touch_presence(get_db_fn, pair_key, user_key, display, field):
    conn = get_db_fn()
    if not conn:
        return
    try:
        cur = conn.cursor()
        if field:
            cur.execute('''
                INSERT INTO pipeline_board_presence (pair_key, user_key, display, field, ts)
                VALUES (%s, %s, %s, %s, NOW())
                ON CONFLICT (pair_key, user_key)
                DO UPDATE SET display = EXCLUDED.display, field = EXCLUDED.field, ts = NOW()
            ''', (pair_key, user_key, display, field))
        else:
            cur.execute(
                'DELETE FROM pipeline_board_presence WHERE pair_key = %s AND user_key = %s',
                (pair_key, user_key),
            )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f'Pipeline board presence touch error: {e}')


def _live_presence(get_db_fn, pair_key, exclude_user_key=None):
    conn = get_db_fn()
    if not conn:
        return []
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            "DELETE FROM pipeline_board_presence WHERE ts < NOW() - (%s || ' seconds')::interval",
            (PRESENCE_TTL_SECONDS,),
        )
        cur.execute(
            'SELECT user_key, display, field FROM pipeline_board_presence '
            'WHERE pair_key = %s AND user_key != %s',
            (pair_key, exclude_user_key or ''),
        )
        rows = [dict(r) for r in cur.fetchall()]
        conn.commit()
        cur.close()
        conn.close()
        return rows
    except Exception as e:
        print(f'Pipeline board presence read error: {e}')
        return []


def init_tables(cur):
    cur.execute('''
        CREATE TABLE IF NOT EXISTS pipeline_board_entries (
            id SERIAL PRIMARY KEY,
            pair_key VARCHAR(100) NOT NULL,
            proposal_number VARCHAR(100),
            property_name VARCHAR(255),
            address VARCHAR(255),
            project TEXT,
            status VARCHAR(30) NOT NULL DEFAULT 'new',
            amount NUMERIC(12, 2),
            sub_pay NUMERIC(12, 2),
            trade_partner VARCHAR(255),
            client_contact VARCHAR(255),
            notes TEXT,
            row_order INTEGER NOT NULL DEFAULT 0,
            archived BOOLEAN NOT NULL DEFAULT FALSE,
            created_by VARCHAR(100),
            created_at TIMESTAMP DEFAULT NOW(),
            updated_by VARCHAR(100),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    cur.execute(
        "CREATE INDEX IF NOT EXISTS idx_pipeline_board_pair ON pipeline_board_entries(pair_key, archived)"
    )
    # CREATE TABLE ... DEFAULT 'new' above is inert for the already-existing
    # production table (Postgres no-ops the whole CREATE TABLE when the table
    # exists) -- the real on-disk default was still 'draft' from before the
    # 2026-08 status changes. Not a functional bug (create_entry hardcodes
    # 'new' explicitly at every INSERT, see below), but misleading if anyone
    # inspects the schema directly. Idempotent, like the CREATE statements
    # above: repeating ALTER COLUMN SET DEFAULT is a no-op each time and
    # touches zero existing rows (metadata only, no table rewrite).
    try:
        cur.execute(
            "ALTER TABLE pipeline_board_entries ALTER COLUMN status SET DEFAULT 'new'"
        )
    except Exception:
        pass
    cur.execute('''
        CREATE TABLE IF NOT EXISTS pipeline_board_presence (
            pair_key VARCHAR(100) NOT NULL,
            user_key VARCHAR(100) NOT NULL,
            display VARCHAR(255),
            field VARCHAR(100),
            ts TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (pair_key, user_key)
        )
    ''')
    # Full-board JSON snapshots so a bad client poll, bad import, or accidental
    # mass archive can be walked back. Same Postgres as live data (if the DB
    # itself is suspended, these don't help until it comes back -- the browser
    # localStorage cache covers that case). Kept small: last N per pair.
    cur.execute('''
        CREATE TABLE IF NOT EXISTS pipeline_board_backups (
            id SERIAL PRIMARY KEY,
            pair_key VARCHAR(100) NOT NULL,
            reason VARCHAR(50) NOT NULL,
            created_by VARCHAR(100),
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            entry_count INTEGER NOT NULL DEFAULT 0,
            payload JSONB NOT NULL
        )
    ''')
    cur.execute(
        'CREATE INDEX IF NOT EXISTS idx_pipeline_board_backups_pair '
        'ON pipeline_board_backups(pair_key, created_at DESC)'
    )


# Segment prefixes import_workbook used to fold into Notes (see below) --
# decided 2026-07-30 this just clutters the board and the field should
# start blank for people to write real notes into. import_workbook no
# longer generates these; this only cleans up rows imported before that.
_JUNK_NOTE_PREFIXES = ('Imported status:', 'Sent:')


def _strip_junk_notes(notes):
    if not notes:
        return None
    parts = [p.strip() for p in notes.split(' | ')]
    kept = [p for p in parts if not any(p.startswith(prefix) for prefix in _JUNK_NOTE_PREFIXES)]
    return ' | '.join(kept).strip() or None


def cleanup_legacy_import_notes(cur):
    """One-time cleanup of the auto-generated 'Imported status: ...' /
    'Sent: ...' text that earlier imports folded into Notes. Safe to run on
    every startup: it only rewrites rows whose stripped text actually
    differs, and since import_workbook stopped generating this text, there's
    nothing left for it to touch after the first successful pass -- it
    isn't a standing filter on notes people type in later."""
    cur.execute(
        "SELECT id, notes FROM pipeline_board_entries "
        "WHERE notes LIKE %s OR notes LIKE %s",
        ('%Imported status:%', '%Sent: 20%'),
    )
    for row in cur.fetchall():
        entry_id = row['id'] if isinstance(row, dict) else row[0]
        notes = row['notes'] if isinstance(row, dict) else row[1]
        cleaned = _strip_junk_notes(notes)
        if cleaned != notes:
            cur.execute(
                'UPDATE pipeline_board_entries SET notes = %s WHERE id = %s',
                (cleaned, entry_id),
            )


def can_import_to_board(users, user_key, pair_key):
    """Bulk Excel import — leadership and up, not everyone with board access.

    Importing was how each board was seeded from the Google Sheet it replaced
    (Thomas, 2026-08-23: "a one time feature that was needed to kick off the
    pipeline board... they shouldn't need to do it again"). Andy's and Rachel's
    boards are already migrated, so for everyone already using the Hub daily
    this control has done its job and only carries risk.

    It is narrowed rather than deleted because the same need returns whenever a
    board is created — a new consultant joins and arrives with a spreadsheet.
    That is a management action, so it sits with leadership.

    Reading and editing a board stay open to the whole roster: that is the
    collaboration the board exists for, it moves one row at a time, and every
    edit snapshots first. A bulk import rewrites a board wholesale, which is a
    different act with a much larger blast radius.
    """
    if not can_access_board(users, user_key, pair_key):
        return False
    return is_leadership(users, user_key)


def can_access_board(users, user_key, pair_key):
    """Any person on the roster may open any live board (2026-08-21).

    Still validates `pair_key` against BOARD_CONSULTANT_SET and `user_key`
    against the roster — "open to everyone" means everyone who works here, and
    an unknown or stale key must still fail closed. That check is what makes
    removing someone from USERS actually revoke their board access.
    """
    if pair_key not in BOARD_CONSULTANT_SET:
        return False
    return bool(user_key) and user_key in (users or {})


def get_pair_key(users, user_key):
    """Default board when opening /pipeline-board with no ?pair=.

    Pure assignment, not permission — everyone can open every board, this
    only decides where they land first. Consultants get their own board;
    primary PMs get their 1:1 consultant; everyone else gets the first board.
    Thomas has no default — he picks via ?pair=.
    """
    user = users.get(user_key, {})
    if user.get('tier') == 'owner':
        return None
    if user.get('role') == 'consultant' and user_key in BOARD_CONSULTANT_SET:
        return user_key
    for consultant_key, pm_key in PRIMARY_PM_FOR_CONSULTANT.items():
        if pm_key == user_key:
            return consultant_key
    for ck in BOARD_CONSULTANTS:
        if can_access_board(users, user_key, ck):
            return ck
    return None


def list_accessible_boards(users, user_key):
    """Ordered list of boards this user can open, for the switcher / dashboard."""
    boards = []
    for ck in BOARD_CONSULTANTS:
        if can_access_board(users, user_key, ck):
            pm_key = PRIMARY_PM_FOR_CONSULTANT.get(ck)
            boards.append({
                'key': ck,
                'consultant_display': _display(users, ck),
                'pm_display': _display(users, pm_key) if pm_key else 'PM',
            })
    return boards


def _display(users, user_key):
    return users.get(user_key, {}).get('display', user_key)


def _norm_property_name(name):
    return ' '.join((name or '').lower().split())


def last_used_client_contact(entries, property_name, exclude_id=None):
    """Client / property-manager name last used on this board for the same property.

    Rachel (2026-08): she expects Client Contact to fill itself the way the
    old sheet did — the *client's* manager, not the PPS PM. Prefer the
    most recently updated row with the same property name (case/spacing
    ignored). Empty contacts and the row being edited are skipped.
    """
    key = _norm_property_name(property_name)
    if not key or not entries:
        return ''
    best = ''
    best_ts = ''
    for entry in entries:
        if exclude_id is not None and str(entry.get('id')) == str(exclude_id):
            continue
        if _norm_property_name(entry.get('property_name')) != key:
            continue
        contact = (entry.get('client_contact') or '').strip()
        if not contact:
            continue
        ts = str(entry.get('updated_at') or '')
        if not best or ts >= best_ts:
            best = contact
            best_ts = ts
    return best


def _row_to_dict(row):
    d = dict(row)
    for k in ('amount', 'sub_pay'):
        if d.get(k) is not None:
            d[k] = float(d[k])
    for k in ('created_at', 'updated_at'):
        if d.get(k):
            d[k] = d[k].isoformat()
    return d


def list_entries(get_db_fn, pair_key):
    """Return (rows, error). On DB failure rows is None and error is a short
    message -- callers must NOT treat that as an empty board (the UI used to
    wipe all rows when get_db failed, because this returned [])."""
    conn = get_db_fn()
    if not conn:
        return None, 'Database unavailable'
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute('''
            SELECT * FROM pipeline_board_entries
            WHERE pair_key = %s AND archived = FALSE
            ORDER BY row_order ASC, id ASC
        ''', (pair_key,))
        rows = [_row_to_dict(r) for r in cur.fetchall()]
        cur.close()
        conn.close()
        return rows, None
    except Exception as e:
        print(f'Pipeline board list error: {e}')
        return None, 'Could not load board'


def _prefix_from_pair_key(pair_key):
    """'andy_potts' -> 'AP', 'rachel_farler' -> 'RF' -- matches the real
    consultant-initials + 2-digit-year + 3-digit-sequence scheme already
    used in both source spreadsheets (AP26001, RF26102) and elsewhere in
    the Hub (TE26xxx for Thomas). Derived from the key, not a lookup table,
    so it keeps working for future pairs with no extra config."""
    parts = [p for p in pair_key.split('_') if p]
    prefix = ''.join(p[0] for p in parts[:2]).upper()
    return prefix or 'PB'


def _next_proposal_number(cur, pair_key):
    """Next number in THIS pair's own sequence (independent of the Hub's
    Proposal Generator numbering, per owner direction 2026-07-29) --
    <prefix><year>NNN, continuing from whatever's already on the board
    (including imported rows) rather than restarting at 001."""
    prefix = _prefix_from_pair_key(pair_key)
    year = f'{datetime.now().year % 100:02d}'
    stem = f'{prefix}{year}'
    cur.execute(
        'SELECT proposal_number FROM pipeline_board_entries '
        'WHERE pair_key = %s AND proposal_number IS NOT NULL',
        (pair_key,),
    )
    pattern = re.compile(rf'^{re.escape(stem)}(\d+)$', re.IGNORECASE)
    max_seq = 0
    for row in cur.fetchall():
        po = (row['proposal_number'] if isinstance(row, dict) else row[0]) or ''
        m = pattern.match(po.strip())
        if m:
            max_seq = max(max_seq, int(m.group(1)))
    return f'{stem}{max_seq + 1:03d}'


def create_entry(get_db_fn, pair_key, user_key):
    conn = get_db_fn()
    if not conn:
        return None
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        proposal_number = _next_proposal_number(cur, pair_key)
        cur.execute('''
            INSERT INTO pipeline_board_entries
                (pair_key, proposal_number, status, row_order, created_by, updated_by)
            VALUES (%s, %s, 'new',
                COALESCE((SELECT MAX(row_order) + 1 FROM pipeline_board_entries WHERE pair_key = %s), 0),
                %s, %s)
            RETURNING *
        ''', (pair_key, proposal_number, pair_key, user_key, user_key))
        row = _row_to_dict(cur.fetchone())
        conn.commit()
        cur.close()
        conn.close()
        return row
    except Exception as e:
        print(f'Pipeline board create error: {e}')
        return None


def _clean_text(value, max_len):
    if value is None:
        return None
    value = str(value).strip()[:max_len]
    return value or None


def _clean_numeric(value):
    if value in (None, ''):
        return None
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def update_entry(get_db_fn, entry_id, pair_key, user_key, fields):
    """Update whitelisted fields on one entry. Returns the updated row or None."""
    sets, params = [], []

    if 'status' in fields:
        status = fields['status']
        if status not in STATUS_VALUES:
            return None, 'Invalid status'
        sets.append('status = %s')
        params.append(status)

    for f in _EDITABLE_TEXT_FIELDS:
        if f in fields:
            sets.append(f'{f} = %s')
            params.append(_clean_text(fields[f], MAX_NOTES if f == 'notes' else MAX_TEXT))

    for f in _EDITABLE_NUMERIC_FIELDS:
        if f in fields:
            sets.append(f'{f} = %s')
            params.append(_clean_numeric(fields[f]))

    if not sets:
        return None, 'No editable fields provided'

    sets.append('updated_by = %s')
    sets.append('updated_at = NOW()')
    params.append(user_key)
    params.extend([entry_id, pair_key])

    conn = get_db_fn()
    if not conn:
        return None, 'Database unavailable'
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(f'''
            UPDATE pipeline_board_entries SET {", ".join(sets)}
            WHERE id = %s AND pair_key = %s AND archived = FALSE
            RETURNING *
        ''', params)
        row = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        if not row:
            return None, 'Entry not found'
        return _row_to_dict(row), None
    except Exception as e:
        print(f'Pipeline board update error: {e}')
        return None, 'Could not save'


def _normalize_header(h):
    if h is None:
        return ''
    text = str(h).strip().lower()
    text = text.replace('$', ' ').replace('?', ' ').replace('#', ' ')
    text = re.sub(r'[^a-z0-9 ]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


# Built directly from the two real sheets reviewed 2026-07-29 (Rachel Farler's
# and another consultant's "2026 PO / Proposals" + its "Bath Tubs Simms" tab),
# not guessed. 'notes:Label' means "keep it, just fold into Notes with this
# label" -- used for real columns (Material Cost, Splits, batch-job dates)
# that don't have their own schema field. Any header NOT in this map at all
# still isn't dropped: see _map_sheet's unmapped-column fallback below.
_HEADER_ALIASES = {
    'po': 'proposal_number',
    'po number': 'proposal_number',
    'proposal number': 'proposal_number',
    'proposal': 'proposal_number',
    'property': 'property_name',
    'address': 'address',
    'project': 'project',
    'job status': 'status',
    'status': 'status',
    'sent': 'sent_or_status',  # value-type sniffed: datetime -> note, text -> status
    'amount': 'amount',
    'cost': 'sub_pay',
    'sub pay': 'sub_pay',
    'material cost': 'notes:Material cost',
    'sub': 'trade_partner',
    'trade partner': 'trade_partner',
    'manager': 'client_contact',  # the *client's* property manager, confirmed from real data
    'company': 'client_contact',
    'client contact': 'client_contact',
    'splits': 'notes:Splits',
    'how many': 'notes:Qty',
    'date start': 'notes:Date start',
    'date end': 'notes:Date end',
    'started': 'notes:Started',
    'done': 'notes:Done',
    'paid out': 'notes:Paid out',
}

# Every real status value seen across both sheets, grouped. Deliberately no
# "IC" entry -- in the sample data every "IC" row was an "Indian Creek"
# property tag left in the wrong column, not a status (verified by checking
# every such row was for that one property). Anything not listed here still
# isn't lost: _normalize_status_value always preserves the raw text in Notes.
_STATUS_KEYWORDS = {
    'sent': {'yes', 'sent', 'proposed', 'propsed', 'verbal so far', 'online forms', 'email', 'work order'},
    'awarded': {'complete', 'completed', 'awarded', 'scheduled'},
    'on_hold': {'on hold', 'delayed'},
    'cancelled': {'cancelled', 'canceled', 'not doing', 'punt', 'duplicate'},
    'draft': {'no', 'none', 'na', 'n/a', '', '.....'},
}


def _normalize_status_value(raw):
    """Best-guess enum mapping, but ALWAYS returns the original text too --
    never silently collapse a real note ('waiting to hear from Desirae')
    into a clean-looking status without keeping what was actually written."""
    text = ('' if raw is None else str(raw)).strip()
    status = 'draft'
    for candidate, keywords in _STATUS_KEYWORDS.items():
        if text.lower() in keywords:
            status = candidate
            break
    return status, text


def _map_sheet(ws):
    """Map one worksheet's header row (row 1) to schema fields.
    Returns {col_index: target}; target is a schema field name, 'status',
    'sent_or_status', or 'notes:<label>'."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    col_map = {}
    unmapped = []
    for i, raw_header in enumerate(header_row):
        norm = _normalize_header(raw_header)
        if not norm:
            continue
        if norm in _HEADER_ALIASES:
            col_map[i] = _HEADER_ALIASES[norm]
        elif i == 0:
            # An unrecognized first column is a job title/description in both
            # real sample sheets (e.g. "9456 Bainwoods - Soffitt / Roof
            # Repair") -- keep it visible as Project rather than Notes.
            col_map[i] = 'project'
        else:
            col_map[i] = f'notes:{raw_header}'
            unmapped.append(str(raw_header))
    return col_map, unmapped


MAX_IMPORT_BYTES = 2 * 1024 * 1024  # 2 MB -- generous for what these sheets actually are
MAX_IMPORT_ROWS = 2000


def import_workbook(get_db_fn, file_obj, pair_key, user_key):
    """Import every sheet of an uploaded .xlsx into one pair's board.
    Never silently drops a column — anything not recognized lands in Notes,
    labeled with its original header, instead of being discarded.

    Size/row caps below: parsing happens on the request path under a plain
    sync worker (see the module docstring and CLAUDE.md 'Hard-won lesson' --
    same risk *class*, much smaller blast radius than that incident since
    it's one request, not the whole process, but still worth a cheap guard
    rather than trusting every future upload to stay small like today's."""
    file_obj.seek(0, 2)
    size = file_obj.tell()
    file_obj.seek(0)
    if size > MAX_IMPORT_BYTES:
        return {'success': False, 'error': f'File is too large (max {MAX_IMPORT_BYTES // (1024*1024)}MB).'}

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        return {'success': False, 'error': f'Could not read that file as an Excel workbook: {e}'}

    if wb.worksheets and (wb.worksheets[0].max_row or 0) > MAX_IMPORT_ROWS:
        return {'success': False, 'error': f'Sheet has too many rows (max {MAX_IMPORT_ROWS}).'}

    conn = get_db_fn()
    if not conn:
        return {'success': False, 'error': 'Database unavailable'}

    imported = 0
    skipped = 0
    duplicates_skipped = 0
    warnings = []

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            'SELECT COALESCE(MAX(row_order), -1) + 1 AS next_order '
            'FROM pipeline_board_entries WHERE pair_key = %s',
            (pair_key,),
        )
        row_order = cur.fetchone()['next_order']

        # Dedup set for "skip if this PO already exists" -- checked against
        # every entry ever created for this pair (archived or not, so a
        # re-import can't resurrect a duplicate of something removed), and
        # updated as we go so two identical rows within the same file (or
        # sheet) are also caught, not just repeats across separate imports.
        cur.execute(
            'SELECT proposal_number FROM pipeline_board_entries '
            'WHERE pair_key = %s AND proposal_number IS NOT NULL',
            (pair_key,),
        )
        existing_pos = {
            r['proposal_number'].strip().upper()
            for r in cur.fetchall() if r['proposal_number']
        }

        # Only the first sheet is treated as real tracking data. Extra tabs in
        # these workbooks have turned out to be ad hoc scratch space (e.g. a
        # "doodle" tab someone used to track one side project with a totally
        # different, inconsistent shape — confirmed 2026-07-29) rather than
        # more of the same table, so importing them blindly does more harm
        # than good. Note it rather than silently ignore it.
        ws = wb.worksheets[0]
        if len(wb.worksheets) > 1:
            skipped_titles = [s.title for s in wb.worksheets[1:]]
            warnings.append(
                f'Only imported the first sheet ("{ws.title}"). '
                f'Additional tab(s) ignored: {", ".join(skipped_titles)}.'
            )

        col_map, unmapped = _map_sheet(ws)
        if unmapped:
            warnings.append(
                f'Columns not recognized, kept in Notes: {", ".join(unmapped)}'
            )
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v in (None, '') for v in row):
                continue
            fields = {}
            notes_parts = []
            has_content = False
            for i, value in enumerate(row):
                if value is None or value == '' or i not in col_map:
                    continue
                target = col_map[i]
                if target == 'sent_or_status' and isinstance(value, (datetime, date)):
                    # A sent-date value counts as real data even though (per
                    # owner direction 2026-07-30) it no longer gets written
                    # into Notes -- that bookkeeping cluttered the board and
                    # the field should start clean for people to write real
                    # notes into. The date itself is a genuine loss on
                    # import; if it turns out to matter, add a real
                    # sent_date column instead of reviving the notes hack.
                    has_content = True
                elif target in ('status', 'sent_or_status'):
                    status, raw_text = _normalize_status_value(value)
                    fields['status'] = status
                    has_content = True
                elif target.startswith('notes:'):
                    notes_parts.append(f'{target.split(":", 1)[1]}: {value}')
                    has_content = True
                elif target in _EDITABLE_NUMERIC_FIELDS:
                    cleaned = _clean_numeric(value)
                    if cleaned is not None:
                        fields[target] = cleaned
                        has_content = True
                else:
                    cleaned = _clean_text(value, MAX_TEXT)
                    if cleaned:
                        fields[target] = cleaned
                        has_content = True

            if not has_content:
                skipped += 1
                continue

            if notes_parts:
                fields['notes'] = _clean_text(' | '.join(notes_parts), MAX_NOTES)

            po_key = (fields.get('proposal_number') or '').strip().upper()
            if po_key and po_key in existing_pos:
                duplicates_skipped += 1
                continue
            if po_key:
                existing_pos.add(po_key)

            cur.execute('''
                INSERT INTO pipeline_board_entries
                    (pair_key, proposal_number, property_name, address, project, status,
                     amount, sub_pay, trade_partner, client_contact, notes, row_order,
                     created_by, updated_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (
                pair_key, fields.get('proposal_number'), fields.get('property_name'),
                fields.get('address'), fields.get('project'), fields.get('status', 'draft'),
                fields.get('amount'), fields.get('sub_pay'), fields.get('trade_partner'),
                fields.get('client_contact'), fields.get('notes'), row_order,
                user_key, user_key,
            ))
            row_order += 1
            imported += 1
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f'Pipeline board import error: {e}')
        return {'success': False, 'error': 'Could not import that file — no rows were added.'}

    return {
        'success': True, 'imported': imported, 'skipped': skipped,
        'duplicates_skipped': duplicates_skipped, 'warnings': warnings,
    }



# Keep enough snapshots that a bad morning of edits is recoverable, without
# unbounded growth. ~40 writes/day * 7 days would be ~280; 60 is a practical
# middle for a 2-pair pilot.
BACKUP_KEEP_PER_PAIR = 60


def _list_all_entries_for_backup(cur, pair_key):
    """Active + archived rows for one pair (full recoverability)."""
    cur.execute(
        'SELECT * FROM pipeline_board_entries '
        'WHERE pair_key = %s ORDER BY row_order ASC, id ASC',
        (pair_key,),
    )
    return [_row_to_dict(r) for r in cur.fetchall()]


def snapshot_board(get_db_fn, pair_key, reason, user_key=None):
    """Persist a full-board JSON snapshot. Best-effort: never fails the
    caller's write if snapshotting itself errors."""
    conn = get_db_fn()
    if not conn:
        return None
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        rows = _list_all_entries_for_backup(cur, pair_key)
        cur.execute(
            '''INSERT INTO pipeline_board_backups
                   (pair_key, reason, created_by, entry_count, payload)
               VALUES (%s, %s, %s, %s, %s::jsonb)
               RETURNING id''',
            (pair_key, (reason or 'manual')[:50], user_key,
             len(rows), json.dumps(rows)),
        )
        backup_id = cur.fetchone()['id']
        # Prune older than keep-N for this pair.
        cur.execute(
            '''DELETE FROM pipeline_board_backups
               WHERE pair_key = %s AND id NOT IN (
                   SELECT id FROM pipeline_board_backups
                   WHERE pair_key = %s
                   ORDER BY created_at DESC, id DESC
                   LIMIT %s
               )''',
            (pair_key, pair_key, BACKUP_KEEP_PER_PAIR),
        )
        conn.commit()
        cur.close()
        conn.close()
        return backup_id
    except Exception as e:
        print(f'Pipeline board snapshot error: {e}')
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        return None


def list_backups(get_db_fn, pair_key, limit=20):
    conn = get_db_fn()
    if not conn:
        return None, 'Database unavailable'
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            '''SELECT id, pair_key, reason, created_by, created_at, entry_count
               FROM pipeline_board_backups
               WHERE pair_key = %s
               ORDER BY created_at DESC, id DESC
               LIMIT %s''',
            (pair_key, limit),
        )
        rows = []
        for r in cur.fetchall():
            d = dict(r)
            if d.get('created_at'):
                d['created_at'] = d['created_at'].isoformat()
            rows.append(d)
        cur.close()
        conn.close()
        return rows, None
    except Exception as e:
        print(f'Pipeline board list_backups error: {e}')
        return None, 'Could not list backups'


def get_backup(get_db_fn, pair_key, backup_id):
    conn = get_db_fn()
    if not conn:
        return None, 'Database unavailable'
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            '''SELECT id, pair_key, reason, created_by, created_at, entry_count, payload
               FROM pipeline_board_backups
               WHERE id = %s AND pair_key = %s''',
            (backup_id, pair_key),
        )
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return None, 'Backup not found'
        d = dict(row)
        if d.get('created_at'):
            d['created_at'] = d['created_at'].isoformat()
        # payload may already be list (psycopg2 jsonb) or need no transform
        return d, None
    except Exception as e:
        print(f'Pipeline board get_backup error: {e}')
        return None, 'Could not load backup'


def restore_from_backup(get_db_fn, pair_key, backup_id, user_key):
    """Replace live board for pair_key with a snapshot.

    Strategy (safe for pilot size):
    1. Snapshot current board first (reason=pre_restore).
    2. Soft-archive every currently active row.
    3. For each payload row: if id still exists, un-archive + overwrite fields;
       else INSERT a new row (ids may not match if sequence moved on).
    Returns (summary_dict, error).
    """
    backup, err = get_backup(get_db_fn, pair_key, backup_id)
    if err:
        return None, err
    payload = backup.get('payload') or []
    if not isinstance(payload, list):
        return None, 'Backup payload is not a list'

    # Safety snapshot of current state before we touch anything.
    snapshot_board(get_db_fn, pair_key, 'pre_restore', user_key)

    conn = get_db_fn()
    if not conn:
        return None, 'Database unavailable'
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            '''UPDATE pipeline_board_entries
               SET archived = TRUE, updated_at = NOW(), updated_by = %s
               WHERE pair_key = %s AND archived = FALSE''',
            (user_key, pair_key),
        )
        restored = 0
        inserted = 0
        for raw in payload:
            if not isinstance(raw, dict):
                continue
            # Only restore non-archived rows from the snapshot into the active board.
            if raw.get('archived'):
                continue
            entry_id = raw.get('id')
            fields = {
                'proposal_number': raw.get('proposal_number'),
                'property_name': raw.get('property_name'),
                'address': raw.get('address'),
                'project': raw.get('project'),
                'status': raw.get('status') if raw.get('status') in STATUS_VALUES else 'draft',
                'amount': raw.get('amount'),
                'sub_pay': raw.get('sub_pay'),
                'trade_partner': raw.get('trade_partner'),
                'client_contact': raw.get('client_contact'),
                'notes': raw.get('notes'),
                'row_order': raw.get('row_order') or 0,
            }
            updated = False
            if entry_id is not None:
                cur.execute(
                    '''UPDATE pipeline_board_entries SET
                        proposal_number = %s, property_name = %s, address = %s,
                        project = %s, status = %s, amount = %s, sub_pay = %s,
                        trade_partner = %s, client_contact = %s, notes = %s,
                        row_order = %s, archived = FALSE,
                        updated_by = %s, updated_at = NOW()
                       WHERE id = %s AND pair_key = %s
                       RETURNING id''',
                    (fields['proposal_number'], fields['property_name'], fields['address'],
                     fields['project'], fields['status'], fields['amount'], fields['sub_pay'],
                     fields['trade_partner'], fields['client_contact'], fields['notes'],
                     fields['row_order'], user_key, entry_id, pair_key),
                )
                if cur.fetchone():
                    updated = True
                    restored += 1
            if not updated:
                cur.execute(
                    '''INSERT INTO pipeline_board_entries
                        (pair_key, proposal_number, property_name, address, project, status,
                         amount, sub_pay, trade_partner, client_contact, notes, row_order,
                         archived, created_by, updated_by)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,FALSE,%s,%s)''',
                    (pair_key, fields['proposal_number'], fields['property_name'],
                     fields['address'], fields['project'], fields['status'],
                     fields['amount'], fields['sub_pay'], fields['trade_partner'],
                     fields['client_contact'], fields['notes'], fields['row_order'],
                     user_key, user_key),
                )
                inserted += 1
                restored += 1
        conn.commit()
        cur.close()
        conn.close()
        snapshot_board(get_db_fn, pair_key, 'post_restore', user_key)
        return {'restored': restored, 'inserted_new': inserted, 'from_backup': backup_id}, None
    except Exception as e:
        print(f'Pipeline board restore error: {e}')
        try:
            conn.rollback()
            conn.close()
        except Exception:
            pass
        return None, 'Restore failed'


def archive_entry(get_db_fn, entry_id, pair_key):
    conn = get_db_fn()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute('''
            UPDATE pipeline_board_entries SET archived = TRUE, updated_at = NOW()
            WHERE id = %s AND pair_key = %s
        ''', (entry_id, pair_key))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        print(f'Pipeline board archive error: {e}')
        return False


def register_routes(app, get_db_fn, users, require_login):
    from flask import jsonify, render_template, request, session, redirect, url_for

    def _current_pair(fallback_query_override_for_admin=True):
        """Resolve the pair_key this request should operate on.

        Anyone who can open more than one board (Adam/Andy, Ben, James,
        Jordan, admin) switches via ?pair=. The override is honored only
        when can_access_board says yes — a stray query string cannot
        open a board the user is not on.
        """
        user_key = session['user_key']
        override = (request.args.get('pair')
                    or request.form.get('pair')
                    or (request.get_json(silent=True) or {}).get('pair'))
        if override and can_access_board(users, user_key, override):
            return user_key, override
        return user_key, get_pair_key(users, user_key)

    @app.route('/pipeline-board')
    @require_login
    def pipeline_board_page():
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return redirect(url_for('dashboard'))
        pm_key = PRIMARY_PM_FOR_CONSULTANT.get(pair_key)
        accessible = list_accessible_boards(users, user_key)
        from hub_usage import record_usage
        record_usage(
            get_db_fn, user_key, 'pipeline', 'open',
            _display(users, pair_key),
        )
        return render_template(
            'pipeline_board.html',
            pair_key=pair_key,
            consultant_display=_display(users, pair_key),
            pm_display=_display(users, pm_key) if pm_key else 'PM',
            statuses=STATUSES,
            completed_statuses=sorted(COMPLETED_STATUSES),
            user_key=user_key,
            user_display=_display(users, user_key),
            accessible_boards=accessible,
            is_admin_preview=(users.get(user_key, {}).get('role') == 'admin'
                              and get_pair_key(users, user_key) != pair_key),
            can_import=can_import_to_board(users, user_key, pair_key),
        )

    @app.route('/api/pipeline-board/entries', methods=['GET', 'POST'])
    @require_login
    def pipeline_board_entries_api():
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return jsonify({'error': 'Not authorized'}), 403
        if request.method == 'GET':
            entries, err = list_entries(get_db_fn, pair_key)
            if err:
                # Critical: never return entries=[] on DB failure — the client
                # used to treat that as "board is empty" and wipe the UI.
                return jsonify({
                    'error': err,
                    'entries': None,
                    'pair_key': pair_key,
                    'db_unavailable': True,
                }), 503
            return jsonify({
                'entries': entries,
                'presence': _live_presence(get_db_fn, pair_key, exclude_user_key=user_key),
                'pair_key': pair_key,
            })
        entry = create_entry(get_db_fn, pair_key, user_key)
        if not entry:
            return jsonify({'error': 'Could not create row'}), 500
        snapshot_board(get_db_fn, pair_key, 'create', user_key)
        return jsonify({'success': True, 'entry': entry})

    @app.route('/api/pipeline-board/entries/<int:entry_id>', methods=['POST'])
    @require_login
    def pipeline_board_update_api(entry_id):
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return jsonify({'error': 'Not authorized'}), 403
        data = request.get_json(silent=True) or {}
        fields = {k: v for k, v in data.items() if k in
                  _EDITABLE_TEXT_FIELDS + _EDITABLE_NUMERIC_FIELDS + ('status',)}
        entry, err = update_entry(get_db_fn, entry_id, pair_key, user_key, fields)
        if err:
            return jsonify({'error': err}), 400
        snapshot_board(get_db_fn, pair_key, 'update', user_key)
        return jsonify({'success': True, 'entry': entry})

    @app.route('/api/pipeline-board/import', methods=['POST'])
    @require_login
    def pipeline_board_import_api():
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return jsonify({'error': 'Not authorized'}), 403
        if not can_import_to_board(users, user_key, pair_key):
            return jsonify({
                'error': 'Importing a spreadsheet is limited to Thomas, Tony, '
                         'Trey and Stephanie. Boards are seeded once when they '
                         'are created — add rows directly instead.'
            }), 403
        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'error': 'No file selected'}), 400
        if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
            return jsonify({'error': 'Please upload an .xlsx (Excel) file'}), 400
        result = import_workbook(get_db_fn, file, pair_key, user_key)
        if not result.get('success'):
            return jsonify(result), 400
        snapshot_board(get_db_fn, pair_key, 'import', user_key)
        return jsonify(result)

    @app.route('/api/pipeline-board/entries/<int:entry_id>/archive', methods=['POST'])
    @require_login
    def pipeline_board_archive_api(entry_id):
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return jsonify({'error': 'Not authorized'}), 403
        ok = archive_entry(get_db_fn, entry_id, pair_key)
        if not ok:
            return jsonify({'error': 'Could not remove row'}), 500
        snapshot_board(get_db_fn, pair_key, 'archive', user_key)
        return jsonify({'success': True})

    @app.route('/api/pipeline-board/presence', methods=['POST'])
    @require_login
    def pipeline_board_presence_api():
        """Heartbeat: 'I'm here, and here's what I'm editing (or nothing)'.
        Called on focus/blur and on a timer while a cell stays focused, from
        the poll loop in pipeline_board.html — replaces the old Socket.IO
        join/cell_focus/disconnect events with plain polling (see module
        docstring for why)."""
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return jsonify({'error': 'Not authorized'}), 403
        data = request.get_json(silent=True) or {}
        field = (data.get('field') or '').strip()[:100] or None
        _touch_presence(get_db_fn, pair_key, user_key, _display(users, user_key), field)
        return jsonify({'success': True, 'presence': _live_presence(get_db_fn, pair_key, exclude_user_key=user_key)})

    @app.route('/api/pipeline-board/backups', methods=['GET', 'POST'])
    @require_login
    def pipeline_board_backups_api():
        """List snapshots (GET) or take a manual snapshot (POST). Admin or pair member."""
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return jsonify({'error': 'Not authorized'}), 403
        if request.method == 'POST':
            backup_id = snapshot_board(get_db_fn, pair_key, 'manual', user_key)
            if not backup_id:
                return jsonify({'error': 'Could not create backup'}), 500
            return jsonify({'success': True, 'backup_id': backup_id})
        backups, err = list_backups(get_db_fn, pair_key)
        if err:
            return jsonify({'error': err}), 503
        return jsonify({'success': True, 'backups': backups, 'pair_key': pair_key})

    @app.route('/api/pipeline-board/backups/<int:backup_id>', methods=['GET'])
    @require_login
    def pipeline_board_backup_get_api(backup_id):
        """Download one full snapshot (JSON). Pair member or admin."""
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return jsonify({'error': 'Not authorized'}), 403
        backup, err = get_backup(get_db_fn, pair_key, backup_id)
        if err:
            return jsonify({'error': err}), 404 if err == 'Backup not found' else 503
        return jsonify({'success': True, 'backup': backup})

    @app.route('/api/pipeline-board/backups/<int:backup_id>/restore', methods=['POST'])
    @require_login
    def pipeline_board_backup_restore_api(backup_id):
        """Admin-only restore: replaces active board from a snapshot."""
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return jsonify({'error': 'Not authorized'}), 403
        if users.get(user_key, {}).get('role') != 'admin':
            return jsonify({'error': 'Only admin can restore a board from backup'}), 403
        summary, err = restore_from_backup(get_db_fn, pair_key, backup_id, user_key)
        if err:
            return jsonify({'error': err}), 400
        return jsonify({'success': True, **summary})

    @app.route('/api/pipeline-board/export', methods=['GET'])
    @require_login
    def pipeline_board_export_api():
        """Live export of active rows (always available to pair members for
        offline/spreadsheet safety copy)."""
        user_key, pair_key = _current_pair()
        if not pair_key or not can_access_board(users, user_key, pair_key):
            return jsonify({'error': 'Not authorized'}), 403
        entries, err = list_entries(get_db_fn, pair_key)
        if err:
            return jsonify({'error': err, 'entries': None}), 503
        # Also take a named snapshot so "I exported" is recoverable server-side.
        snapshot_board(get_db_fn, pair_key, 'export', user_key)
        return jsonify({
            'success': True,
            'pair_key': pair_key,
            'exported_at': datetime.now().isoformat(),
            'entries': entries,
        })
